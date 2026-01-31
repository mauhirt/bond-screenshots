#!/usr/bin/env python3
"""
Test sequential alignment matching.

Strategy: Walk through CSV rows in order. For each CSV row, try to match it
to the next unmatched Excel row. If the issuer name prefixes match, pair them.
If not, this CSV row is a duplicate/extra — skip it.

This works because:
1. Both Excel and CSV are in the same Bloomberg terminal order
2. The CSV has extra duplicate rows at screenshot boundaries
3. The duplicates are "insertions" in the sequence
"""
import csv
import re
import openpyxl

# Load original Excel
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    amt = ws_orig.cell(row=row, column=5).value
    formula = ws_orig.cell(row=row, column=2).value or ''
    m = re.search(r'"([^"]+)\s+Muni"', str(formula))
    bb_id = m.group(1) if m else ''
    excel_rows.append({'issuer': issuer, 'amt': amt, 'bb_id': bb_id, 'idx': row - 2})

# Load CSV
csv_rows = []
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    for row_data in reader:
        if len(row_data) >= 10:
            cusip = row_data[0].strip()
            if cusip.startswith('TH '):
                cusip = cusip[3:]
            if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                if not all(f.strip() == '--' for f in row_data[:5]):
                    csv_rows.append({
                        'cusip': cusip,
                        'issuer': row_data[2].strip(),
                        'fields': row_data,
                        'csv_idx': len(csv_rows)
                    })

print(f"Excel rows: {len(excel_rows)}")
print(f"CSV rows: {len(csv_rows)}")


def issuer_match(excel_issuer, csv_issuer, min_prefix=8):
    """Check if issuer names match (prefix comparison with OCR tolerance)."""
    e = excel_issuer.lower().strip()
    c = csv_issuer.lower().strip()
    if not e or not c:
        return False

    # Direct prefix match
    prefix_len = 0
    for a, b in zip(e, c):
        if a == b:
            prefix_len += 1
        else:
            break

    if prefix_len >= min_prefix:
        return True

    # Try with OCR-tolerant comparison (first 12 chars)
    e12 = e[:12]
    c12 = c[:12]
    matches = sum(1 for a, b in zip(e12, c12) if a == b)
    return matches >= min(len(e12), len(c12)) * 0.7


# Sequential matching with lookahead
matched = {}  # excel_idx -> csv_idx
matched_csv = set()
excel_ptr = 0  # Next Excel row to match

for ci, c in enumerate(csv_rows):
    if excel_ptr >= len(excel_rows):
        break

    e = excel_rows[excel_ptr]

    if issuer_match(e['issuer'], c['issuer']):
        matched[excel_ptr] = ci
        matched_csv.add(ci)
        excel_ptr += 1
    else:
        # CSV row doesn't match current Excel row.
        # Check if it matches a nearby Excel row (in case CSV is missing a row)
        # But first, this is likely a duplicate - skip it
        # However, if the next few CSV rows also don't match, we might need to skip an Excel row
        pass

print(f"\nSimple sequential: {len(matched)} matched, {len(excel_rows) - len(matched)} unmatched Excel rows")
print(f"  Excel pointer stopped at: {excel_ptr}/{len(excel_rows)}")

# Show where alignment breaks
if excel_ptr < len(excel_rows):
    print(f"\nAlignment broke at Excel row {excel_ptr}:")
    print(f"  Excel: '{excel_rows[excel_ptr]['issuer'][:50]}'")
    # Show nearby CSV rows
    last_matched_csv = max(matched.values()) if matched else 0
    print(f"  Last matched CSV idx: {last_matched_csv}")
    for ci in range(last_matched_csv - 2, min(last_matched_csv + 10, len(csv_rows))):
        if ci >= 0:
            marker = ' <<<' if ci == last_matched_csv else ''
            print(f"  CSV[{ci}]: '{csv_rows[ci]['issuer'][:50]}' (CUSIP: {csv_rows[ci]['cusip']}){marker}")


# Try with lookahead: if current CSV doesn't match, look ahead up to 5 CSV rows
print("\n\n=== Sequential matching WITH lookahead ===")
matched2 = {}
matched_csv2 = set()
excel_ptr2 = 0
skipped_csv = 0
skipped_excel = 0

ci = 0
while ci < len(csv_rows) and excel_ptr2 < len(excel_rows):
    e = excel_rows[excel_ptr2]
    c = csv_rows[ci]

    if issuer_match(e['issuer'], c['issuer']):
        matched2[excel_ptr2] = ci
        matched_csv2.add(ci)
        excel_ptr2 += 1
        ci += 1
    else:
        # Look ahead in CSV (up to 5 rows) for a match
        found = False
        for ahead in range(1, 6):
            if ci + ahead < len(csv_rows):
                if issuer_match(e['issuer'], csv_rows[ci + ahead]['issuer']):
                    # Skip the non-matching CSV rows (duplicates)
                    skipped_csv += ahead
                    ci += ahead
                    matched2[excel_ptr2] = ci
                    matched_csv2.add(ci)
                    excel_ptr2 += 1
                    ci += 1
                    found = True
                    break

        if not found:
            # Look ahead in Excel (up to 3 rows) - maybe Excel has a row the CSV skipped
            found_excel = False
            for ea in range(1, 4):
                if excel_ptr2 + ea < len(excel_rows):
                    if issuer_match(excel_rows[excel_ptr2 + ea]['issuer'], c['issuer']):
                        skipped_excel += ea
                        excel_ptr2 += ea
                        matched2[excel_ptr2] = ci
                        matched_csv2.add(ci)
                        excel_ptr2 += 1
                        ci += 1
                        found_excel = True
                        break

            if not found_excel:
                # Can't match - skip CSV row
                ci += 1
                skipped_csv += 1

print(f"Matched: {len(matched2)}/{len(excel_rows)}")
print(f"Unmatched Excel rows: {len(excel_rows) - len(matched2)}")
print(f"Skipped CSV rows: {skipped_csv}")
print(f"Skipped Excel rows: {skipped_excel}")
print(f"Excel pointer: {excel_ptr2}/{len(excel_rows)}")

# Verify issuer alignment
mismatches = 0
for ei, ci in sorted(matched2.items()):
    e_issuer = excel_rows[ei]['issuer']
    c_issuer = csv_rows[ci]['issuer']
    if not issuer_match(e_issuer, c_issuer, min_prefix=5):
        mismatches += 1
        if mismatches <= 20:
            print(f"  MISMATCH: Excel[{ei}] '{e_issuer[:40]}' vs CSV[{ci}] '{c_issuer[:30]}'")

print(f"\nIssuer mismatches in matched pairs: {mismatches}")

# Show unmatched Excel rows
unmatched_excel = [i for i in range(len(excel_rows)) if i not in matched2]
if unmatched_excel:
    print(f"\nUnmatched Excel rows (first 30):")
    for ei in unmatched_excel[:30]:
        print(f"  Excel[{ei}]: '{excel_rows[ei]['issuer'][:50]}' (BB_ID: {excel_rows[ei]['bb_id']})")
