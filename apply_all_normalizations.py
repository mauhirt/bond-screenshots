#!/usr/bin/env python3
"""
Apply ALL normalizations to the v3 output in one pass:
1. Tax Prov → valid Bloomberg categories
2. Fin Typ → NEW MONEY / REFUNDING
3. BICS Level 2 → valid Bloomberg categories
4. Industry → valid Bloomberg categories
5. ESG Project Categories → valid Bloomberg categories
6. Project Subcategory → valid Bloomberg categories
7. Date fixes (swap issue/maturity if reversed, clear non-date strings)
"""
import openpyxl
import re
from datetime import datetime

# ── Valid Categories ──

VALID_TAX_PROV = [
    'AMT/ST TAX-EXEMPT', 'AMT/ST TAXABLE', 'FED & ST TAX-EXEMPT',
    'FED AMT FOR INDIVIDUALS', 'FED BQ', 'FED BQ/ST TAX-EXEMPT',
    'FED BQ/ST TAXABLE', 'FED TAX-EXEMPT', 'FED TAX-EXEMPT/ST TAXABLE',
    'FED TAXABLE', 'FED TAXABLE/ST TAX-EXEMPT', 'FED TAXABLE/ST TAXABLE',
]

VALID_INDUSTRY = [
    'APPROP', 'ARPT', 'BONDBK', 'CCRC', 'CDD', 'CHRT', 'CMNTYC', 'DEV',
    'EDU', 'EDULEASE', 'GARVEE', 'GASFWD', 'GO', 'GODIST', 'GOVLEASE',
    'GOVTGTD', 'HGR', 'HOSP', 'HOTELTAX', 'INCTAX', 'LMFH', 'LNGBDL',
    'LNPOOL', 'MDD', 'MEL', 'MISC', 'MISCTAX', 'MUNUTIL', 'NA', 'NFPCULT',
    'NFPRO', 'PILOT', 'PORTS', 'PUBPWR', 'PUBTRAN', 'PUBWTR', 'SALESTAX',
    'SCD', 'SCO', 'SELFAPP', 'SMFH', 'SOLWST', 'SPLASMT', 'STDHSG', 'TIF',
    'TOLL', 'TRIBES', 'TXMUD', 'WSGTD', 'WTRSWR',
]

VALID_BICS = [
    'Education', 'Financing', 'General Government', 'Health Care',
    'Housing', 'NA', 'Post Employment', 'Public Services',
    'Transportation', 'Utilities',
]

ESG_BASE = [
    'Sustainable Water', 'Energy Efficiency', 'Green Buildings',
    'Renewable Energy', 'Natural Resource Management', 'Pollution Control',
    'Circular Economy', 'Clean Transportation', 'Climate Change Adaptation',
    'Biodiversity',
]

VALID_SUBCAT_BASE = {
    'Bioenergy', 'BREEAM Certified', 'Circular Design and Production',
    'Circular Value Recovery', 'Conservation', 'Energy Star Certified',
    'Energy Storage', 'Geothermal', 'Green House Gas Control',
    'Greenhouse Gas Control', 'Hydro', 'Hydrogen', 'Infrastructure',
    'Information Support', 'LEED Certified', 'Marine', 'Multimodal',
    'Non Motorized', 'Plumbing System', 'Pollution Control', 'Public',
    'Rail (Non Passenger)', 'Smart Grids', 'Soil Remediation', 'Solar',
    'Sustainable Forestry', 'Vehicles', 'Waste Management',
    'WELL Certified', 'Wind',
}


# ── Tax Prov Normalization ──

def clean_tax_prov(raw):
    if not raw:
        return None
    s = str(raw).strip().upper()
    if not s or s == '--' or s == 'NA':
        return None

    # Direct match
    for valid in VALID_TAX_PROV:
        if s == valid:
            return valid

    # Pattern matching
    # FED & ST TAX-EXEMPT
    if re.match(r'^FED\s*[&R]\s*ST\s*TAX[\s-]*(EX[EI]?MPT|CBMPT|EXEHPT)', s):
        return 'FED & ST TAX-EXEMPT'
    if re.match(r'^FE[OD]\s*[&R]\s*S[TL]\s*TAX[\s-]*EX', s):
        return 'FED & ST TAX-EXEMPT'
    if re.match(r'^FD\s*[&R]\s*ST\s*TAX', s):
        return 'FED & ST TAX-EXEMPT'

    # FED TAX-EXEMPT
    if re.match(r'^FED\s*TAX[\s-]*(EX[EI]?MPT|CBMPT|EXEHPT)$', s):
        return 'FED TAX-EXEMPT'
    if re.match(r'^FE[OD]\s*TAX[\s-]*(EX[EI]?MPT|EXEHPT)$', s):
        return 'FED TAX-EXEMPT'
    if re.match(r'^FED?\s*TAX[\s-]*EX[A-Z]*$', s):
        return 'FED TAX-EXEMPT'

    # FED TAX-EXEMPT/ST TAXABLE
    if re.match(r'^FED\s*TAX[\s-]*EX[A-Z]*/\s*ST\s*TAX(ABLE|ARLE)', s):
        return 'FED TAX-EXEMPT/ST TAXABLE'
    if 'TAX-EXEMPT/ST TAXAB' in s or 'TAX-EXEMPT/ST TAXABL' in s:
        return 'FED TAX-EXEMPT/ST TAXABLE'

    # FED TAXABLE
    if re.match(r'^FED\s*TAXAB', s) and '/ST' not in s:
        return 'FED TAXABLE'
    if re.match(r'^FE[OD]\s*TAXAB', s) and '/ST' not in s:
        return 'FED TAXABLE'

    # FED TAXABLE/ST TAX-EXEMPT
    if re.match(r'^FED\s*TAXAB.*/?S?T?\s*TAX[\s-]*EX', s):
        if '/ST TAX' in s or 'ST TAX-EX' in s or 'STTAX' in s:
            return 'FED TAXABLE/ST TAX-EXEMPT'
    if 'TAXABLE/ST TAX-EX' in s or 'TXBL&/ST TAX-EX' in s:
        return 'FED TAXABLE/ST TAX-EXEMPT'
    if 'TAXABLE/ST TAX-EXEMPT' in s:
        return 'FED TAXABLE/ST TAX-EXEMPT'

    # FED TAXABLE/ST TAXABLE
    if re.match(r'^FED\s*TAXAB.*/?S?T?\s*TAXAB', s):
        return 'FED TAXABLE/ST TAXABLE'

    # FED BQ
    if re.match(r'^FED\s*B[QUO]$', s):
        return 'FED BQ'
    if re.match(r'^FE[OD]\s*B[QUO]$', s):
        return 'FED BQ'

    # FED BQ/ST TAX-EXEMPT
    if re.match(r'^FED?\s*B[QUOV].*/?.*ST.*TAX[\s-]*EX', s):
        return 'FED BQ/ST TAX-EXEMPT'
    if re.match(r'^FD?\s*BY?\s*ST\s*TAX[\s-]*EX', s):
        return 'FED BQ/ST TAX-EXEMPT'

    # FED BQ/ST TAXABLE
    if re.match(r'^FED\s*B[QUO].*/?.*ST.*TAXAB', s):
        return 'FED BQ/ST TAXABLE'

    # FED AMT FOR INDIVIDUALS
    if 'AMT FOR' in s or 'AHT FOR' in s:
        return 'FED AMT FOR INDIVIDUALS'
    if re.match(r'^FED\s*AMT', s) and 'ST' not in s:
        return 'FED AMT FOR INDIVIDUALS'

    # AMT/ST TAX-EXEMPT
    if re.match(r'^AMT/?ST\s*TAX[\s-]*EX', s):
        return 'AMT/ST TAX-EXEMPT'
    if re.match(r'^AHT/?ST\s*TAX[\s-]*EX', s):
        return 'AMT/ST TAX-EXEMPT'
    if re.match(r'^AMT/ST TAX', s) and 'EXEMPT' in s:
        return 'AMT/ST TAX-EXEMPT'

    # AMT/ST TAXABLE
    if re.match(r'^AMT/?ST\s*TAXAB', s):
        return 'AMT/ST TAXABLE'

    # Broader patterns
    if 'FED' in s and 'BQ' in s and 'TAX-EX' in s:
        return 'FED BQ/ST TAX-EXEMPT'
    if 'FED' in s and 'BQ' in s and 'TAXAB' in s:
        return 'FED BQ/ST TAXABLE'
    if 'FED' in s and 'BQ' in s:
        return 'FED BQ'
    if 'FED' in s and 'ST TAX-EX' in s and 'TAXAB' not in s:
        if 'TAXABLE' in s.split('ST')[0] if 'ST' in s else False:
            return 'FED TAXABLE/ST TAX-EXEMPT'
        return 'FED & ST TAX-EXEMPT'
    if 'FED' in s and 'TAX-EX' in s:
        return 'FED TAX-EXEMPT'
    if 'FED' in s and 'TAXAB' in s:
        return 'FED TAXABLE'
    if 'AMT' in s and 'TAX-EX' in s:
        return 'AMT/ST TAX-EXEMPT'

    return None  # Unknown


# ── Fin Typ Normalization ──

def clean_fin_typ(raw):
    if not raw:
        return None
    s = str(raw).strip().upper()
    if not s or s == '--':
        return None
    if s in ('NEW MONEY',):
        return 'NEW MONEY'
    if s in ('REFUNDING',):
        return 'REFUNDING'
    if 'NEH' in s or 'NEW' in s or 'NEU' in s or 'NEN' in s:
        if 'MONEY' in s or 'HONEY' in s or 'HONEY' in s or 'MNEY' in s or 'HMEY' in s or 'HANEY' in s or 'HANCY' in s:
            return 'NEW MONEY'
    if 'REF' in s or 'REH' in s or 'REW' in s:
        if 'FUND' in s or 'FINANC' in s or 'FINAN' in s or 'FINMNG' in s or 'FINCING' in s or 'FINIDG' in s or 'FIN&MNG' in s or 'FINANG' in s:
            return 'REFUNDING'
        if 'MONEY' in s:
            return 'NEW MONEY'  # REF MONEY -> NEW MONEY? Actually this is ambiguous
    return None


# ── BICS Normalization ──

def clean_bics(raw):
    if not raw:
        return None
    s = str(raw).strip().rstrip('.')
    if not s or s == '--' or s.upper() == 'NA':
        return None
    for valid in VALID_BICS:
        if s.lower().startswith(valid.lower()[:5]):
            return valid
    return None


# ── Industry Normalization ──

def levenshtein(s1, s2):
    if len(s1) < len(s2):
        return levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev = list(range(len(s2) + 1))
    for i, c1 in enumerate(s1):
        curr = [i + 1]
        for j, c2 in enumerate(s2):
            curr.append(min(prev[j+1]+1, curr[j]+1, prev[j]+(c1 != c2)))
        prev = curr
    return prev[len(s2)]

def clean_industry(raw):
    if not raw:
        return None
    s = str(raw).strip().upper().rstrip('.')
    if not s or s == '--' or s == 'NA':
        return None

    # Direct match
    if s in VALID_INDUSTRY:
        return s

    # Known OCR patterns
    if 'TRSWR' in s or 'TRSUR' in s or 'TRSHR' in s or 'MRSWR' in s:
        return 'WTRSWR'
    if s.startswith('MTRSUR') or s.startswith('INTRSUR') or s.startswith('MFRSUR'):
        return 'WTRSWR'
    if s.startswith('GASFM') or s.startswith('GASFW') or s.startswith('GASF'):
        return 'GASFWD'
    if s.startswith('GODIS') or s.startswith('GDDIS') or s.startswith('GXDIS'):
        return 'GODIST'
    if s.startswith('WSGTD') or s.startswith('MSGTD') or s.startswith('HSGTD'):
        return 'WSGTD'
    if s.startswith('PUBW') or s.startswith('RUBW'):
        return 'PUBWTR'
    if s.startswith('PUBT') or s.startswith('RUBT'):
        return 'PUBTRAN'
    if s.startswith('PUBP') or s.startswith('RUBP'):
        return 'PUBPWR'
    if s.startswith('MUNUTIL') or s.startswith('HUNUTIL') or s.startswith('NUNUTIL'):
        return 'MUNUTIL'
    if s.startswith('SALES') or s.startswith('SALFSTAX') or s.startswith('SALEST'):
        return 'SALESTAX'
    if s.startswith('MISCT') or s.startswith('HISCT'):
        return 'MISCTAX'
    if 'GOVTGTD' in s or 'GOVTG' in s:
        return 'GOVTGTD'
    if s.startswith('HOTEL') or s.startswith('HOTELTAS'):
        return 'HOTELTAX'
    if s.startswith('SELFAPP') or s.startswith('SELFA'):
        return 'SELFAPP'
    if s.startswith('SPLASM') or s.startswith('SPLAS'):
        return 'SPLASMT'
    if s.startswith('STDHSG') or s.startswith('STDH'):
        return 'STDHSG'
    if s.startswith('SOLWS') or s.startswith('SOLW'):
        return 'SOLWST'
    if s.startswith('GARVEE') or s.startswith('GARV'):
        return 'GARVEE'
    if s.startswith('NFPCULT') or s.startswith('NFPC'):
        return 'NFPCULT'
    if s.startswith('NFPRO') or s.startswith('NFPR'):
        return 'NFPRO'
    if s.startswith('BONDBK') or s.startswith('BOND'):
        return 'BONDBK'
    if s.startswith('EDULEASE') or s.startswith('EDUL'):
        return 'EDULEASE'
    if s.startswith('GOVLEASE') or s.startswith('GOVL'):
        return 'GOVLEASE'
    if s.startswith('LNGBDL') or s.startswith('LNGB'):
        return 'LNGBDL'
    if s.startswith('LNPOOL') or s.startswith('LNPO'):
        return 'LNPOOL'
    if s.startswith('TXMUD') or s.startswith('TXMU'):
        return 'TXMUD'
    if s.startswith('TRIBE') or s.startswith('TRIB'):
        return 'TRIBES'
    if s.startswith('INCTAX') or s.startswith('INCT'):
        return 'INCTAX'

    # Filter out Yes/No leaks and other non-industry values
    if s in ('YES', 'NO', 'YES YES'):
        return None

    # Levenshtein fallback
    best_dist = 999
    best_match = None
    for valid in VALID_INDUSTRY:
        d = levenshtein(s, valid)
        if d < best_dist:
            best_dist = d
            best_match = valid
    max_dist = 3 if len(s) <= 6 else 4
    if best_dist <= max_dist:
        return best_match

    return None


# ── ESG Project Categories Normalization ──

ESG_EXPANSIONS = {
    'Green B': 'Green Buildings', 'Climat': 'Climate Change Adaptation',
    'Sustaina': 'Sustainable Water', 'Ener.': 'Energy Efficiency',
    'Gree': 'Green Buildings', 'POL': 'Pollution Control',
    'Renewa': 'Renewable Energy', 'Sust': 'Sustainable Water',
    'Circular': 'Circular Economy', 'Clean': 'Clean Transportation',
    'Biodi': 'Biodiversity', 'Natur': 'Natural Resource Management',
}

def clean_esg_category(raw):
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s == '--' or s.upper() == 'NA':
        return None

    # Strip STATE/CITY prefixes
    s = re.sub(r'^(STATE|CITY|CTTY|CNTY|ST|COUNTY)\s+', '', s, flags=re.IGNORECASE)

    # Split on separators
    parts = re.split(r'[;,/]', s)
    cleaned = []
    for part in parts:
        p = part.strip().rstrip('.')
        if not p:
            continue

        # Check for "greenhouse" or "bioenergy" first
        if 'greenhouse' in p.lower() or 'greenh' in p.lower():
            cleaned.append('Renewable Energy')
            continue
        if 'bioen' in p.lower():
            cleaned.append('Renewable Energy')
            continue

        # Direct match
        matched = False
        for base in ESG_BASE:
            if p.lower() == base.lower():
                cleaned.append(base)
                matched = True
                break
        if matched:
            continue

        # Expansion match
        for prefix, full in ESG_EXPANSIONS.items():
            if p.lower().startswith(prefix.lower()):
                cleaned.append(full)
                matched = True
                break
        if matched:
            continue

        # Partial match (first 6 chars)
        for base in ESG_BASE:
            if len(p) >= 4 and base.lower().startswith(p.lower()[:6]):
                cleaned.append(base)
                matched = True
                break
        if not matched and len(p) > 3:
            cleaned.append(p)  # Keep as-is if not matched

    if not cleaned:
        return None

    # Deduplicate
    seen = set()
    result = []
    for c in cleaned:
        if c not in seen:
            seen.add(c)
            result.append(c)

    return ', '.join(result)


# ── Project Subcategory Normalization ──

def normalize_subcat_single(s):
    if not s:
        return None
    u = s.strip().upper()
    for valid in VALID_SUBCAT_BASE:
        if u == valid.upper():
            return valid
    # OCR fixes
    if u in ('LED CERTIFIED', 'LEE CERTIFIED'):
        return 'LEED Certified'
    if 'RENEWABLE GAS' in u or 'GREEN HOUSE GAS' in u or 'GREENHOUSE GAS' in u:
        return 'Green House Gas Control'
    if 'RAIL' in u and 'PASSENGER' in u:
        return 'Rail (Non Passenger)'
    if u == 'RAIL':
        return 'Rail (Non Passenger)'
    if u.startswith('SOLAR (INCL') or u.startswith('SOLAR(INCL'):
        return 'Solar'
    if u.startswith('CIRCULAR DESIGN'):
        return 'Circular Design and Production'
    if u.startswith('CIRCULAR VALUE'):
        return 'Circular Value Recovery'
    if u.startswith('CIRCULAR WASTE') or u.startswith('CIRC'):
        return 'Circular Value Recovery'
    if u.startswith('WASTE RED') or u.startswith('WASTE MAN'):
        return 'Waste Management'
    if u.startswith('NON MOTOR'):
        return 'Non Motorized'
    if u.startswith('ENERGY STAR'):
        return 'Energy Star Certified'
    if u.startswith('ENERGY STOR'):
        return 'Energy Storage'
    if u.startswith('BREEAM'):
        return 'BREEAM Certified'
    if u.startswith('WELL'):
        return 'WELL Certified'
    if u.startswith('SUSTAINAB') and 'FOREST' in u:
        return 'Sustainable Forestry'
    if u.startswith('SUSTAINAB') and 'INFRA' in u:
        return 'Infrastructure'
    if u.startswith('INFO'):
        return 'Information Support'
    if u.startswith('POLLUT'):
        return 'Pollution Control'
    if u.startswith('SMART'):
        return 'Smart Grids'
    if u.startswith('SOIL'):
        return 'Soil Remediation'
    if u.startswith('MULTI'):
        return 'Multimodal'
    if u.startswith('PLUMB'):
        return 'Plumbing System'
    if 'SUB.' in u and 'ENERGY' in u:
        return 'Energy Storage'
    if u == 'PUBLIC':
        return 'Public'
    return None

def clean_subcategory(raw):
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s.upper() == 'NA':
        return None
    if s in ('Yes Yes', 'No', 'STAT.', '#NAME?'):
        return None

    # Normalize separators
    normalized = s.replace(';', ',')
    normalized = re.sub(r'(?<!\()/(?!\))', ',', normalized)
    normalized = re.sub(r'(?<=[a-z])-(?=[A-Z])', ', ', normalized)
    normalized = re.sub(r'(?<=[A-Z])-(?=[A-Z])', ', ', normalized)

    parts = [p.strip() for p in normalized.split(',') if p.strip()]
    expanded = []
    for part in parts:
        matched = normalize_subcat_single(part)
        if matched:
            expanded.append(matched)
            continue
        if part.upper().startswith('PUBLIC '):
            rest = part[7:].strip()
            rest_norm = normalize_subcat_single(rest)
            if rest_norm:
                expanded.append('Public')
                expanded.append(rest_norm)
                continue
        if part.upper().startswith('SOLAR '):
            rest = part[6:].strip()
            rest_norm = normalize_subcat_single(rest)
            if rest_norm:
                expanded.append('Solar')
                expanded.append(rest_norm)
                continue
        norm = normalize_subcat_single(part)
        if norm:
            expanded.append(norm)

    if not expanded:
        return None
    seen = set()
    deduped = []
    for p in expanded:
        if p not in seen:
            seen.add(p)
            deduped.append(p)
    return ', '.join(deduped)


# ── Main Application ──

print("Loading Excel...")
wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

stats = {
    'tax_prov': 0, 'fin_typ': 0, 'bics': 0, 'industry': 0,
    'esg': 0, 'subcat': 0, 'date_fix': 0, 'date_swap': 0
}

for row in range(2, ws.max_row + 1):
    # Tax Prov (col 8)
    val = ws.cell(row=row, column=8).value
    if val:
        cleaned = clean_tax_prov(val)
        if cleaned != str(val).strip():
            ws.cell(row=row, column=8).value = cleaned
            stats['tax_prov'] += 1

    # Fin Typ (col 9)
    val = ws.cell(row=row, column=9).value
    if val:
        cleaned = clean_fin_typ(val)
        if cleaned != str(val).strip():
            ws.cell(row=row, column=9).value = cleaned
            stats['fin_typ'] += 1

    # BICS Level 2 (col 10)
    val = ws.cell(row=row, column=10).value
    if val:
        cleaned = clean_bics(val)
        if cleaned != str(val).strip():
            ws.cell(row=row, column=10).value = cleaned
            stats['bics'] += 1

    # Industry (col 17)
    val = ws.cell(row=row, column=17).value
    if val:
        cleaned = clean_industry(val)
        if cleaned != str(val).strip():
            ws.cell(row=row, column=17).value = cleaned
            stats['industry'] += 1

    # ESG Project Categories (col 19)
    val = ws.cell(row=row, column=19).value
    if val:
        cleaned = clean_esg_category(val)
        if cleaned != str(val).strip():
            ws.cell(row=row, column=19).value = cleaned
            stats['esg'] += 1

    # Project Subcategory (col 20)
    val = ws.cell(row=row, column=20).value
    if val:
        cleaned = clean_subcategory(val)
        if cleaned != str(val).strip():
            ws.cell(row=row, column=20).value = cleaned
            stats['subcat'] += 1

    # Date fixes (col 6 = Issue Date, col 7 = Maturity)
    issue_date = ws.cell(row=row, column=6).value
    maturity = ws.cell(row=row, column=7).value

    # Clear non-date strings
    if issue_date and isinstance(issue_date, str):
        ws.cell(row=row, column=6).value = None
        stats['date_fix'] += 1
    if maturity and isinstance(maturity, str):
        ws.cell(row=row, column=7).value = None
        stats['date_fix'] += 1

    # Swap if Issue Date > Maturity
    issue_date = ws.cell(row=row, column=6).value
    maturity = ws.cell(row=row, column=7).value
    if isinstance(issue_date, datetime) and isinstance(maturity, datetime):
        if issue_date > maturity:
            ws.cell(row=row, column=6).value = maturity
            ws.cell(row=row, column=7).value = issue_date
            stats['date_swap'] += 1

wb.save('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')

print(f"\nNormalization results:")
for key, count in stats.items():
    print(f"  {key}: {count} fixes")

# Verify all categorical columns
print("\n\nValidation:")
wb2 = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws2 = wb2.active

from collections import Counter

# Tax Prov
tax_vals = Counter()
tax_invalid = 0
for row in range(2, ws2.max_row + 1):
    v = ws2.cell(row=row, column=8).value
    if v:
        tax_vals[str(v)] += 1
        if str(v) not in VALID_TAX_PROV:
            tax_invalid += 1

print(f"\nTax Prov: {sum(tax_vals.values())} non-null, {len(tax_vals)} unique, {tax_invalid} invalid")
if tax_invalid:
    for v, c in tax_vals.most_common():
        if v not in VALID_TAX_PROV:
            print(f"  INVALID: [{c}] '{v}'")

# Industry
ind_vals = Counter()
ind_invalid = 0
for row in range(2, ws2.max_row + 1):
    v = ws2.cell(row=row, column=17).value
    if v:
        ind_vals[str(v)] += 1
        if str(v) not in VALID_INDUSTRY:
            ind_invalid += 1

print(f"\nIndustry: {sum(ind_vals.values())} non-null, {len(ind_vals)} unique, {ind_invalid} invalid")
if ind_invalid:
    for v, c in ind_vals.most_common():
        if v not in VALID_INDUSTRY:
            print(f"  INVALID: [{c}] '{v}'")

# Fin Typ
fin_vals = Counter()
for row in range(2, ws2.max_row + 1):
    v = ws2.cell(row=row, column=9).value
    if v:
        fin_vals[str(v)] += 1
fin_invalid = sum(c for v, c in fin_vals.items() if v not in ('NEW MONEY', 'REFUNDING'))
print(f"\nFin Typ: {sum(fin_vals.values())} non-null, {fin_invalid} invalid")
if fin_invalid:
    for v, c in fin_vals.most_common():
        if v not in ('NEW MONEY', 'REFUNDING'):
            print(f"  INVALID: [{c}] '{v}'")

# BICS
bics_vals = Counter()
for row in range(2, ws2.max_row + 1):
    v = ws2.cell(row=row, column=10).value
    if v:
        bics_vals[str(v)] += 1
bics_invalid = sum(c for v, c in bics_vals.items() if v not in VALID_BICS)
print(f"\nBICS: {sum(bics_vals.values())} non-null, {bics_invalid} invalid")
if bics_invalid:
    for v, c in bics_vals.most_common():
        if v not in VALID_BICS:
            print(f"  INVALID: [{c}] '{v}'")
