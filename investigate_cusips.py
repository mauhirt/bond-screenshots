#!/usr/bin/env python3
"""Investigate CUSIP issues: compare CSV CUSIPs, BB_IDs, and output Excel CUSIPs."""
import openpyxl
import re
import csv

# 1. Extract BB_IDs from original Excel formulas
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

bb_ids = {}  # row -> bb_id
for row in range(2, ws_orig.max_row + 1):
    # BB_ID is in the formula strings - try multiple columns
    for col in range(1, 21):
        val = ws_orig.cell(row=row, column=col).value
        if val and isinstance(val, str) and 'Muni' in val:
            m = re.search(r'"([^"]+)\s+Muni"', val)
            if m:
                bb_ids[row] = m.group(1)
                break

print(f"Original Excel: {ws_orig.max_row - 1} data rows, {len(bb_ids)} BB_IDs extracted")

# 2. Load output Excel
wb_out = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws_out = wb_out.active

print(f"Output Excel: {ws_out.max_row - 1} data rows")

# 3. Show first 20 rows side by side: BB_ID from original vs CUSIP in output
print(f"\nFirst 30 rows comparison:")
print(f"{'Row':>4} {'BB_ID (orig)':>15} {'CUSIP (output)':>15} {'Match?':>8} {'Issuer':>40}")
for row in range(2, 32):
    bb_id = bb_ids.get(row, 'N/A')
    cusip = str(ws_out.cell(row=row, column=1).value or 'N/A').strip()
    issuer = str(ws_out.cell(row=row, column=3).value or '')[:40]

    # Check if BB_ID is a prefix/substring of CUSIP (Bloomberg IDs are often CUSIP without check digit)
    # Actually BB_IDs for munis can be quite different from CUSIPs
    match = '?'
    if bb_id != 'N/A' and cusip != 'N/A':
        # Remove common OCR substitutions and compare
        if bb_id in cusip or cusip.startswith(bb_id):
            match = 'YES'
        elif bb_id[:6] == cusip[:6]:
            match = 'CLOSE'
        else:
            match = 'NO'

    print(f"{row:4d} {bb_id:>15} {cusip:>15} {match:>8} {issuer}")

# 4. Load CSV and check what CUSIPs are in the first column
print(f"\n\nCSV first column (first 30 data rows):")
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    header = next(reader)
    print(f"CSV headers: {header[:5]}")
    for i, row_data in enumerate(reader):
        if i >= 30:
            break
        if row_data:
            print(f"  CSV row {i+1}: col0='{row_data[0][:20]}' col1='{row_data[1][:20] if len(row_data) > 1 else 'N/A'}' col2='{row_data[2][:20] if len(row_data) > 2 else 'N/A'}'")

# 5. Check the populate script to understand how CUSIP was assigned
print("\n\n=== Checking how CUSIPs were populated ===")
# Look at the relationship between BB_ID and the CUSIPs that ended up in the output
# For the first 10 rows, show BB_ID and all output columns
for row in range(2, 12):
    bb_id = bb_ids.get(row, 'N/A')
    cusip_out = ws_out.cell(row=row, column=1).value
    state_out = ws_out.cell(row=row, column=2).value
    issuer_out = ws_out.cell(row=row, column=3).value
    yield_out = ws_out.cell(row=row, column=4).value
    amt_out = ws_out.cell(row=row, column=5).value

    issuer_orig = ws_orig.cell(row=row, column=3).value
    amt_orig = ws_orig.cell(row=row, column=5).value

    print(f"\nRow {row}:")
    print(f"  BB_ID:         {bb_id}")
    print(f"  CUSIP (out):   {cusip_out}")
    print(f"  State (out):   {state_out}")
    print(f"  Issuer (out):  {str(issuer_out)[:50]}")
    print(f"  Issuer (orig): {str(issuer_orig)[:50]}")
    print(f"  Yield (out):   {yield_out}")
    print(f"  Amt (out):     {amt_out}")
    print(f"  Amt (orig):    {amt_orig}")
