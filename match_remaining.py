#!/usr/bin/env python3
"""
Match remaining 252 Excel rows using broader issuer name + CUSIP matching.
Also check if these issuers exist in the CSV at all.
"""
import csv
import re
import openpyxl
from collections import defaultdict

# (reuse functions from build_matching_v3.py)
OCR_PAIRS = set()
_pairs = [
    ('0','X'),('0','O'),('0','D'),
    ('1','I'),('1','L'),('1','7'),
    ('3','2'),('3','8'),
    ('4','8'),('4','A'),
    ('5','S'),('5','F'),('5','6'),
    ('6','G'),('6','8'),
    ('7','T'),
    ('8','B'),
    ('9','G'),('9','Q'),
    ('D','K'),('D','H'),
    ('E','F'),
    ('K','H'),('K','X'),
    ('M','N'),('M','W'),
    ('P','R'),
    ('U','V'),('W','V'),
]
for a, b in _pairs:
    OCR_PAIRS.add((a, b))
    OCR_PAIRS.add((b, a))

def cusip_check_digit(base8):
    values = []
    for ch in base8.upper():
        if ch.isdigit():
            values.append(int(ch))
        elif ch.isalpha():
            values.append(ord(ch) - ord('A') + 10)
        else:
            values.append(0)
    total = 0
    for i, v in enumerate(values):
        if i % 2 == 1:
            v *= 2
        total += v // 10 + v % 10
    return str((10 - (total % 10)) % 10)

def cusip_ocr_score(correct9, csv_cusip):
    best = 0
    def score_pair(a, b):
        s = 0
        for x, y in zip(a, b):
            if x == y: s += 3
            elif x.upper() == y.upper(): s += 2
            elif (x, y) in OCR_PAIRS: s += 1
        return s

    clen = len(csv_cusip)
    if clen == 9:
        best = score_pair(correct9, csv_cusip)
    elif clen == 8:
        best = score_pair(correct9[:8], csv_cusip)
    elif clen == 10:
        for skip in range(10):
            trimmed = csv_cusip[:skip] + csv_cusip[skip+1:]
            best = max(best, score_pair(correct9, trimmed))
    elif clen == 7:
        best = score_pair(correct9[:7], csv_cusip)
    else:
        ml = min(9, clen)
        best = score_pair(correct9[:ml], csv_cusip[:ml])
    return best

def issuer_prefix_match(e, c):
    e = e.lower().strip()
    c = c.lower().strip().rstrip('.')
    match = 0
    for a, b in zip(e, c):
        if a == b: match += 1
        else: break
    return match

def fuzzy_issuer_match(e, c, threshold=0.5):
    """Fuzzy issuer match: check if enough chars match (not just prefix)."""
    e = e.lower().strip()[:20]
    c = c.lower().strip().rstrip('.')[:20]
    if not e or not c:
        return 0
    matches = sum(1 for a, b in zip(e, c) if a == b)
    return matches / max(len(e), len(c))

# Load data
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    amt = ws_orig.cell(row=row, column=5).value
    formula = ws_orig.cell(row=row, column=2).value or ''
    m = re.search(r'"([^"]+)\s+Muni"', str(formula))
    bb_id = m.group(1) if m else ''
    correct_cusip = ''
    if bb_id and len(bb_id) >= 8:
        correct_cusip = bb_id[:8] + cusip_check_digit(bb_id[:8])
    excel_rows.append({
        'issuer': issuer, 'amt': amt, 'bb_id': bb_id,
        'correct_cusip': correct_cusip, 'idx': len(excel_rows)
    })

csv_rows = []
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    for row_data in reader:
        if len(row_data) >= 10:
            cusip = row_data[0].strip()
            if cusip.startswith('TH '):
                cusip = cusip[3:]
            if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                if not all(f.strip() == '--' for f in row_data[:5]):
                    csv_rows.append({
                        'cusip': cusip,
                        'issuer': row_data[2].strip(),
                        'fields': row_data,
                        'csv_idx': len(csv_rows)
                    })

# First, check which unmatched issuers exist in the CSV
# Group CSV by issuer first word
csv_by_word = defaultdict(list)
for ci, c in enumerate(csv_rows):
    words = c['issuer'].lower().strip().rstrip('.').split()
    if words:
        csv_by_word[words[0]].append(ci)

# Unmatched issuers (from previous run)
unmatched_issuers = [
    'Massachusetts Clean Water Trust/The',
    'Oklahoma Water Resources Board',
    'Metropolitan Transportation Authority',
    'Ohio Water Development Authority Water P',
    'Middleton-Cross Plains Area School Distr',
    'Bay Area Toll Authority',
    'Eco Maine',
    'Anchor Bay School District',
]

print("Checking if unmatched issuers exist in CSV:")
for iss in unmatched_issuers:
    first_word = iss.lower().split()[0]
    candidates = csv_by_word.get(first_word, [])
    if candidates:
        sample = [csv_rows[ci]['issuer'][:40] for ci in candidates[:5]]
        print(f"  '{iss[:40]}' -> found {len(candidates)} CSV rows: {sample}")
    else:
        # Try second word or partial
        found = False
        for word, indices in csv_by_word.items():
            if word[:4] == first_word[:4]:
                sample = [csv_rows[ci]['issuer'][:40] for ci in indices[:3]]
                print(f"  '{iss[:40]}' -> approx match '{word}': {len(indices)} rows: {sample}")
                found = True
                break
        if not found:
            # Search all CSV for any fuzzy match
            best_score = 0
            best_ci = -1
            for ci, c in enumerate(csv_rows):
                score = fuzzy_issuer_match(iss, c['issuer'])
                if score > best_score:
                    best_score = score
                    best_ci = ci
            if best_ci >= 0:
                print(f"  '{iss[:40]}' -> best fuzzy ({best_score:.0%}): '{csv_rows[best_ci]['issuer'][:40]}'")
            else:
                print(f"  '{iss[:40]}' -> NOT FOUND in CSV")

# Now do a broader matching pass for ALL Excel rows
# Phase 1: strict matching (from build_matching_v3.py results - we redo it here)
csv_by_issuer = defaultdict(list)
for ci, c in enumerate(csv_rows):
    key = c['issuer'].lower().strip()[:6]
    csv_by_issuer[key].append(ci)

csv_by_cusip4 = defaultdict(list)
for ci, c in enumerate(csv_rows):
    csv_by_cusip4[c['cusip'][:4]].append(ci)

print("\n\nPhase 1: Strict matching (CUSIP + issuer)...")
all_pairs = []
for ei, e in enumerate(excel_rows):
    if not e['correct_cusip']:
        continue
    candidates = set()
    issuer_key = e['issuer'].lower().strip()[:6]
    for key, indices in csv_by_issuer.items():
        if key[:4] == issuer_key[:4]:
            candidates.update(indices)
    cusip4 = e['correct_cusip'][:4]
    for key, indices in csv_by_cusip4.items():
        compat = sum(1 for a, b in zip(cusip4, key) if a == b or (a, b) in OCR_PAIRS) >= 2
        if compat:
            candidates.update(indices)
    for ci in candidates:
        c = csv_rows[ci]
        cusip_sc = cusip_ocr_score(e['correct_cusip'], c['cusip'])
        issuer_sc = issuer_prefix_match(e['issuer'], c['issuer'])
        total = cusip_sc + min(issuer_sc, 15) * 2
        if total >= 25 and (cusip_sc >= 12 or issuer_sc >= 8):
            all_pairs.append((total, cusip_sc, issuer_sc, ei, ci))

all_pairs.sort(key=lambda x: (-x[0], -x[1]))
matched = {}
matched_csv = set()
for total, cusip_sc, issuer_sc, ei, ci in all_pairs:
    if ei in matched or ci in matched_csv:
        continue
    matched[ei] = ci
    matched_csv.add(ci)

print(f"Phase 1: {len(matched)} matched")

# Phase 2: For unmatched, use issuer-only matching (search ALL CSV rows)
print("\nPhase 2: Issuer-focused matching for remaining rows...")
unmatched_ei = [i for i in range(len(excel_rows)) if i not in matched]
print(f"  Unmatched: {len(unmatched_ei)}")

phase2_pairs = []
for ei in unmatched_ei:
    e = excel_rows[ei]
    if not e['correct_cusip']:
        continue
    for ci, c in enumerate(csv_rows):
        if ci in matched_csv:
            continue
        # Check issuer similarity (fuzzy)
        issuer_score = fuzzy_issuer_match(e['issuer'], c['issuer'], threshold=0.4)
        if issuer_score < 0.4:
            continue
        cusip_sc = cusip_ocr_score(e['correct_cusip'], c['cusip'])
        prefix_sc = issuer_prefix_match(e['issuer'], c['issuer'])
        total = cusip_sc + prefix_sc * 2
        if total >= 15:  # Lower threshold
            phase2_pairs.append((total, cusip_sc, prefix_sc, issuer_score, ei, ci))

phase2_pairs.sort(key=lambda x: (-x[0], -x[3]))
new_matches = 0
for total, cusip_sc, prefix_sc, issuer_sc, ei, ci in phase2_pairs:
    if ei in matched or ci in matched_csv:
        continue
    matched[ei] = ci
    matched_csv.add(ci)
    new_matches += 1

print(f"Phase 2: {new_matches} new matches, Total: {len(matched)}")

# Phase 3: For still unmatched, try very broad search
unmatched_ei2 = [i for i in range(len(excel_rows)) if i not in matched]
print(f"\nPhase 3: Broad search for remaining {len(unmatched_ei2)} rows...")

phase3_pairs = []
for ei in unmatched_ei2:
    e = excel_rows[ei]
    if not e['correct_cusip']:
        continue
    best_score = 0
    best_ci = -1
    best_issuer_match = 0
    for ci, c in enumerate(csv_rows):
        if ci in matched_csv:
            continue
        cusip_sc = cusip_ocr_score(e['correct_cusip'], c['cusip'])
        issuer_sc = fuzzy_issuer_match(e['issuer'], c['issuer'])
        total = cusip_sc + issuer_sc * 30  # Weight issuer heavily
        if total > best_score:
            best_score = total
            best_ci = ci
            best_issuer_match = issuer_sc
    if best_ci >= 0 and best_issuer_match >= 0.35:
        phase3_pairs.append((best_score, best_issuer_match, ei, best_ci))

phase3_pairs.sort(key=lambda x: -x[0])
new_matches3 = 0
for score, issuer_sc, ei, ci in phase3_pairs:
    if ei in matched or ci in matched_csv:
        continue
    matched[ei] = ci
    matched_csv.add(ci)
    new_matches3 += 1

print(f"Phase 3: {new_matches3} new matches, Total: {len(matched)}")

# Final stats
unmatched_final = [i for i in range(len(excel_rows)) if i not in matched]
print(f"\nFinal: {len(matched)}/{len(excel_rows)} matched, {len(unmatched_final)} unmatched")

# Validate all matches
issuer_ok = 0
issuer_close = 0
issuer_bad = 0
for ei, ci in matched.items():
    score = fuzzy_issuer_match(excel_rows[ei]['issuer'], csv_rows[ci]['issuer'])
    prefix = issuer_prefix_match(excel_rows[ei]['issuer'], csv_rows[ci]['issuer'])
    if prefix >= 8 or score >= 0.7:
        issuer_ok += 1
    elif prefix >= 4 or score >= 0.5:
        issuer_close += 1
    else:
        issuer_bad += 1
        if issuer_bad <= 15:
            print(f"  BAD: [{ei}] '{excel_rows[ei]['issuer'][:35]}' vs '{csv_rows[ci]['issuer'][:25]}' "
                  f"(cusip={excel_rows[ei]['correct_cusip']} csv={csv_rows[ci]['cusip']})")

print(f"\nIssuer validation: {issuer_ok} good, {issuer_close} close, {issuer_bad} bad")

# Show remaining unmatched
print(f"\nRemaining unmatched ({len(unmatched_final)}):")
iss_groups = defaultdict(int)
for ei in unmatched_final:
    iss_groups[excel_rows[ei]['issuer'][:45]] += 1
for iss, cnt in sorted(iss_groups.items(), key=lambda x: -x[1])[:15]:
    print(f"  [{cnt:3d}] {iss}")
