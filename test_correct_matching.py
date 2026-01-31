#!/usr/bin/env python3
"""
Test new matching: correct CUSIP (from BB_ID) vs OCR'd CSV CUSIP.

For each Excel row:
1. Compute correct CUSIP = BB_ID[:8] + check_digit
2. Compare against all CSV CUSIPs using OCR-aware scoring
3. Pick the best match (with strict threshold)
"""
import csv
import re
import openpyxl
from collections import defaultdict

OCR_PAIRS = {
    ('0','X'),('X','0'),('0','O'),('O','0'),('0','D'),('D','0'),
    ('1','I'),('I','1'),('1','L'),('L','1'),('1','7'),('7','1'),
    ('3','2'),('2','3'),('3','8'),('8','3'),
    ('4','8'),('8','4'),('4','A'),('A','4'),
    ('5','S'),('S','5'),('5','F'),('F','5'),('5','6'),('6','5'),
    ('6','G'),('G','6'),('6','8'),('8','6'),
    ('7','T'),('T','7'),
    ('8','B'),('B','8'),
    ('9','G'),('G','9'),('9','Q'),('Q','9'),
    ('D','K'),('K','D'),('D','H'),('H','D'),
    ('E','F'),('F','E'),
    ('K','H'),('H','K'),('K','X'),('X','K'),
    ('M','N'),('N','M'),('M','W'),('W','M'),
    ('P','R'),('R','P'),
    ('U','V'),('V','U'),('W','V'),('V','W'),
}

def cusip_check_digit(base8):
    """Compute CUSIP check digit for an 8-character base."""
    values = []
    for ch in base8.upper():
        if ch.isdigit():
            values.append(int(ch))
        elif ch.isalpha():
            values.append(ord(ch) - ord('A') + 10)
        elif ch == '*':
            values.append(36)
        elif ch == '@':
            values.append(37)
        elif ch == '#':
            values.append(38)
        else:
            values.append(0)
    total = 0
    for i, v in enumerate(values):
        if i % 2 == 1:
            v *= 2
        total += v // 10 + v % 10
    return str((10 - (total % 10)) % 10)


def ocr_score(correct, csv_cusip):
    """Score OCR similarity between correct CUSIP and CSV CUSIP.
    Returns (score, max_possible)."""
    # Try multiple alignments for different CSV CUSIP lengths
    best_score = 0
    best_max = 0

    # Direct comparison (9 vs 9, or truncated)
    c9 = correct[:9]

    # Case 1: CSV CUSIP is 9 chars (normal)
    if len(csv_cusip) == 9:
        score = 0
        for a, b in zip(c9, csv_cusip):
            if a == b:
                score += 3
            elif a.upper() == b.upper():
                score += 2
            elif (a, b) in OCR_PAIRS or (b, a) in OCR_PAIRS:
                score += 1
        best_score = max(best_score, score)
        best_max = 27

    # Case 2: CSV CUSIP is 8 chars (missing check digit)
    elif len(csv_cusip) == 8:
        score = 0
        for a, b in zip(c9[:8], csv_cusip):
            if a == b:
                score += 3
            elif a.upper() == b.upper():
                score += 2
            elif (a, b) in OCR_PAIRS or (b, a) in OCR_PAIRS:
                score += 1
        best_score = max(best_score, score)
        best_max = 24

    # Case 3: CSV CUSIP is 10 chars (extra OCR char) - try removing each position
    elif len(csv_cusip) == 10:
        for skip in range(10):
            trimmed = csv_cusip[:skip] + csv_cusip[skip+1:]
            score = 0
            for a, b in zip(c9, trimmed):
                if a == b:
                    score += 3
                elif a.upper() == b.upper():
                    score += 2
                elif (a, b) in OCR_PAIRS or (b, a) in OCR_PAIRS:
                    score += 1
            if score > best_score:
                best_score = score
                best_max = 27

    # Case 4: Other lengths - compare first min(len) chars
    else:
        min_len = min(9, len(csv_cusip))
        score = 0
        for a, b in zip(c9[:min_len], csv_cusip[:min_len]):
            if a == b:
                score += 3
            elif a.upper() == b.upper():
                score += 2
            elif (a, b) in OCR_PAIRS or (b, a) in OCR_PAIRS:
                score += 1
        best_score = max(best_score, score)
        best_max = min_len * 3

    return best_score, best_max


# Load Excel BB_IDs
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    amt = ws_orig.cell(row=row, column=5).value
    formula = ws_orig.cell(row=row, column=2).value or ''
    m = re.search(r'"([^"]+)\s+Muni"', str(formula))
    bb_id = m.group(1) if m else ''
    correct_cusip = (bb_id[:8] + cusip_check_digit(bb_id[:8])) if bb_id and len(bb_id) >= 8 else ''
    excel_rows.append({
        'issuer': issuer, 'amt': amt, 'bb_id': bb_id,
        'correct_cusip': correct_cusip, 'idx': row - 2
    })

# Load CSV
csv_rows = []
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    for row_data in reader:
        if len(row_data) >= 10:
            cusip = row_data[0].strip()
            if cusip.startswith('TH '):
                cusip = cusip[3:]
            if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                if not all(f.strip() == '--' for f in row_data[:5]):
                    csv_rows.append({
                        'cusip': cusip,
                        'issuer': row_data[2].strip(),
                        'fields': row_data,
                        'csv_idx': len(csv_rows)
                    })

print(f"Excel rows: {len(excel_rows)}")
print(f"CSV rows: {len(csv_rows)}")
print(f"Excel rows with correct CUSIPs: {sum(1 for e in excel_rows if e['correct_cusip'])}")

# Match: for each Excel row, find best CSV match
print("\nMatching...")
matched = {}  # excel_idx -> csv_idx
matched_csv = set()

# Build pairs with scores
pairs = []
for ei, e in enumerate(excel_rows):
    if not e['correct_cusip']:
        continue
    for ci, c in enumerate(csv_rows):
        score, max_score = ocr_score(e['correct_cusip'], c['cusip'])
        if max_score > 0 and score >= max_score * 0.65:  # At least 65% match
            pairs.append((score, max_score, ei, ci))

# Sort by score descending, then assign greedily
pairs.sort(key=lambda x: (-x[0], x[1]))
for score, max_score, ei, ci in pairs:
    if ei in matched or ci in matched_csv:
        continue
    matched[ei] = ci
    matched_csv.add(ci)

print(f"Matched: {len(matched)}/{len(excel_rows)}")

# Analyze match quality
exact_cusip = 0
close_cusip = 0
for ei, ci in matched.items():
    correct = excel_rows[ei]['correct_cusip']
    csv_c = csv_rows[ci]['cusip']
    if correct == csv_c:
        exact_cusip += 1
    elif correct[:6] == csv_c[:6]:
        close_cusip += 1

print(f"  Exact CUSIP matches: {exact_cusip}")
print(f"  First-6-match: {close_cusip}")

# Show score distribution
from collections import Counter
score_dist = Counter()
for score, max_score, ei, ci in pairs:
    if ei in matched and matched[ei] == ci:
        pct = int(score / max_score * 100)
        score_dist[pct] += 1

print(f"\nMatch quality distribution:")
for pct in sorted(score_dist.keys(), reverse=True):
    print(f"  {pct}%+: {score_dist[pct]}")

# Show unmatched rows
unmatched = [i for i in range(len(excel_rows)) if i not in matched]
print(f"\nUnmatched Excel rows: {len(unmatched)}")
if unmatched:
    for ei in unmatched[:20]:
        e = excel_rows[ei]
        print(f"  [{ei}] BB_ID: {e['bb_id']}, Correct CUSIP: {e['correct_cusip']}, Issuer: {e['issuer'][:40]}")

# Verify issuer alignment on matches
issuer_match_count = 0
issuer_mismatch_count = 0
for ei, ci in matched.items():
    e_issuer = excel_rows[ei]['issuer'].lower()[:12]
    c_issuer = csv_rows[ci]['issuer'].lower()[:12]
    if e_issuer == c_issuer:
        issuer_match_count += 1
    else:
        issuer_mismatch_count += 1
        if issuer_mismatch_count <= 10:
            print(f"  Issuer mismatch: Excel '{excel_rows[ei]['issuer'][:40]}' vs CSV '{csv_rows[ci]['issuer'][:30]}' "
                  f"(correct={excel_rows[ei]['correct_cusip']} csv={csv_rows[ci]['cusip']})")

print(f"\nIssuer validation: {issuer_match_count} match, {issuer_mismatch_count} mismatch")
