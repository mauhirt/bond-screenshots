#!/usr/bin/env python3
"""
Build correct matching using combined CUSIP + Issuer name scoring.

Strategy:
1. Compute correct CUSIP from BB_ID for each Excel row
2. Score each (Excel, CSV) pair using:
   - OCR-aware CUSIP similarity (0-27 points)
   - Issuer name prefix match (0-15 points, weighted x2 = 0-30 points)
3. Greedy assignment: highest total scores first
4. Validate via issuer name consistency
"""
import csv
import re
import openpyxl
from collections import defaultdict

OCR_PAIRS = set()
_pairs = [
    ('0','X'),('0','O'),('0','D'),
    ('1','I'),('1','L'),('1','7'),
    ('3','2'),('3','8'),
    ('4','8'),('4','A'),
    ('5','S'),('5','F'),('5','6'),
    ('6','G'),('6','8'),
    ('7','T'),
    ('8','B'),
    ('9','G'),('9','Q'),
    ('D','K'),('D','H'),
    ('E','F'),
    ('K','H'),('K','X'),
    ('M','N'),('M','W'),
    ('P','R'),
    ('U','V'),('W','V'),
]
for a, b in _pairs:
    OCR_PAIRS.add((a, b))
    OCR_PAIRS.add((b, a))

def cusip_check_digit(base8):
    values = []
    for ch in base8.upper():
        if ch.isdigit():
            values.append(int(ch))
        elif ch.isalpha():
            values.append(ord(ch) - ord('A') + 10)
        else:
            values.append(0)
    total = 0
    for i, v in enumerate(values):
        if i % 2 == 1:
            v *= 2
        total += v // 10 + v % 10
    return str((10 - (total % 10)) % 10)

def cusip_ocr_score(correct9, csv_cusip):
    """OCR-aware CUSIP comparison. Returns score out of 27 max."""
    best = 0

    def score_pair(a, b):
        s = 0
        for x, y in zip(a, b):
            if x == y:
                s += 3
            elif x.upper() == y.upper():
                s += 2
            elif (x, y) in OCR_PAIRS:
                s += 1
        return s

    clen = len(csv_cusip)
    if clen == 9:
        best = score_pair(correct9, csv_cusip)
    elif clen == 8:
        # Compare first 8 chars
        best = score_pair(correct9[:8], csv_cusip)
    elif clen == 10:
        # Try removing each char
        for skip in range(10):
            trimmed = csv_cusip[:skip] + csv_cusip[skip+1:]
            s = score_pair(correct9, trimmed)
            best = max(best, s)
    elif clen == 7:
        best = score_pair(correct9[:7], csv_cusip)
    else:
        ml = min(9, clen)
        best = score_pair(correct9[:ml], csv_cusip[:ml])

    return best

def issuer_prefix_match(excel_issuer, csv_issuer):
    """Return number of matching prefix characters (case-insensitive)."""
    e = excel_issuer.lower().strip()
    c = csv_issuer.lower().strip()
    # Strip trailing dots from CSV
    c = c.rstrip('.')
    match = 0
    for a, b in zip(e, c):
        if a == b:
            match += 1
        else:
            break
    return match

# Load Excel
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    amt = ws_orig.cell(row=row, column=5).value
    formula = ws_orig.cell(row=row, column=2).value or ''
    m = re.search(r'"([^"]+)\s+Muni"', str(formula))
    bb_id = m.group(1) if m else ''
    correct_cusip = ''
    if bb_id and len(bb_id) >= 8:
        correct_cusip = bb_id[:8] + cusip_check_digit(bb_id[:8])
    excel_rows.append({
        'issuer': issuer, 'amt': amt, 'bb_id': bb_id,
        'correct_cusip': correct_cusip, 'idx': len(excel_rows)
    })

# Load CSV
csv_rows = []
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    for row_data in reader:
        if len(row_data) >= 10:
            cusip = row_data[0].strip()
            if cusip.startswith('TH '):
                cusip = cusip[3:]
            if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                if not all(f.strip() == '--' for f in row_data[:5]):
                    csv_rows.append({
                        'cusip': cusip,
                        'issuer': row_data[2].strip(),
                        'fields': row_data,
                        'csv_idx': len(csv_rows)
                    })

print(f"Excel rows: {len(excel_rows)}")
print(f"CSV rows: {len(csv_rows)}")

# Pre-compute issuer prefix groups for efficiency
# Group CSV rows by first 6 chars of issuer (case-insensitive)
csv_by_issuer = defaultdict(list)
for ci, c in enumerate(csv_rows):
    key = c['issuer'].lower().strip()[:6]
    csv_by_issuer[key].append(ci)

# Also group by first 4 chars of CUSIP for faster lookup
csv_by_cusip_prefix = defaultdict(list)
for ci, c in enumerate(csv_rows):
    key = c['cusip'][:4]
    csv_by_cusip_prefix[key].append(ci)

print("Building candidate pairs...")

# For each Excel row, find candidate CSV rows efficiently
all_pairs = []
for ei, e in enumerate(excel_rows):
    if not e['correct_cusip']:
        continue

    # Get candidates:
    # 1. Same issuer prefix (first 6 chars)
    issuer_key = e['issuer'].lower().strip()[:6]
    candidates = set()
    for key, indices in csv_by_issuer.items():
        if key[:4] == issuer_key[:4]:  # looser prefix for grouping
            candidates.update(indices)

    # 2. Same CUSIP prefix (first 4 chars, considering OCR)
    cusip4 = e['correct_cusip'][:4]
    for key, indices in csv_by_cusip_prefix.items():
        # Check if key is OCR-compatible with cusip4
        compat = sum(1 for a, b in zip(cusip4, key) if a == b or (a, b) in OCR_PAIRS) >= 2
        if compat:
            candidates.update(indices)

    # Score all candidates
    for ci in candidates:
        c = csv_rows[ci]
        cusip_score = cusip_ocr_score(e['correct_cusip'], c['cusip'])
        issuer_score = issuer_prefix_match(e['issuer'], c['issuer'])

        # Combined score: CUSIP (0-27) + Issuer prefix (0-15, weighted x2 = 0-30)
        total = cusip_score + min(issuer_score, 15) * 2

        # Minimum thresholds
        if total >= 25 and (cusip_score >= 12 or issuer_score >= 8):
            all_pairs.append((total, cusip_score, issuer_score, ei, ci))

print(f"Candidate pairs: {len(all_pairs)}")

# Sort by total score descending
all_pairs.sort(key=lambda x: (-x[0], -x[1]))

# Greedy assignment
matched = {}
matched_csv = set()

for total, cusip_sc, issuer_sc, ei, ci in all_pairs:
    if ei in matched or ci in matched_csv:
        continue
    matched[ei] = ci
    matched_csv.add(ci)

print(f"\nMatched: {len(matched)}/{len(excel_rows)}")

# Validate matches
issuer_ok = 0
issuer_bad = 0
bad_examples = []
for ei, ci in matched.items():
    e_iss = excel_rows[ei]['issuer'].lower()[:10]
    c_iss = csv_rows[ci]['issuer'].lower().rstrip('.')[:10]
    if e_iss == c_iss:
        issuer_ok += 1
    else:
        # Check if first 6 chars match
        if e_iss[:6] == c_iss[:6]:
            issuer_ok += 1
        else:
            issuer_bad += 1
            if len(bad_examples) < 20:
                bad_examples.append((ei, ci, excel_rows[ei]['issuer'][:40], csv_rows[ci]['issuer'][:30],
                                    excel_rows[ei]['correct_cusip'], csv_rows[ci]['cusip']))

print(f"Issuer validation: {issuer_ok} match, {issuer_bad} mismatch")
for ei, ci, e_iss, c_iss, corr, csv_c in bad_examples:
    print(f"  [{ei}] '{e_iss}' vs '{c_iss}' (correct={corr} csv={csv_c})")

# Show unmatched stats
unmatched = [i for i in range(len(excel_rows)) if i not in matched]
print(f"\nUnmatched Excel rows: {len(unmatched)}")
# Group unmatched by issuer to see patterns
iss_groups = defaultdict(int)
for ei in unmatched:
    iss_groups[excel_rows[ei]['issuer'][:40]] += 1
for iss, cnt in sorted(iss_groups.items(), key=lambda x: -x[1])[:20]:
    print(f"  [{cnt:3d}] {iss}")
