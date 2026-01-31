#!/usr/bin/env python3
"""
Populate green bonds Excel v3:
- Correct CUSIPs computed from BB_IDs (BB_ID + check digit)
- Multi-phase matching: CUSIP OCR similarity + Issuer name
- CSV parsing with BICS-anchor approach
- All normalizations applied inline

Key fix: BB_IDs ARE CUSIP bases. The original v2 script compared
BB_IDs against CSV CUSIPs using the wrong scoring, resulting in
~90% incorrect row matches. This v3 uses correct CUSIPs derived
from BB_IDs and matches them to OCR'd CSV CUSIPs.
"""
import csv
import re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# ── Constants ──

EXCEL_HEADERS = [
    'CUSIP', 'State Code', 'Issuer Name', 'Yield at Issue', 'Amt Issued',
    'Issue Date', 'Maturity', 'Tax Prov', 'Fin Typ', 'BICS Level 2',
    'Self-reported Green', 'Mgmt of Proc', 'ESG Reporting',
    'ESG Assurance Providers', 'Proj Sel Proc', 'ESG Framework',
    'Industry', 'Issuer Type', 'ESG Project Categories', 'Project Subcategory'
]

BICS_PREFIXES = ['Utilit', 'Financ', 'Educat', 'Trans', 'General', 'Health',
                 'Housin', 'Post E', 'Public', 'NA']

FIN_TYP_PATTERNS = [
    'NEW MONEY', 'NEH MONEY', 'NEH MNEY', 'NEH HANEY', 'NEH HANCY',
    'NEH HMEY', 'NEH HONEY', 'NEW HONEY', 'REH MONEY', 'REW MONEY',
    'REF MONEY', 'NEH MONEY..', 'REFUNDING', 'REFINANCING', 'REFINANC',
    'REFINANC.', 'REFINANC..', 'REFIN&MNG', 'REFINMNG', 'REFINCING',
    'REFINANG', 'REFINIDG', 'REFUNDING..', 'REV',
]

ESG_PROJECT_KEYWORDS = [
    'Sustainable', 'Energy', 'Clean', 'Pollution', 'Green', 'Biodiversity',
    'Climate', 'Renewable', 'LEED', 'Conservation', 'Biogas',
    'Infrastructure', 'Public', 'Natural', 'Water',
]

# Extended OCR pairs (including newly observed errors)
OCR_PAIRS = set()
_pairs = [
    ('0','X'),('0','O'),('0','D'),('0','Q'),
    ('1','I'),('1','L'),('1','7'),('1','J'),
    ('3','2'),('3','8'),('3','5'),
    ('4','8'),('4','A'),('4','9'),
    ('5','S'),('5','F'),('5','6'),('5','3'),
    ('6','G'),('6','8'),('6','A'),('6','5'),
    ('7','T'),('7','1'),
    ('8','B'),('8','6'),('8','3'),
    ('9','G'),('9','Q'),('9','4'),
    ('A','4'),('A','6'),
    ('D','K'),('D','H'),('D','0'),
    ('E','F'),('E','C'),
    ('F','5'),('F','P'),
    ('G','6'),('G','9'),('G','C'),
    ('H','K'),('H','D'),('H','N'),
    ('J','1'),('J','U'),
    ('K','X'),('K','H'),('K','D'),
    ('L','1'),('L','I'),
    ('M','N'),('M','W'),('M','H'),
    ('N','M'),('N','H'),
    ('O','0'),('O','Q'),('O','D'),
    ('P','R'),('P','F'),
    ('Q','9'),('Q','0'),('Q','G'),
    ('R','P'),
    ('S','5'),
    ('T','7'),
    ('U','V'),('U','J'),
    ('V','U'),('V','W'),
    ('W','M'),('W','V'),('W','A'),
    ('X','K'),('X','0'),
    ('Y','V'),
    ('Z','2'),
]
for a, b in _pairs:
    OCR_PAIRS.add((a, b))
    OCR_PAIRS.add((b, a))


# ── CUSIP Functions ──

def cusip_check_digit(base8):
    """Compute CUSIP check digit (Luhn mod-10 variant)."""
    values = []
    for ch in base8.upper():
        if ch.isdigit():
            values.append(int(ch))
        elif ch.isalpha():
            values.append(ord(ch) - ord('A') + 10)
        elif ch == '*':
            values.append(36)
        elif ch == '@':
            values.append(37)
        elif ch == '#':
            values.append(38)
        else:
            values.append(0)
    total = 0
    for i, v in enumerate(values):
        if i % 2 == 1:
            v *= 2
        total += v // 10 + v % 10
    return str((10 - (total % 10)) % 10)


def cusip_ocr_score(correct9, csv_cusip):
    """OCR-aware comparison between correct CUSIP and CSV CUSIP."""
    def score_pair(a, b):
        s = 0
        for x, y in zip(a, b):
            if x == y:
                s += 3
            elif x.upper() == y.upper():
                s += 2
            elif (x, y) in OCR_PAIRS:
                s += 1
        return s

    best = 0
    clen = len(csv_cusip)

    if clen == 9:
        best = score_pair(correct9, csv_cusip)
    elif clen == 8:
        best = score_pair(correct9[:8], csv_cusip)
    elif clen == 10:
        for skip in range(10):
            trimmed = csv_cusip[:skip] + csv_cusip[skip+1:]
            best = max(best, score_pair(correct9, trimmed))
    elif clen == 7:
        best = score_pair(correct9[:7], csv_cusip)
    elif clen in (11, 12):
        for skip in range(clen):
            for skip2 in range(skip+1, clen):
                trimmed = csv_cusip[:skip] + csv_cusip[skip+1:skip2] + csv_cusip[skip2+1:]
                if len(trimmed) == 9:
                    best = max(best, score_pair(correct9, trimmed))
    else:
        ml = min(9, clen)
        best = score_pair(correct9[:ml], csv_cusip[:ml])

    return best


def issuer_prefix_score(excel_issuer, csv_issuer):
    """Score issuer name prefix similarity."""
    e = excel_issuer.lower().strip()
    c = csv_issuer.lower().strip().rstrip('.')
    match = 0
    for a, b in zip(e, c):
        if a == b:
            match += 1
        else:
            break
    return match


def fuzzy_issuer_score(excel_issuer, csv_issuer):
    """Fuzzy issuer match: ratio of matching chars in first 20."""
    e = excel_issuer.lower().strip()[:20]
    c = csv_issuer.lower().strip().rstrip('.')[:20]
    if not e or not c:
        return 0
    matches = sum(1 for a, b in zip(e, c) if a == b)
    return matches / max(len(e), len(c))


# ── CSV Parsing (same BICS-anchor approach) ──

def clean_cusip_field(val):
    s = val.strip()
    if s.startswith('TH '):
        s = s[3:].strip()
    return s

def starts_with_bics(val):
    clean = val.strip().rstrip('.')
    return any(clean.startswith(p) for p in BICS_PREFIXES)

def looks_like_fin_typ(val):
    v = val.strip().upper().rstrip('.')
    for pat in FIN_TYP_PATTERNS:
        if v.startswith(pat.upper().rstrip('.')):
            return True
    return v == '--'

def is_yes_no_dash(val):
    return val.strip() in ('Yes', 'No', '--')

def looks_like_esg_project(val):
    return any(kw.lower() in val.lower() for kw in ESG_PROJECT_KEYWORDS)

def split_bics_merge(bics_raw):
    val = bics_raw.strip()
    m = re.match(r'^(.+?)\s+--\s+(Yes|No)$', val)
    if m: return m.group(1).strip(), m.group(2)
    m = re.match(r'^(.+?)\s+-\s+(Yes|No)$', val)
    if m: return m.group(1).strip(), m.group(2)
    m = re.match(r'^(.+?\.\.+)(Yes|No)$', val)
    if m: return m.group(1).strip(), m.group(2)
    m = re.match(r'^(Utilit\w*|Financ\w*|Educat\w*|Trans\w*|General\s*\w*|Health\w*|Housin\w*|Public\w*)(Yes|No)$', val)
    if m: return m.group(1).strip(), m.group(2)
    return val, None

def extract_fin_typ_from_merged(val):
    v = val.strip()
    for pat in sorted(FIN_TYP_PATTERNS, key=len, reverse=True):
        idx = v.upper().find(pat.upper())
        if idx > 0:
            before = v[:idx].rstrip(' -')
            after = v[idx:]
            if any(kw in before.upper() for kw in ['TAX', 'FED', 'AMT', 'EXEMPT', 'EXMPT', 'ST']):
                return before, after
    return val, None

def parse_csv_row(fields):
    """Parse CSV row into 20-column structure."""
    if len(fields) < 10:
        return None

    cusip = clean_cusip_field(fields[0])
    base = [cusip] + list(fields[1:7])
    rest = list(fields[7:])

    # Find BICS anchor
    bics_idx = None
    for i, f in enumerate(rest):
        if starts_with_bics(f):
            bics_idx = i
            break
    if bics_idx is None:
        for i, f in enumerate(rest):
            bv, _ = split_bics_merge(f)
            if starts_with_bics(bv):
                bics_idx = i
                break

    if bics_idx is None:
        if len(rest) >= 9:
            tax_prov = rest[0]
            fin_typ = rest[1]
            bics = rest[2]
            yesno = []
            idx = 3
            while idx < len(rest) and len(yesno) < 6:
                if is_yes_no_dash(rest[idx]):
                    yesno.append(rest[idx].strip())
                    idx += 1
                else:
                    break
            while len(yesno) < 6:
                yesno.append('--')
            remaining = rest[idx:]
            industry = remaining[0] if remaining else '--'
            esg_project = remaining[1] if len(remaining) > 1 else '--'
            project_subcat = remaining[2] if len(remaining) > 2 else '--'
            result = base + [tax_prov, fin_typ, bics] + yesno[:6] + [
                industry, '--', esg_project, project_subcat]
            return {i: result[i] for i in range(20)}
        return None

    pre_bics = rest[:bics_idx]
    bics_raw = rest[bics_idx]
    bics_value, extra_yesno = split_bics_merge(bics_raw)
    post_bics = rest[bics_idx + 1:]

    if len(pre_bics) == 0:
        tax_prov, fin_typ = '--', '--'
    elif len(pre_bics) == 1:
        if looks_like_fin_typ(pre_bics[0]):
            tax_prov, fin_typ = '--', pre_bics[0]
        else:
            tax_prov, fin_typ = pre_bics[0], '--'
    elif len(pre_bics) == 2:
        tax_prov, fin_typ = pre_bics[0], pre_bics[1]
        if not looks_like_fin_typ(fin_typ):
            mt, ef = extract_fin_typ_from_merged(fin_typ)
            if ef:
                tax_prov = tax_prov + ' ' + mt
                fin_typ = ef
            else:
                tax_prov = pre_bics[0] + ' ' + pre_bics[1]
                fin_typ = '--'
    else:
        if looks_like_fin_typ(pre_bics[-1]):
            tax_prov = ' '.join(pre_bics[:-1])
            fin_typ = pre_bics[-1]
        else:
            mt, ef = extract_fin_typ_from_merged(pre_bics[-1])
            if ef:
                tax_prov = ' '.join(pre_bics[:-1]) + ' ' + mt
                fin_typ = ef
            else:
                tax_prov = ' '.join(pre_bics)
                fin_typ = '--'

    yesno = []
    if extra_yesno:
        yesno.append(extra_yesno)
    remaining = []
    collecting = True
    for f in post_bics:
        if collecting and is_yes_no_dash(f) and len(yesno) < 6:
            yesno.append(f.strip())
        else:
            collecting = False
            remaining.append(f)
    while len(yesno) < 6:
        yesno.append('--')

    industry = remaining[0] if remaining else '--'
    issuer_type = '--'
    esg_project = '--'
    project_subcat = '--'
    if len(remaining) >= 4:
        if looks_like_esg_project(remaining[1]):
            esg_project = remaining[1]
            project_subcat = remaining[2]
        elif remaining[1] == '--' or (remaining[1].isupper() and len(remaining[1]) <= 12):
            issuer_type = remaining[1]
            esg_project = remaining[2]
            project_subcat = remaining[3]
        else:
            esg_project = remaining[1]
            project_subcat = remaining[2]
    elif len(remaining) == 3:
        if remaining[1] == '--' or (remaining[1].isupper() and not looks_like_esg_project(remaining[1])):
            issuer_type = remaining[1]
            esg_project = remaining[2]
        else:
            esg_project = remaining[1]
            project_subcat = remaining[2]
    elif len(remaining) == 2:
        esg_project = remaining[1]

    result = base + [tax_prov, fin_typ, bics_value] + yesno[:6] + [
        industry, issuer_type, esg_project, project_subcat]
    return {i: result[i] for i in range(min(20, len(result)))}


# ── Data Cleaning ──

def clean_text(val):
    if not val or val.strip() == '--':
        return None
    return val.strip()

def clean_yield(val):
    if not val or val.strip() == '--':
        return None
    s = val.strip().rstrip('.')
    s = re.sub(r'%$', '', s)
    try:
        return float(s)
    except ValueError:
        return val

def clean_date(val):
    if not val or val.strip() == '--':
        return None
    s = val.strip()
    for fmt in ['%m/%d/%Y', '%m/%d/%y']:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None  # Return None instead of string for unparseable dates

def clean_yes_no(val):
    if not val or val.strip() == '--':
        return None
    v = val.strip()
    return v if v in ('Yes', 'No') else None


# ── Main ──

def main():
    orig_path = '/home/user/bond-screenshots/green bonds excel.xlsx'
    csv_path = '/home/user/bond-screenshots/claude_table_output_2025_new.csv'
    output_path = '/home/user/bond-screenshots/green_bonds_2025_final.xlsx'

    # Load original Excel
    print("Loading original Excel...")
    orig_wb = load_workbook(orig_path)
    orig_ws = orig_wb.active

    excel_rows = []
    for row in range(2, orig_ws.max_row + 1):
        issuer = (orig_ws.cell(row=row, column=3).value or '').strip()
        amt = orig_ws.cell(row=row, column=5).value
        formula = orig_ws.cell(row=row, column=2).value or ''
        m = re.search(r'"([^"]+)\s+Muni"', str(formula))
        bb_id = m.group(1) if m else ''
        correct_cusip = ''
        if bb_id and len(bb_id) >= 8:
            correct_cusip = bb_id[:8] + cusip_check_digit(bb_id[:8])
        elif bb_id:
            correct_cusip = bb_id  # Short BB_ID, use as-is

        excel_rows.append({
            'idx': row - 2, 'row': row, 'bb_id': bb_id,
            'correct_cusip': correct_cusip,
            'issuer': issuer, 'amt': amt
        })

    print(f"Original Excel rows: {len(excel_rows)}")

    # Load and parse CSV
    print("Loading CSV...")
    csv_all = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader, 1):
            if len(row) >= 10:
                cusip = row[0].strip()
                if cusip.startswith('TH '):
                    cusip = cusip[3:]
                if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                    if not all(f.strip() == '--' for f in row[:5]):
                        parsed = parse_csv_row(row)
                        csv_all.append({
                            'idx': len(csv_all), 'line': i, 'cusip': cusip,
                            'issuer': row[2].strip(), 'fields': row,
                            'parsed': parsed
                        })

    print(f"CSV data rows: {len(csv_all)}")

    # ── Multi-phase matching (correct CUSIP → CSV CUSIP) ──

    matched = {}  # excel_idx -> csv_idx
    matched_csv = set()

    # Build lookup indices
    csv_by_issuer = defaultdict(list)
    for ci, c in enumerate(csv_all):
        key = c['issuer'].lower().strip()[:6]
        csv_by_issuer[key].append(ci)

    csv_by_cusip4 = defaultdict(list)
    for ci, c in enumerate(csv_all):
        csv_by_cusip4[c['cusip'][:4]].append(ci)

    # Phase 1: Strict CUSIP + Issuer matching
    print("Phase 1: Strict matching...")
    all_pairs = []
    for ei, e in enumerate(excel_rows):
        if not e['correct_cusip'] or len(e['correct_cusip']) < 8:
            continue

        candidates = set()
        # By issuer prefix
        issuer_key = e['issuer'].lower().strip()[:6]
        for key, indices in csv_by_issuer.items():
            if key[:4] == issuer_key[:4]:
                candidates.update(indices)
        # By CUSIP prefix (with OCR tolerance)
        cusip4 = e['correct_cusip'][:4]
        for key, indices in csv_by_cusip4.items():
            compat = sum(1 for a, b in zip(cusip4, key)
                        if a == b or (a, b) in OCR_PAIRS) >= 2
            if compat:
                candidates.update(indices)

        for ci in candidates:
            c = csv_all[ci]
            cusip_sc = cusip_ocr_score(e['correct_cusip'], c['cusip'])
            issuer_sc = issuer_prefix_score(e['issuer'], c['issuer'])
            total = cusip_sc + min(issuer_sc, 15) * 2
            if total >= 25 and (cusip_sc >= 12 or issuer_sc >= 8):
                all_pairs.append((total, cusip_sc, issuer_sc, ei, ci))

    all_pairs.sort(key=lambda x: (-x[0], -x[1]))
    for total, cusip_sc, issuer_sc, ei, ci in all_pairs:
        if ei in matched or ci in matched_csv:
            continue
        matched[ei] = ci
        matched_csv.add(ci)

    print(f"  Phase 1: {len(matched)} matched")

    # Phase 2: Broader search for remaining
    print("Phase 2: Broader matching...")
    unmatched_ei = [i for i in range(len(excel_rows)) if i not in matched]
    phase2_pairs = []
    for ei in unmatched_ei:
        e = excel_rows[ei]
        if not e['correct_cusip'] or len(e['correct_cusip']) < 8:
            continue
        for ci, c in enumerate(csv_all):
            if ci in matched_csv:
                continue
            issuer_sc = fuzzy_issuer_score(e['issuer'], c['issuer'])
            if issuer_sc < 0.35:
                continue
            cusip_sc = cusip_ocr_score(e['correct_cusip'], c['cusip'])
            prefix_sc = issuer_prefix_score(e['issuer'], c['issuer'])
            total = cusip_sc + prefix_sc * 2
            if total >= 15:
                phase2_pairs.append((total, cusip_sc, prefix_sc, issuer_sc, ei, ci))

    phase2_pairs.sort(key=lambda x: (-x[0], -x[3]))
    new2 = 0
    for total, cusip_sc, prefix_sc, issuer_sc, ei, ci in phase2_pairs:
        if ei in matched or ci in matched_csv:
            continue
        matched[ei] = ci
        matched_csv.add(ci)
        new2 += 1

    print(f"  Phase 2: {new2} new, Total: {len(matched)}")

    # Phase 3: Very broad search
    print("Phase 3: Broad search...")
    unmatched_ei2 = [i for i in range(len(excel_rows)) if i not in matched]
    phase3_pairs = []
    for ei in unmatched_ei2:
        e = excel_rows[ei]
        if not e['correct_cusip']:
            continue
        best_score = 0
        best_ci = -1
        best_issuer = 0
        for ci, c in enumerate(csv_all):
            if ci in matched_csv:
                continue
            cusip_sc = cusip_ocr_score(e['correct_cusip'], c['cusip'])
            issuer_sc = fuzzy_issuer_score(e['issuer'], c['issuer'])
            total = cusip_sc + issuer_sc * 30
            if total > best_score:
                best_score = total
                best_ci = ci
                best_issuer = issuer_sc
        if best_ci >= 0 and best_issuer >= 0.3:
            phase3_pairs.append((best_score, best_issuer, ei, best_ci))

    phase3_pairs.sort(key=lambda x: -x[0])
    new3 = 0
    for score, issuer_sc, ei, ci in phase3_pairs:
        if ei in matched or ci in matched_csv:
            continue
        matched[ei] = ci
        matched_csv.add(ci)
        new3 += 1

    print(f"  Phase 3: {new3} new, Total: {len(matched)}")

    unmatched_count = len(excel_rows) - len(matched)
    print(f"Unmatched Excel rows: {unmatched_count}")

    # ── Build Output Excel ──
    print("Building output Excel...")
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Municipals'

    # Headers
    header_font = Font(bold=True)
    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = out_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    date_format = 'MM/DD/YYYY'
    matched_rows = 0
    unmatched_rows = 0

    for out_row, e in enumerate(excel_rows, 2):
        ei = e['idx']
        correct_cusip = e['correct_cusip']
        excel_issuer = e['issuer']
        excel_amt = e['amt']

        # CUSIP: ALWAYS use correct CUSIP from BB_ID
        out_ws.cell(row=out_row, column=1, value=correct_cusip)
        # Issuer: ALWAYS from original Excel
        out_ws.cell(row=out_row, column=3, value=excel_issuer)
        # Amt: ALWAYS from original Excel
        out_ws.cell(row=out_row, column=5, value=excel_amt)

        if ei in matched:
            ci = matched[ei]
            c = csv_all[ci]
            p = c['parsed']
            matched_rows += 1

            if p:
                # Col 2: State from CSV
                out_ws.cell(row=out_row, column=2, value=clean_text(p.get(1, '')))
                # Col 4: Yield from CSV
                yield_val = clean_yield(p.get(3, ''))
                cell = out_ws.cell(row=out_row, column=4, value=yield_val)
                if isinstance(yield_val, float):
                    cell.number_format = '0.000'
                # Col 6: Issue Date
                dt = clean_date(p.get(5, ''))
                cell = out_ws.cell(row=out_row, column=6, value=dt)
                if isinstance(dt, datetime):
                    cell.number_format = date_format
                # Col 7: Maturity
                mat = clean_date(p.get(6, ''))
                cell = out_ws.cell(row=out_row, column=7, value=mat)
                if isinstance(mat, datetime):
                    cell.number_format = date_format
                # Col 8: Tax Prov
                out_ws.cell(row=out_row, column=8, value=clean_text(p.get(7, '')))
                # Col 9: Fin Typ
                out_ws.cell(row=out_row, column=9, value=clean_text(p.get(8, '')))
                # Col 10: BICS Level 2
                out_ws.cell(row=out_row, column=10, value=clean_text(p.get(9, '')))
                # Col 11-16: Yes/No
                for j in range(6):
                    out_ws.cell(row=out_row, column=11+j, value=clean_yes_no(p.get(10+j, '')))
                # Col 17: Industry
                out_ws.cell(row=out_row, column=17, value=clean_text(p.get(16, '')))
                # Col 18: Issuer Type
                out_ws.cell(row=out_row, column=18, value=None)
                # Col 19: ESG Project Categories
                out_ws.cell(row=out_row, column=19, value=clean_text(p.get(18, '')))
                # Col 20: Project Subcategory
                out_ws.cell(row=out_row, column=20, value=clean_text(p.get(19, '')))
            else:
                unmatched_rows += 1
                matched_rows -= 1
        else:
            unmatched_rows += 1

    # Auto-adjust column widths
    for col in range(1, 21):
        max_width = len(EXCEL_HEADERS[col-1])
        for row in range(2, min(50, out_ws.max_row+1)):
            v = out_ws.cell(row=row, column=col).value
            if v:
                max_width = max(max_width, len(str(v)))
        out_ws.column_dimensions[out_ws.cell(row=1, column=col).column_letter].width = min(max_width+2, 40)

    out_wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"Total rows (with header): {out_ws.max_row}")
    print(f"Data rows: {out_ws.max_row - 1}")
    print(f"Matched from CSV: {matched_rows}")
    print(f"Unmatched (CUSIP + Issuer + Amt only): {unmatched_rows}")


if __name__ == '__main__':
    main()
