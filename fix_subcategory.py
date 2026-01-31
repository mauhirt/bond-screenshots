#!/usr/bin/env python3
"""Normalize Project Subcategory (col 20) to valid Bloomberg values.

Valid base subcategory values (from Bloomberg screenshot):
  Bioenergy, BREEAM Certified, Circular Design and Production,
  Circular Value Recovery, Conservation, Energy Star Certified,
  Energy Storage, Geothermal, Green House Gas Control,
  Greenhouse Gas Control, Hydro, Hydrogen, Infrastructure,
  Information Support, LEED Certified, Marine, Multimodal,
  Non Motorized, Plumbing System, Pollution Control, Public,
  Rail (Non Passenger), Smart Grids, Soil Remediation, Solar,
  Sustainable Forestry, Vehicles, Waste Management, WELL Certified,
  Wind, NA

Values can be single or comma-separated combinations.
"""
import openpyxl
import re

VALID_BASE = {
    'Bioenergy', 'BREEAM Certified', 'Circular Design and Production',
    'Circular Value Recovery', 'Conservation', 'Energy Star Certified',
    'Energy Storage', 'Geothermal', 'Green House Gas Control',
    'Greenhouse Gas Control', 'Hydro', 'Hydrogen', 'Infrastructure',
    'Information Support', 'LEED Certified', 'Marine', 'Multimodal',
    'Non Motorized', 'Plumbing System', 'Pollution Control', 'Public',
    'Rail (Non Passenger)', 'Smart Grids', 'Soil Remediation', 'Solar',
    'Sustainable Forestry', 'Vehicles', 'Waste Management',
    'WELL Certified', 'Wind',
}

# Values to clear entirely (leaked Yes/No values, junk)
CLEAR_VALUES = {
    'Yes Yes', 'No', 'STAT.', '#NAME?',
}


def normalize_single(part):
    """Normalize a single subcategory token to its valid form."""
    s = part.strip()
    if not s:
        return None

    # Case-insensitive lookup
    s_upper = s.upper()

    # Direct matches (case-insensitive)
    for valid in VALID_BASE:
        if s_upper == valid.upper():
            return valid

    # OCR fixes
    # LED -> LEED
    if s_upper in ('LED CERTIFIED', 'LEE CERTIFIED', 'LEED CERTIFED'):
        return 'LEED Certified'

    # Renewable Gas Control -> Green House Gas Control (OCR error)
    if 'RENEWABLE GAS' in s_upper or s_upper == 'RENEWABLE GAS CONTROL':
        return 'Green House Gas Control'

    # Green House Gas Control variants
    if 'GREEN HOUSE GAS' in s_upper or 'GREENHOUSE GAS' in s_upper:
        return 'Green House Gas Control'

    # Rail variants
    if 'RAIL' in s_upper and 'PASSENGER' in s_upper:
        return 'Rail (Non Passenger)'
    if s_upper == 'RAIL':
        return 'Rail (Non Passenger)'

    # Truncated values
    if s_upper.startswith('SOLAR (INCL') or s_upper.startswith('SOLAR(INCL'):
        return 'Solar'
    if s_upper.startswith('CIRCULAR DESIGN'):
        return 'Circular Design and Production'
    if s_upper.startswith('CIRCULAR VALUE'):
        return 'Circular Value Recovery'
    if s_upper.startswith('CIRCULAR WASTE') or s_upper.startswith('CIRC'):
        return 'Circular Value Recovery'
    if s_upper.startswith('WASTE REDUCTION') or s_upper.startswith('WASTE MANAGE'):
        return 'Waste Management'
    if s_upper.startswith('NON MOTOR') or s_upper.startswith('NONMOTOR'):
        return 'Non Motorized'
    if s_upper.startswith('ENERGY STAR'):
        return 'Energy Star Certified'
    if s_upper.startswith('ENERGY STOR'):
        return 'Energy Storage'
    if s_upper.startswith('BREEAM'):
        return 'BREEAM Certified'
    if s_upper.startswith('WELL'):
        return 'WELL Certified'
    if s_upper.startswith('SUSTAINAB') and 'FOREST' in s_upper:
        return 'Sustainable Forestry'
    if s_upper.startswith('SUSTAINAB') and 'INFRA' in s_upper:
        return 'Infrastructure'
    if s_upper.startswith('INFO'):
        return 'Information Support'
    if s_upper.startswith('POLLUT'):
        return 'Pollution Control'
    if s_upper.startswith('SMART'):
        return 'Smart Grids'
    if s_upper.startswith('SOIL'):
        return 'Soil Remediation'
    if s_upper.startswith('MULTI'):
        return 'Multimodal'
    if s_upper.startswith('PLUMB'):
        return 'Plumbing System'

    # Sub.and/or Energy Storage -> Energy Storage
    if 'SUB.' in s_upper and 'ENERGY' in s_upper:
        return 'Energy Storage'

    # PUBLIC case normalization
    if s_upper == 'PUBLIC':
        return 'Public'

    return None  # Unknown - will flag for review


def clean_subcategory(raw):
    """Clean and normalize a Project Subcategory value."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.upper() == 'NA':
        return None

    # Clear junk/leaked values
    if s in CLEAR_VALUES:
        return None

    # Split on separators: semicolons, slashes (not inside parens), commas, hyphens between words
    # But first handle compound entries joined by space only (e.g., "Public Rail (Non Passenger)")
    # Strategy: split on ;,/ first, then check for space-joined pairs

    # Normalize separators: replace ; and / with ,
    normalized = s
    # Replace semicolons
    normalized = normalized.replace(';', ',')
    # Replace / but not inside parentheses and not in "and/or"
    normalized = re.sub(r'(?<!\()/(?!\))', ',', normalized)
    # Replace hyphens used as separators (e.g., "Public-Rail") but not inside valid names
    # Only replace hyphens that are between two capitalized words
    normalized = re.sub(r'(?<=[a-z])-(?=[A-Z])', ', ', normalized)
    normalized = re.sub(r'(?<=[A-Z])-(?=[A-Z])', ', ', normalized)

    # Split on commas
    parts = [p.strip() for p in normalized.split(',') if p.strip()]

    # Handle space-joined compounds (e.g., "Public Rail (Non Passenger)" -> "Public", "Rail (Non Passenger)")
    # and "Public LEED Certified" -> "Public", "LEED Certified"
    # and "Public Greenhouse Gas Control" -> "Public", "Greenhouse Gas Control"
    expanded = []
    for part in parts:
        # Check if it's already a valid single value
        matched = normalize_single(part)
        if matched:
            expanded.append(matched)
            continue

        # Try to split space-joined values
        # Known patterns where a space joins two subcategories
        space_split = False

        # "Public X" pattern
        if part.upper().startswith('PUBLIC '):
            rest = part[7:].strip()  # after "Public "
            rest_norm = normalize_single(rest)
            if rest_norm:
                expanded.append('Public')
                expanded.append(rest_norm)
                space_split = True

        # "Solar Wind" without comma -> "Solar", "Wind"
        if not space_split and part.upper().startswith('SOLAR '):
            rest = part[6:].strip()
            # "Solar Wind/Energy Storage" already split by / above
            rest_norm = normalize_single(rest)
            if rest_norm:
                expanded.append('Solar')
                expanded.append(rest_norm)
                space_split = True

        if not space_split:
            # Try the part as-is, even if not matching
            expanded.append(part)

    # Normalize each part
    result_parts = []
    for part in expanded:
        # Check if already normalized
        if part in VALID_BASE:
            result_parts.append(part)
        else:
            norm = normalize_single(part)
            if norm:
                result_parts.append(norm)
            else:
                # Keep as-is but flag
                result_parts.append(f"?{part}")

    if not result_parts:
        return None

    # Deduplicate while preserving order
    seen = set()
    deduped = []
    for p in result_parts:
        if p not in seen:
            seen.add(p)
            deduped.append(p)

    return ', '.join(deduped)


# Load Excel
wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

fixes = 0
cleared = 0
unknown = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=20).value
    if val is None:
        continue
    s = str(val).strip()
    if not s:
        continue

    new_val = clean_subcategory(val)

    if new_val is None and val is not None:
        ws.cell(row=row, column=20).value = None
        cleared += 1
        fixes += 1
    elif new_val is not None and new_val != s:
        if '?' in new_val:
            unknown.append((row, s, new_val))
        ws.cell(row=row, column=20).value = new_val
        fixes += 1

wb.save('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')

print(f"Total fixes applied: {fixes}")
print(f"Cleared (junk/leaked): {cleared}")

if unknown:
    print(f"\nUnknown/flagged values ({len(unknown)}):")
    for row, orig, new in unknown:
        print(f"  Row {row}: '{orig}' -> '{new}'")

# Verify final distribution
from collections import Counter
wb2 = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws2 = wb2.active
counter = Counter()
non_null = 0
for row in range(2, ws2.max_row + 1):
    val = ws2.cell(row=row, column=20).value
    if val is not None and str(val).strip():
        counter[str(val).strip()] += 1
        non_null += 1

print(f"\nFinal distribution ({non_null} non-null values, {len(counter)} unique):")
for val, cnt in counter.most_common():
    marker = ''
    # Check each part is valid
    parts = [p.strip() for p in val.split(',')]
    for p in parts:
        if p.startswith('?') or (p not in VALID_BASE and p not in ('NA',)):
            marker = ' *** INVALID ***'
            break
    print(f"  [{cnt:3d}] '{val}'{marker}")
