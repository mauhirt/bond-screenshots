#!/usr/bin/env python3
"""
Reconcile Claude's formatted output with Gemini's more accurate data extraction.

Strategy (matching the user's manual reconciliation pattern on rows 416-918):
- Match Gemini → Claude rows by CUSIP (1539 direct matches)
- Issue Date & Maturity: Take from Gemini, swap DD/MM when both ≤ 12
  (Gemini extracts DD/MM for ambiguous dates from Bloomberg terminal)
- Yield: Keep Claude's values (user kept Claude's yields in manual recon)
- CUSIP, Issuer Name, Amt Issued: Keep from Claude/original template
- Tax Prov, Fin Typ, BICS, Industry: Take Gemini's raw values, apply
  Claude's normalization. Fall back to Claude if normalization fails.
- Yes/No, ESG, Subcategory: Keep Claude's normalized values
- Unmatched (no Gemini data): Keep existing Claude data, force issue year to 2025
- Post-fix: Force all issue date years to 2025 (all issuance was 2025)
"""
import csv
import re
from datetime import datetime
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, Alignment

# Import normalization functions from apply_all_normalizations.py
import sys
sys.path.insert(0, '/home/user/bond-screenshots')

# We'll inline the needed normalization functions here for clarity
# ── Normalization helpers (from apply_all_normalizations.py) ──

VALID_TAX_PROV = [
    'AMT/ST TAX-EXEMPT', 'AMT/ST TAXABLE', 'FED & ST TAX-EXEMPT',
    'FED AMT FOR INDIVIDUALS', 'FED BQ', 'FED BQ/ST TAX-EXEMPT',
    'FED BQ/ST TAXABLE', 'FED TAX-EXEMPT', 'FED TAX-EXEMPT/ST TAXABLE',
    'FED TAXABLE', 'FED TAXABLE/ST TAX-EXEMPT', 'FED TAXABLE/ST TAXABLE',
]

VALID_BICS = [
    'Education', 'Financing', 'General Government', 'Health Care',
    'Housing', 'NA', 'Post Employment', 'Public Services',
    'Transportation', 'Utilities',
]

VALID_INDUSTRY = [
    'APPROP', 'ARPT', 'BONDBK', 'CCRC', 'CDD', 'CHRT', 'CMNTYC', 'DEV',
    'EDU', 'EDULEASE', 'GARVEE', 'GASFWD', 'GO', 'GODIST', 'GOVLEASE',
    'GOVTGTD', 'HGR', 'HOSP', 'HOTELTAX', 'INCTAX', 'LMFH', 'LNGBDL',
    'LNPOOL', 'MDD', 'MEL', 'MISC', 'MISCTAX', 'MUNUTIL', 'NA', 'NFPCULT',
    'NFPRO', 'PILOT', 'PORTS', 'PUBPWR', 'PUBTRAN', 'PUBWTR', 'SALESTAX',
    'SCD', 'SCO', 'SELFAPP', 'SMFH', 'SOLWST', 'SPLASMT', 'STDHSG', 'TIF',
    'TOLL', 'TRIBES', 'TXMUD', 'WSGTD', 'WTRSWR',
]


def normalize_bics(raw):
    """Normalize BICS value to valid category."""
    if not raw:
        return None
    s = str(raw).strip().rstrip('.')
    if not s or s == '--' or s.upper() == 'NA':
        return None
    for valid in VALID_BICS:
        if s.lower().startswith(valid.lower()[:5]):
            return valid
    return None


def normalize_industry(raw):
    """Normalize Industry value."""
    if not raw:
        return None
    s = str(raw).strip().upper().rstrip('.')
    if not s or s == '--' or s == 'NA':
        return None
    if s in VALID_INDUSTRY:
        return s
    # Known OCR patterns
    mappings = {
        'TRSWR': 'WTRSWR', 'TRSUR': 'WTRSWR', 'GASFM': 'GASFWD',
        'GASFW': 'GASFWD', 'GODIS': 'GODIST', 'WSGTD': 'WSGTD',
        'PUBW': 'PUBWTR', 'PUBT': 'PUBTRAN', 'PUBP': 'PUBPWR',
        'MUNUTIL': 'MUNUTIL', 'SALES': 'SALESTAX', 'MISCT': 'MISCTAX',
        'HOTEL': 'HOTELTAX', 'SELFA': 'SELFAPP', 'SPLAS': 'SPLASMT',
        'STDH': 'STDHSG', 'SOLW': 'SOLWST', 'GARV': 'GARVEE',
        'NFPC': 'NFPCULT', 'NFPR': 'NFPRO', 'BOND': 'BONDBK',
        'EDUL': 'EDULEASE', 'GOVL': 'GOVLEASE', 'LNGB': 'LNGBDL',
        'LNPO': 'LNPOOL', 'TXMU': 'TXMUD', 'TRIB': 'TRIBES',
        'INCT': 'INCTAX',
    }
    for prefix, target in mappings.items():
        if s.startswith(prefix):
            return target
    if s in ('YES', 'NO', 'YES YES'):
        return None
    # Levenshtein fallback
    def levenshtein(s1, s2):
        if len(s1) < len(s2): return levenshtein(s2, s1)
        if len(s2) == 0: return len(s1)
        prev = list(range(len(s2) + 1))
        for i, c1 in enumerate(s1):
            curr = [i + 1]
            for j, c2 in enumerate(s2):
                curr.append(min(prev[j+1]+1, curr[j]+1, prev[j]+(c1!=c2)))
            prev = curr
        return prev[len(s2)]
    best_dist, best_match = 999, None
    for v in VALID_INDUSTRY:
        d = levenshtein(s, v)
        if d < best_dist:
            best_dist, best_match = d, v
    max_dist = 3 if len(s) <= 6 else 4
    return best_match if best_dist <= max_dist else None


def normalize_tax_prov(raw):
    """Normalize Tax Prov value."""
    if not raw:
        return None
    s = str(raw).strip().upper()
    if not s or s == '--' or s == 'NA':
        return None
    for valid in VALID_TAX_PROV:
        if s == valid:
            return valid
    # Pattern matching
    if re.match(r'^FED\s*[&R]\s*ST\s*TAX[\s-]*(EX|CB)', s) or re.match(r'^FE[OD]\s*[&R]\s*S[TL]\s*TAX[\s-]*EX', s):
        return 'FED & ST TAX-EXEMPT'
    if re.match(r'^FED\s*TAX[\s-]*(EX[EI]?MPT|CBMPT|EXEHPT)$', s) or re.match(r'^FE[OD]\s*TAX[\s-]*EX', s):
        return 'FED TAX-EXEMPT'
    if 'TAX-EXEMPT/ST TAXAB' in s or re.match(r'^FED\s*TAX[\s-]*EX[A-Z]*/\s*ST\s*TAX(ABLE|ARLE)', s):
        return 'FED TAX-EXEMPT/ST TAXABLE'
    if re.match(r'^FED\s*TAXAB', s) and '/ST' not in s:
        return 'FED TAXABLE'
    if 'TAXABLE/ST TAX-EX' in s:
        return 'FED TAXABLE/ST TAX-EXEMPT'
    if re.match(r'^FED\s*B[QUO]$', s):
        return 'FED BQ'
    if re.match(r'^FED?\s*B[QUOV].*/?.*ST.*TAX[\s-]*EX', s):
        return 'FED BQ/ST TAX-EXEMPT'
    if re.match(r'^FED\s*B[QUO].*/?.*ST.*TAXAB', s):
        return 'FED BQ/ST TAXABLE'
    if 'AMT FOR' in s or re.match(r'^FED\s*AMT', s):
        return 'FED AMT FOR INDIVIDUALS'
    if re.match(r'^AMT/?ST\s*TAX[\s-]*EX', s):
        return 'AMT/ST TAX-EXEMPT'
    # Broader patterns
    if 'FED' in s and 'BQ' in s and 'TAX-EX' in s: return 'FED BQ/ST TAX-EXEMPT'
    if 'FED' in s and 'BQ' in s: return 'FED BQ'
    if 'FED' in s and 'ST TAX-EX' in s: return 'FED & ST TAX-EXEMPT'
    if 'FED' in s and 'TAX-EX' in s: return 'FED TAX-EXEMPT'
    if 'FED' in s and 'TAXAB' in s: return 'FED TAXABLE'
    if 'AMT' in s and 'TAX-EX' in s: return 'AMT/ST TAX-EXEMPT'
    return None


def normalize_fin_typ(raw):
    """Normalize Fin Typ value."""
    if not raw:
        return None
    s = str(raw).strip().upper()
    if s in ('NEW MONEY',): return 'NEW MONEY'
    if s in ('REFUNDING',): return 'REFUNDING'
    if any(k in s for k in ('NEW', 'NEH', 'NEU', 'NEN')):
        if any(k in s for k in ('MONEY', 'HONEY', 'MNEY', 'HMEY', 'HANEY', 'HANCY')):
            return 'NEW MONEY'
    if any(k in s for k in ('REF', 'REH', 'REW')):
        if any(k in s for k in ('FUND', 'FINANC', 'FINAN', 'FINMNG', 'FINCING')):
            return 'REFUNDING'
    return None


# ── Date handling ──

def parse_gemini_date(date_str, swap_ambiguous=True):
    """Parse Gemini date, swapping DD/MM to MM/DD for ambiguous dates.

    Gemini extracts dates from Bloomberg terminal which uses DD/MM format.
    When day > 12, Gemini auto-corrects to MM/DD (only valid interpretation).
    When both <= 12, Gemini keeps DD/MM order -> we need to swap.
    """
    if not date_str or not date_str.strip():
        return None
    s = date_str.strip()

    # Parse MM/DD/YYYY format
    parts = s.split('/')
    if len(parts) != 3:
        return None

    try:
        p1, p2, year = int(parts[0]), int(parts[1]), int(parts[2])
    except ValueError:
        return None

    if year < 100:
        year += 2000

    if swap_ambiguous and p1 <= 12 and p2 <= 12:
        # Ambiguous date: swap DD/MM -> MM/DD
        month, day = p2, p1
    else:
        month, day = p1, p2

    try:
        return datetime(year, month, day)
    except ValueError:
        # If swap produced invalid date, try original
        try:
            return datetime(year, p1, p2)
        except ValueError:
            return None


# ── Main Reconciliation ──

def main():
    claude_path = '/home/user/bond-screenshots/green_bonds_2025_final.xlsx'
    gemini_path = '/home/user/bond-screenshots/Green_Bonds_2025_gemini.csv'
    output_path = '/home/user/bond-screenshots/green_bonds_2025_final.xlsx'

    # Load Claude output
    print("Loading Claude output...")
    claude_wb = openpyxl.load_workbook(claude_path)
    claude_ws = claude_wb.active

    # Load Gemini CSV
    print("Loading Gemini CSV...")
    with open(gemini_path, 'r') as f:
        gemini_all = list(csv.reader(f))

    gemini_header = gemini_all[0]
    print(f"Gemini: {len(gemini_all)-1} data rows, columns: {len(gemini_header)}")
    # Gemini cols: 0=Row, 1=CUSIP, 2=Issuer, 3=Yield, 4=Amt, 5=Issue, 6=Mat,
    #              7=TaxProv, 8=FinTyp, 9=BICS, 10-15=Yes/No, 16=Industry,
    #              17=KestrelScore, 18=ESG, 19=Subcat, 20=IssuerType

    # Build Gemini lookup by CUSIP (first occurrence for duplicates)
    gemini_by_cusip = {}
    gemini_dupes = 0
    for r in gemini_all[1:]:
        cusip = r[1].strip()
        if cusip and cusip not in gemini_by_cusip:
            gemini_by_cusip[cusip] = r
        elif cusip:
            gemini_dupes += 1

    print(f"Gemini unique CUSIPs: {len(gemini_by_cusip)} (+ {gemini_dupes} duplicates)")

    # Reconcile
    stats = {
        'matched': 0, 'unmatched': 0,
        'issue_updated': 0, 'issue_swapped': 0, 'issue_kept': 0,
        'mat_updated': 0, 'mat_swapped': 0, 'mat_kept': 0,
        'tax_updated': 0, 'bics_updated': 0, 'industry_updated': 0,
        'fin_typ_updated': 0, 'state_updated': 0,
        'yesno_updated': 0, 'issuer_type_updated': 0,
        'esg_cat_updated': 0, 'subcat_updated': 0,
    }

    date_format = 'MM/DD/YYYY'

    for row in range(2, claude_ws.max_row + 1):
        cusip = claude_ws.cell(row=row, column=1).value
        if not cusip:
            continue

        gemini_row = gemini_by_cusip.get(cusip)
        if not gemini_row:
            stats['unmatched'] += 1
            continue

        stats['matched'] += 1

        # ── Issue Date: Take from Gemini with DD/MM swap ──
        g_issue_str = gemini_row[5].strip() if len(gemini_row) > 5 else ''
        if g_issue_str:
            new_issue = parse_gemini_date(g_issue_str, swap_ambiguous=True)
            if new_issue:
                old_issue = claude_ws.cell(row=row, column=6).value
                claude_ws.cell(row=row, column=6).value = new_issue
                claude_ws.cell(row=row, column=6).number_format = date_format
                stats['issue_updated'] += 1
                # Check if swap was applied
                parts = g_issue_str.split('/')
                if len(parts) == 3:
                    try:
                        p1, p2 = int(parts[0]), int(parts[1])
                        if p1 <= 12 and p2 <= 12:
                            stats['issue_swapped'] += 1
                    except:
                        pass
            else:
                stats['issue_kept'] += 1
        else:
            stats['issue_kept'] += 1

        # ── Maturity: Take from Gemini with DD/MM swap ──
        g_mat_str = gemini_row[6].strip() if len(gemini_row) > 6 else ''
        if g_mat_str:
            new_mat = parse_gemini_date(g_mat_str, swap_ambiguous=True)
            if new_mat:
                claude_ws.cell(row=row, column=7).value = new_mat
                claude_ws.cell(row=row, column=7).number_format = date_format
                stats['mat_updated'] += 1
                parts = g_mat_str.split('/')
                if len(parts) == 3:
                    try:
                        p1, p2 = int(parts[0]), int(parts[1])
                        if p1 <= 12 and p2 <= 12:
                            stats['mat_swapped'] += 1
                    except:
                        pass
            else:
                stats['mat_kept'] += 1
        else:
            stats['mat_kept'] += 1

        # ── Yield: Keep Claude's (user kept Claude's yields in manual recon) ──
        # No change needed

        # ── State: Extract from Gemini issuer prefix (e.g., "AR City of...") ──
        g_issuer = gemini_row[2].strip() if len(gemini_row) > 2 else ''
        if len(g_issuer) >= 2 and g_issuer[2:3] == ' ':
            g_state = g_issuer[:2].upper()
            if g_state.isalpha() and len(g_state) == 2:
                old_state = claude_ws.cell(row=row, column=2).value
                if old_state != g_state:
                    claude_ws.cell(row=row, column=2).value = g_state
                    stats['state_updated'] += 1

        # ── Tax Prov: Try Gemini's value, normalize it ──
        g_tax = gemini_row[7].strip() if len(gemini_row) > 7 else ''
        if g_tax and g_tax != '--':
            normalized = normalize_tax_prov(g_tax)
            if normalized:
                old_tax = claude_ws.cell(row=row, column=8).value
                if old_tax != normalized:
                    claude_ws.cell(row=row, column=8).value = normalized
                    stats['tax_updated'] += 1

        # ── Fin Typ: Try Gemini's value ──
        g_fin = gemini_row[8].strip() if len(gemini_row) > 8 else ''
        if g_fin and g_fin != '--':
            normalized = normalize_fin_typ(g_fin)
            if normalized:
                old_fin = claude_ws.cell(row=row, column=9).value
                if old_fin != normalized:
                    claude_ws.cell(row=row, column=9).value = normalized
                    stats['fin_typ_updated'] += 1

        # ── BICS: Try Gemini's value ──
        g_bics = gemini_row[9].strip() if len(gemini_row) > 9 else ''
        if g_bics and g_bics != '--':
            normalized = normalize_bics(g_bics)
            if normalized:
                old_bics = claude_ws.cell(row=row, column=10).value
                if old_bics != normalized:
                    claude_ws.cell(row=row, column=10).value = normalized
                    stats['bics_updated'] += 1

        # ── Industry: Try Gemini's value ──
        g_ind = gemini_row[16].strip() if len(gemini_row) > 16 else ''
        if g_ind and g_ind != '--':
            normalized = normalize_industry(g_ind)
            if normalized:
                old_ind = claude_ws.cell(row=row, column=17).value
                if old_ind != normalized:
                    claude_ws.cell(row=row, column=17).value = normalized
                    stats['industry_updated'] += 1

        # ── Yes/No columns: Use Gemini where strictly "Yes" or "No" ──
        # Gemini cols 10-15 map to Excel cols 11-16
        # Gemini has column-shift contamination (industry codes, ESG cats in
        # Yes/No fields), so only accept strict "Yes"/"No" values.
        yesno_mapping = [
            (10, 11),  # Self-reported Green
            (11, 12),  # Mgmt of Proc
            (12, 13),  # ESG Reporting
            (13, 14),  # ESG Assurance Providers
            (14, 15),  # Proj Sel Proc
            (15, 16),  # ESG Framework
        ]
        for g_col, e_col in yesno_mapping:
            gv = gemini_row[g_col].strip() if len(gemini_row) > g_col else ''
            if gv in ('Yes', 'No'):
                old_val = str(claude_ws.cell(row=row, column=e_col).value or '').strip()
                if old_val != gv:
                    claude_ws.cell(row=row, column=e_col).value = gv
                    stats['yesno_updated'] += 1

        # ── Issuer Type: Extract from Gemini issuer name prefix ──
        # Gemini issuer format: "XX Entity of Name..." e.g., "AR City of Little Rock"
        if len(g_issuer) >= 3 and g_issuer[2:3] == ' ':
            rest = g_issuer[3:]
            issuer_type = ''
            if rest.startswith('City & County') or rest.startswith('City and County'):
                issuer_type = 'County'
            elif rest.startswith('City'):
                issuer_type = 'City'
            elif rest.startswith('County'):
                issuer_type = 'County'
            elif rest.startswith('State'):
                issuer_type = 'State'
            elif rest.startswith('Town'):
                issuer_type = 'Town'
            elif rest.startswith('Village'):
                issuer_type = 'Village'
            elif rest.startswith('District'):
                issuer_type = 'District'
            if issuer_type:
                old_it = str(claude_ws.cell(row=row, column=18).value or '').strip()
                if old_it != issuer_type:
                    claude_ws.cell(row=row, column=18).value = issuer_type
                    stats['issuer_type_updated'] += 1

        # ── ESG Project Categories: Gemini col 17 (labeled Kestrel Score) ──
        # Gemini's column 17 actually contains ESG Project Categories data.
        # Needs cleanup: strip "CITY...", "STATE" prefixes, fix "|" → ", "
        g_esg_raw = gemini_row[17].strip() if len(gemini_row) > 17 else ''
        if g_esg_raw and g_esg_raw != '--':
            # Clean up prefixes and separators
            esg_clean = g_esg_raw
            # Remove entity prefixes like "CITY... ", "STATE "
            esg_clean = re.sub(r'^(CITY|STATE|COUNTY|TOWN)\S*\s+', '', esg_clean)
            # Fix separator: "|" → ", "
            esg_clean = esg_clean.replace('|', ', ')
            # Fix truncation: "Ren..." → "Renewable Energy", etc.
            esg_clean = re.sub(r'Ren\w*\.{2,}', 'Renewable Energy', esg_clean)
            esg_clean = re.sub(r'Clean Trans\w*\.{2,}', 'Clean Transportation', esg_clean)
            esg_clean = re.sub(r'Sust\w*\.{2,}', 'Sustainable Water', esg_clean)
            esg_clean = re.sub(r'Pollu\w*\.{2,}', 'Pollution Control', esg_clean)
            esg_clean = re.sub(r'Energ\w*Eff\w*\.{2,}', 'Energy Efficiency', esg_clean)
            esg_clean = esg_clean.strip().rstrip('.')
            # Only use if it contains valid ESG category keywords
            valid_esg_kw = ['Sustainable Water', 'Energy Efficiency', 'Clean Transportation',
                           'Renewable Energy', 'Pollution Control']
            if any(kw in esg_clean for kw in valid_esg_kw):
                old_esg = str(claude_ws.cell(row=row, column=19).value or '').strip()
                if old_esg != esg_clean:
                    claude_ws.cell(row=row, column=19).value = esg_clean
                    stats['esg_cat_updated'] += 1

        # ── Project Subcategory: Combine Gemini col 18 + col 19 ──
        # Manual reconciliation shows subcategory = col18 + ", " + col19
        g_sub1 = gemini_row[18].strip() if len(gemini_row) > 18 else ''
        g_sub2 = gemini_row[19].strip() if len(gemini_row) > 19 else ''
        # Filter out contaminated values (industry codes, ESG categories)
        valid_subcats = ['Infrastructure', 'Energy Storage', 'Public', 'Solar',
                        'Wind', 'Rail (Non Passenger)', 'Conservation',
                        'LEED Certified', 'Greenhouse Gas Control',
                        'Bioer', 'Waste Reduction']
        sub_parts = []
        for sv in [g_sub1, g_sub2]:
            if sv and any(vc in sv for vc in valid_subcats):
                # Also handle combined values with "|"
                for piece in sv.replace('|', ', ').split(', '):
                    piece = piece.strip().rstrip('.')
                    # Fix truncation
                    piece = re.sub(r'Infrastructur\b', 'Infrastructure', piece)
                    if piece and any(vc in piece for vc in valid_subcats):
                        if piece not in sub_parts:
                            sub_parts.append(piece)
        if sub_parts:
            subcat_val = ', '.join(sub_parts)
            old_sub = str(claude_ws.cell(row=row, column=20).value or '').strip()
            if old_sub != subcat_val:
                claude_ws.cell(row=row, column=20).value = subcat_val
                stats['subcat_updated'] += 1

    # ── Post-processing: Force all issue dates to year 2025 ──
    # All issuance in this dataset was for 2025. Claude's OCR often got the
    # year wrong (2020, 2006, etc.). For unmatched CUSIPs (no Gemini data)
    # and matched CUSIPs where Gemini had garbage dates, force year to 2025.
    year_fixes = 0
    for row in range(2, claude_ws.max_row + 1):
        issue_dt = claude_ws.cell(row=row, column=6).value
        if isinstance(issue_dt, datetime) and issue_dt.year != 2025:
            try:
                fixed = issue_dt.replace(year=2025)
                claude_ws.cell(row=row, column=6).value = fixed
                claude_ws.cell(row=row, column=6).number_format = date_format
                year_fixes += 1
            except ValueError:
                pass  # Feb 29 edge case

    # Save
    claude_wb.save(output_path)

    print(f"\n=== Reconciliation Results ===")
    print(f"Matched by CUSIP: {stats['matched']}")
    print(f"Unmatched (kept Claude data): {stats['unmatched']}")
    print(f"\nIssue Date: {stats['issue_updated']} updated ({stats['issue_swapped']} had DD/MM swap)")
    print(f"Maturity: {stats['mat_updated']} updated ({stats['mat_swapped']} had DD/MM swap)")
    print(f"Issue date year forced to 2025: {year_fixes}")
    print(f"\nState: {stats['state_updated']} updated from Gemini")
    print(f"Tax Prov: {stats['tax_updated']} updated from Gemini")
    print(f"Fin Typ: {stats['fin_typ_updated']} updated from Gemini")
    print(f"BICS: {stats['bics_updated']} updated from Gemini")
    print(f"Industry: {stats['industry_updated']} updated from Gemini")
    print(f"Yes/No fields: {stats['yesno_updated']} updated from Gemini")
    print(f"Issuer Type: {stats['issuer_type_updated']} updated from Gemini")
    print(f"ESG Project Categories: {stats['esg_cat_updated']} updated from Gemini")
    print(f"Project Subcategory: {stats['subcat_updated']} updated from Gemini")
    print(f"\nSaved: {output_path}")

    # Verify issue date years
    issue_years = Counter()
    for row in range(2, claude_ws.max_row + 1):
        dt = claude_ws.cell(row=row, column=6).value
        if isinstance(dt, datetime):
            issue_years[dt.year] += 1
    print(f"\nIssue date year distribution: {dict(sorted(issue_years.items()))}")


if __name__ == '__main__':
    main()
