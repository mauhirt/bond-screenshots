#!/usr/bin/env python3
"""Test if positional matching after dedup aligns CSV rows with Excel rows."""
import csv
import re
import openpyxl

# Load original Excel issuer names and BB_IDs
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    amt = ws_orig.cell(row=row, column=5).value
    formula = ws_orig.cell(row=row, column=2).value or ''
    m = re.search(r'"([^"]+)\s+Muni"', str(formula))
    bb_id = m.group(1) if m else ''
    excel_rows.append({'issuer': issuer, 'amt': amt, 'bb_id': bb_id})

print(f"Excel rows: {len(excel_rows)}")

# Load CSV and deduplicate
csv_rows = []
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    for row_data in reader:
        if len(row_data) >= 10:
            cusip = row_data[0].strip()
            if cusip.startswith('TH '):
                cusip = cusip[3:]
            if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                if not all(f.strip() == '--' for f in row_data[:5]):
                    csv_rows.append({
                        'cusip': cusip,
                        'issuer': row_data[2].strip(),
                        'amt': row_data[4].strip() if len(row_data) > 4 else '',
                        'fields': row_data
                    })

print(f"Total CSV rows (before dedup): {len(csv_rows)}")

# Deduplicate: remove rows with same CUSIP that appear consecutively (screenshot boundary dupes)
deduped = [csv_rows[0]]
for i in range(1, len(csv_rows)):
    if csv_rows[i]['cusip'] != csv_rows[i-1]['cusip']:
        deduped.append(csv_rows[i])
    # else: skip duplicate

print(f"CSV rows after dedup: {len(deduped)}")

# Now check positional alignment
print(f"\nPositional comparison (first 40 rows):")
print(f"{'Pos':>4} {'Excel Issuer':>40} {'CSV Issuer':>30} {'Match?':>8} {'Excel BB_ID':>12} {'CSV CUSIP':>12}")
matches = 0
mismatches = 0
for i in range(min(40, len(excel_rows), len(deduped))):
    e_issuer = excel_rows[i]['issuer'][:40]
    c_issuer = deduped[i]['issuer'][:30]

    # Check if issuers match (first 10 chars, case insensitive)
    e_prefix = e_issuer.lower()[:15]
    c_prefix = c_issuer.lower()[:15]
    match = 'YES' if e_prefix == c_prefix else 'NO'
    if match == 'YES':
        matches += 1
    else:
        mismatches += 1

    print(f"{i+1:4d} {e_issuer:>40} {c_issuer:>30} {match:>8} {excel_rows[i]['bb_id']:>12} {deduped[i]['cusip']:>12}")

# Overall stats
print(f"\n\nFull positional alignment check:")
total_match = 0
total_mismatch = 0
first_mismatch = None
for i in range(min(len(excel_rows), len(deduped))):
    e_prefix = excel_rows[i]['issuer'].lower()[:15]
    c_prefix = deduped[i]['issuer'].lower()[:15]
    if e_prefix == c_prefix:
        total_match += 1
    else:
        total_mismatch += 1
        if first_mismatch is None:
            first_mismatch = i

print(f"Matching issuers: {total_match}")
print(f"Mismatching issuers: {total_mismatch}")
if first_mismatch is not None:
    print(f"First mismatch at position: {first_mismatch}")
    i = first_mismatch
    print(f"  Excel: '{excel_rows[i]['issuer'][:50]}' (BB_ID: {excel_rows[i]['bb_id']})")
    print(f"  CSV:   '{deduped[i]['issuer'][:50]}' (CUSIP: {deduped[i]['cusip']})")
    # Show surrounding rows
    for j in range(max(0, i-2), min(len(deduped), i+5)):
        e = excel_rows[j]['issuer'][:40] if j < len(excel_rows) else 'N/A'
        c = deduped[j]['issuer'][:40] if j < len(deduped) else 'N/A'
        marker = ' <<<' if j == i else ''
        print(f"  [{j}] Excel: {e:>40} | CSV: {c:>40}{marker}")
