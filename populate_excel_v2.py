#!/usr/bin/env python3
"""
Populate green bonds Excel using:
- Original Excel as skeleton (1825 rows, correct Issuer Names + Amt Issued)
- CSV extraction for remaining columns (CUSIP, State, Yield, dates, etc.)
- Multi-phase fuzzy matching (BB_ID → CUSIP OCR mapping + issuer name)
- Deduplication: exactly 1825 output rows
"""

import csv
import re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

# ── Constants ──

EXCEL_HEADERS = [
    'CUSIP', 'State Code', 'Issuer Name', 'Yield at Issue', 'Amt Issued',
    'Issue Date', 'Maturity', 'Tax Prov', 'Fin Typ', 'BICS Level 2',
    'Self-reported Green', 'Mgmt of Proc', 'ESG Reporting',
    'ESG Assurance Providers', 'Proj Sel Proc', 'ESG Framework',
    'Industry', 'Issuer Type', 'ESG Project Categories', 'Project Subcategory'
]

BICS_PREFIXES = ['Utilit', 'Financ', 'Educat', 'Trans', 'General']

FIN_TYP_PATTERNS = [
    'NEW MONEY', 'NEH MONEY', 'NEH MNEY', 'NEH HANEY', 'NEH HANCY',
    'NEH HMEY', 'NEH HONEY', 'NEW HONEY', 'REH MONEY', 'REW MONEY',
    'REF MONEY', 'NEH MONEY..', 'REFUNDING', 'REFINANCING', 'REFINANC',
    'REFINANC.', 'REFINANC..', 'REFIN&MNG', 'REFINMNG', 'REFINCING',
    'REFINANG', 'REFINIDG', 'REFUNDING..', 'REV',
]

ESG_PROJECT_KEYWORDS = [
    'Sustainable', 'Energy', 'Clean', 'Pollution', 'Green', 'Biodiversity',
    'Climate', 'Renewable', 'LEED', 'Conservation', 'Biogas',
    'Infrastructure', 'Public', 'Natural', 'Water',
]

OCR_PAIRS = {
    ('0','X'),('X','0'),('0','O'),('O','0'),('0','D'),('D','0'),
    ('1','I'),('I','1'),('1','L'),('L','1'),('1','7'),('7','1'),
    ('3','2'),('2','3'),('3','8'),('8','3'),
    ('4','8'),('8','4'),('4','A'),('A','4'),
    ('5','S'),('S','5'),('5','F'),('F','5'),('5','6'),('6','5'),
    ('6','G'),('G','6'),('6','8'),('8','6'),
    ('7','T'),('T','7'),
    ('8','B'),('B','8'),
    ('9','G'),('G','9'),('9','Q'),('Q','9'),
    ('D','K'),('K','D'),('D','H'),('H','D'),
    ('E','F'),('F','E'),
    ('K','H'),('H','K'),('K','X'),('X','K'),
    ('M','N'),('N','M'),('M','W'),('W','M'),
    ('P','R'),('R','P'),
    ('U','V'),('V','U'),('W','V'),('V','W'),
}

# ── Matching functions ──

def bb_to_cusip_score(bb_id, cusip):
    """OCR-aware fuzzy score between Bloomberg ID and extracted CUSIP."""
    bb8 = bb_id[:8].ljust(8)
    c8 = cusip[:8].ljust(8) if len(cusip) >= 8 else cusip.ljust(8)
    score = 0
    for a, b in zip(bb8, c8):
        if a == b:
            score += 3
        elif a.upper() == b.upper():
            score += 2
        elif (a, b) in OCR_PAIRS:
            score += 1
    return score


def issuer_prefix_score(excel_issuer, csv_issuer):
    """Score issuer name prefix similarity."""
    e = excel_issuer.lower().strip()[:15]
    c = csv_issuer.lower().strip()[:15]
    if not e or not c:
        return 0
    match = 0
    for a, b in zip(e, c):
        if a == b:
            match += 1
        else:
            break
    return match


# ── CSV row parsing (same logic as populate_excel.py) ──

def clean_cusip(val):
    s = val.strip()
    if s.startswith('TH '):
        s = s[3:].strip()
    return s

def starts_with_bics(val):
    clean = val.strip().rstrip('.')
    return any(clean.startswith(p) for p in BICS_PREFIXES)

def looks_like_fin_typ(val):
    v = val.strip().upper().rstrip('.')
    for pat in FIN_TYP_PATTERNS:
        if v.startswith(pat.upper().rstrip('.')):
            return True
    if v == '--':
        return True
    return False

def is_yes_no_dash(val):
    return val.strip() in ('Yes', 'No', '--')

def looks_like_esg_project(val):
    return any(kw.lower() in val.lower() for kw in ESG_PROJECT_KEYWORDS)

def split_bics_merge(bics_raw):
    val = bics_raw.strip()
    m = re.match(r'^(.+?)\s+--\s+(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)
    m = re.match(r'^(.+?)\s+-\s+(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)
    m = re.match(r'^(.+?\.\.+)(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)
    m = re.match(r'^(Utilit\w*|Financ\w*|Educat\w*|Trans\w*|General\s*\w*)(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)
    return val, None

def extract_fin_typ_from_merged(val):
    v = val.strip()
    for pat in sorted(FIN_TYP_PATTERNS, key=len, reverse=True):
        idx = v.upper().find(pat.upper())
        if idx > 0:
            before = v[:idx].rstrip(' -')
            after = v[idx:]
            if any(kw in before.upper() for kw in ['TAX', 'FED', 'AMT', 'EXEMPT', 'EXMPT', 'ST']):
                return before, after
    return val, None

def parse_csv_row(fields):
    """Parse CSV row into 20-column structure. Returns dict or None."""
    if len(fields) < 10:
        return None

    cusip = clean_cusip(fields[0])
    base = [cusip] + list(fields[1:7])
    rest = list(fields[7:])

    # Find BICS anchor
    bics_idx = None
    for i, f in enumerate(rest):
        if starts_with_bics(f):
            bics_idx = i
            break
    if bics_idx is None:
        for i, f in enumerate(rest):
            bv, _ = split_bics_merge(f)
            if starts_with_bics(bv):
                bics_idx = i
                break

    if bics_idx is None:
        # Positional fallback
        if len(rest) >= 9:
            tax_prov = rest[0]
            fin_typ = rest[1]
            bics = rest[2]
            yesno_start = 3
            yesno = []
            idx = yesno_start
            while idx < len(rest) and len(yesno) < 6:
                if is_yes_no_dash(rest[idx]):
                    yesno.append(rest[idx].strip())
                    idx += 1
                else:
                    break
            while len(yesno) < 6:
                yesno.append('--')
            remaining = rest[idx:]
            industry = remaining[0] if remaining else '--'
            esg_project = remaining[1] if len(remaining) > 1 else '--'
            project_subcat = remaining[2] if len(remaining) > 2 else '--'
            result = base + [tax_prov, fin_typ, bics] + yesno[:6] + [
                industry, '--', esg_project, project_subcat]
            return {i: result[i] for i in range(20)}
        return None

    pre_bics = rest[:bics_idx]
    bics_raw = rest[bics_idx]
    bics_value, extra_yesno = split_bics_merge(bics_raw)
    post_bics = rest[bics_idx + 1:]

    # Extract Tax Prov and Fin Typ
    if len(pre_bics) == 0:
        tax_prov, fin_typ = '--', '--'
    elif len(pre_bics) == 1:
        if looks_like_fin_typ(pre_bics[0]):
            tax_prov, fin_typ = '--', pre_bics[0]
        else:
            tax_prov, fin_typ = pre_bics[0], '--'
    elif len(pre_bics) == 2:
        tax_prov, fin_typ = pre_bics[0], pre_bics[1]
        if not looks_like_fin_typ(fin_typ):
            mt, ef = extract_fin_typ_from_merged(fin_typ)
            if ef:
                tax_prov = tax_prov + ' ' + mt
                fin_typ = ef
            else:
                tax_prov = pre_bics[0] + ' ' + pre_bics[1]
                fin_typ = '--'
    else:
        if looks_like_fin_typ(pre_bics[-1]):
            tax_prov = ' '.join(pre_bics[:-1])
            fin_typ = pre_bics[-1]
        else:
            mt, ef = extract_fin_typ_from_merged(pre_bics[-1])
            if ef:
                tax_prov = ' '.join(pre_bics[:-1]) + ' ' + mt
                fin_typ = ef
            else:
                tax_prov = ' '.join(pre_bics)
                fin_typ = '--'

    # Yes/No block
    yesno = []
    if extra_yesno:
        yesno.append(extra_yesno)
    remaining = []
    collecting = True
    for f in post_bics:
        if collecting and is_yes_no_dash(f) and len(yesno) < 6:
            yesno.append(f.strip())
        else:
            collecting = False
            remaining.append(f)
    while len(yesno) < 6:
        yesno.append('--')

    # Post-block fields
    industry = remaining[0] if remaining else '--'
    issuer_type = '--'
    esg_project = '--'
    project_subcat = '--'
    if len(remaining) >= 4:
        if looks_like_esg_project(remaining[1]):
            esg_project = remaining[1]
            project_subcat = remaining[2]
        elif remaining[1] == '--' or (remaining[1].isupper() and len(remaining[1]) <= 12):
            issuer_type = remaining[1]
            esg_project = remaining[2]
            project_subcat = remaining[3]
        else:
            esg_project = remaining[1]
            project_subcat = remaining[2]
    elif len(remaining) == 3:
        if remaining[1] == '--' or (remaining[1].isupper() and not looks_like_esg_project(remaining[1])):
            issuer_type = remaining[1]
            esg_project = remaining[2]
        else:
            esg_project = remaining[1]
            project_subcat = remaining[2]
    elif len(remaining) == 2:
        if looks_like_esg_project(remaining[1]):
            esg_project = remaining[1]
        else:
            esg_project = remaining[1]

    result = base + [tax_prov, fin_typ, bics_value] + yesno[:6] + [
        industry, issuer_type, esg_project, project_subcat]
    return {i: result[i] for i in range(min(20, len(result)))}


# ── Data cleaning ──

def clean_text(val):
    if not val or val.strip() == '--':
        return None
    return val.strip()

def clean_yield(val):
    if not val or val.strip() == '--':
        return None
    s = val.strip().rstrip('.')
    s = re.sub(r'%$', '', s)
    try:
        return float(s)
    except ValueError:
        return val

def clean_date(val):
    if not val or val.strip() == '--':
        return None
    s = val.strip()
    for fmt in ['%m/%d/%Y', '%m/%d/%y']:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return val

def clean_yes_no(val):
    if not val or val.strip() == '--':
        return None
    v = val.strip()
    return v if v in ('Yes', 'No') else None


# ── Main ──

def main():
    orig_path = '/home/user/bond-screenshots/green bonds excel.xlsx'
    csv_path = '/home/user/bond-screenshots/claude_table_output_2025_new.csv'
    output_path = '/home/user/bond-screenshots/green_bonds_2025_final.xlsx'

    # Load original Excel (ground truth)
    print("Loading original Excel...")
    orig_wb = load_workbook(orig_path)
    orig_ws = orig_wb.active

    excel_rows = []
    for row in range(2, orig_ws.max_row + 1):
        issuer = (orig_ws.cell(row=row, column=3).value or '').strip()
        amt = orig_ws.cell(row=row, column=5).value
        formula = orig_ws.cell(row=row, column=2).value or ''
        m = re.search(r'"([^"]+)\s+Muni"', str(formula))
        bb_id = m.group(1) if m else ''
        excel_rows.append({'idx': row-2, 'row': row, 'bb_id': bb_id,
                          'issuer': issuer, 'amt': amt})

    print(f"Original Excel rows: {len(excel_rows)}")

    # Load and parse CSV
    print("Loading CSV...")
    csv_all = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader, 1):
            if len(row) >= 10:
                cusip = row[0].strip()
                if cusip.startswith('TH '):
                    cusip = cusip[3:]
                if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                    if not all(f.strip() == '--' for f in row[:5]):
                        parsed = parse_csv_row(row)
                        csv_all.append({
                            'idx': len(csv_all), 'line': i, 'cusip': cusip,
                            'issuer': row[2].strip(), 'fields': row, 'parsed': parsed
                        })

    print(f"CSV data rows: {len(csv_all)}")

    # ── Multi-phase matching ──
    matched = {}  # excel_idx -> csv_idx
    matched_csv = set()

    # Phase 1: BB_ID → CUSIP (OCR-aware, high confidence)
    print("Phase 1: BB_ID matching...")
    pairs = []
    for ei, e in enumerate(excel_rows):
        bb = e['bb_id']
        if not bb:
            continue
        for ci, c in enumerate(csv_all):
            score = bb_to_cusip_score(bb, c['cusip'])
            if score >= 10:
                pairs.append((score, ei, ci))

    pairs.sort(key=lambda x: -x[0])
    for score, ei, ci in pairs:
        if ei in matched or ci in matched_csv:
            continue
        matched[ei] = ci
        matched_csv.add(ci)
    print(f"  Matched: {len(matched)}")

    # Phase 2: Issuer name + BB_ID for remaining
    print("Phase 2: Issuer + BB_ID matching...")
    csv_by_issuer = defaultdict(list)
    for c in csv_all:
        if c['idx'] not in matched_csv:
            key = c['issuer'].lower()[:10]
            csv_by_issuer[key].append(c)

    new_matches = 0
    for e in excel_rows:
        if e['idx'] in matched:
            continue
        e_key = e['issuer'].lower()[:10]
        best_score = 0
        best_ci = None
        for csv_key, candidates in csv_by_issuer.items():
            issuer_sim = issuer_prefix_score(e_key, csv_key)
            if issuer_sim < 5:
                continue
            for c in candidates:
                if c['idx'] in matched_csv:
                    continue
                bb_score = bb_to_cusip_score(e['bb_id'], c['cusip']) if e['bb_id'] else 0
                total = issuer_sim * 2 + bb_score
                if total > best_score:
                    best_score = total
                    best_ci = c['idx']
        if best_ci is not None and best_score >= 15:
            matched[e['idx']] = best_ci
            matched_csv.add(best_ci)
            new_matches += 1
    print(f"  New matches: {new_matches}, Total: {len(matched)}")

    unmatched_count = len(excel_rows) - len(matched)
    print(f"Unmatched Excel rows (will use BB_ID as CUSIP): {unmatched_count}")

    # ── Build output Excel ──
    print("Building output Excel...")
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = 'Municipals'

    # Headers
    header_font = Font(bold=True)
    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = out_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    date_format = 'MM/DD/YYYY'
    matched_rows = 0
    unmatched_rows = 0

    for out_row, e in enumerate(excel_rows, 2):
        ei = e['idx']

        # Ground truth from original Excel
        excel_issuer = e['issuer']
        excel_amt = e['amt']
        bb_id = e['bb_id']

        if ei in matched:
            # Use CSV data for most columns
            ci = matched[ei]
            c = csv_all[ci]
            p = c['parsed']
            matched_rows += 1

            if p:
                # Col 1: CUSIP from CSV
                out_ws.cell(row=out_row, column=1, value=clean_text(p.get(0, '')))
                # Col 2: State from CSV
                out_ws.cell(row=out_row, column=2, value=clean_text(p.get(1, '')))
                # Col 3: Issuer Name from ORIGINAL EXCEL
                out_ws.cell(row=out_row, column=3, value=excel_issuer)
                # Col 4: Yield from CSV
                yield_val = clean_yield(p.get(3, ''))
                cell = out_ws.cell(row=out_row, column=4, value=yield_val)
                if isinstance(yield_val, float):
                    cell.number_format = '0.000'
                # Col 5: Amt Issued from ORIGINAL EXCEL
                out_ws.cell(row=out_row, column=5, value=excel_amt)
                # Col 6: Issue Date from CSV
                dt = clean_date(p.get(5, ''))
                cell = out_ws.cell(row=out_row, column=6, value=dt)
                if isinstance(dt, datetime):
                    cell.number_format = date_format
                # Col 7: Maturity from CSV
                mat = clean_date(p.get(6, ''))
                cell = out_ws.cell(row=out_row, column=7, value=mat)
                if isinstance(mat, datetime):
                    cell.number_format = date_format
                # Col 8: Tax Prov
                out_ws.cell(row=out_row, column=8, value=clean_text(p.get(7, '')))
                # Col 9: Fin Typ
                out_ws.cell(row=out_row, column=9, value=clean_text(p.get(8, '')))
                # Col 10: BICS Level 2
                out_ws.cell(row=out_row, column=10, value=clean_text(p.get(9, '')))
                # Col 11-16: Yes/No
                for j in range(6):
                    out_ws.cell(row=out_row, column=11+j, value=clean_yes_no(p.get(10+j, '')))
                # Col 17: Industry
                out_ws.cell(row=out_row, column=17, value=clean_text(p.get(16, '')))
                # Col 18: Issuer Type (leave empty)
                out_ws.cell(row=out_row, column=18, value=None)
                # Col 19: ESG Project Categories
                out_ws.cell(row=out_row, column=19, value=clean_text(p.get(18, '')))
                # Col 20: Project Subcategory
                out_ws.cell(row=out_row, column=20, value=clean_text(p.get(19, '')))
            else:
                # Parsed failed - use BB_ID as CUSIP
                out_ws.cell(row=out_row, column=1, value=bb_id)
                out_ws.cell(row=out_row, column=3, value=excel_issuer)
                out_ws.cell(row=out_row, column=5, value=excel_amt)
        else:
            # Unmatched: use BB_ID as CUSIP, Excel data for Issuer/Amt
            unmatched_rows += 1
            out_ws.cell(row=out_row, column=1, value=bb_id)
            out_ws.cell(row=out_row, column=3, value=excel_issuer)
            out_ws.cell(row=out_row, column=5, value=excel_amt)

    # Auto-adjust column widths
    for col in range(1, 21):
        max_width = len(EXCEL_HEADERS[col-1])
        for row in range(2, min(50, out_ws.max_row+1)):
            v = out_ws.cell(row=row, column=col).value
            if v:
                max_width = max(max_width, len(str(v)))
        out_ws.column_dimensions[out_ws.cell(row=1, column=col).column_letter].width = min(max_width+2, 40)

    out_wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"Total rows (with header): {out_ws.max_row}")
    print(f"Data rows: {out_ws.max_row - 1}")
    print(f"Matched from CSV: {matched_rows}")
    print(f"Unmatched (BB_ID only): {unmatched_rows}")


if __name__ == '__main__':
    main()
