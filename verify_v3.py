#!/usr/bin/env python3
"""Verify the v3 output: CUSIPs, state consistency, data completeness."""
import openpyxl
import re
from collections import Counter

wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

print(f"Total rows: {ws.max_row - 1}")

# 1. CUSIP validation
valid_9 = 0
other_len = Counter()
for row in range(2, ws.max_row + 1):
    cusip = str(ws.cell(row=row, column=1).value or '').strip()
    if len(cusip) == 9 and re.match(r'^[A-Z0-9]{9}$', cusip):
        valid_9 += 1
    else:
        other_len[len(cusip)] += 1

print(f"\nCUSIP validation:")
print(f"  Valid 9-char: {valid_9}")
for length, cnt in sorted(other_len.items()):
    print(f"  {length}-char: {cnt}")

# 2. Data completeness
cols = ['CUSIP', 'State Code', 'Issuer Name', 'Yield at Issue', 'Amt Issued',
        'Issue Date', 'Maturity', 'Tax Prov', 'Fin Typ', 'BICS Level 2',
        'Self-reported Green', 'Mgmt of Proc', 'ESG Reporting',
        'ESG Assurance Providers', 'Proj Sel Proc', 'ESG Framework',
        'Industry', 'Issuer Type', 'ESG Project Categories', 'Project Subcategory']

for col_idx in range(1, 21):
    non_null = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=col_idx).value
        if val is not None and str(val).strip():
            non_null += 1
    print(f"  Col {col_idx:2d} ({cols[col_idx-1]:25s}): {non_null:5d}/{ws.max_row-1}")

# 3. State consistency check (using explicit state codes in issuer names)
state_in_name = {'AR': ' AR ', 'CA': ' CA ', 'NY': ' NY ', 'NJ': ' NJ ',
                 'TX': ' TX ', 'OH': ' OH ', 'PA': ' PA ', 'FL': ' FL ',
                 'IL': ' IL ', 'MA': ' MA ', 'CT': ' CT ', 'MI': ' MI '}
consistent = 0
inconsistent = 0
checked = 0
for row in range(2, ws.max_row + 1):
    state = ws.cell(row=row, column=2).value
    issuer = ws.cell(row=row, column=3).value or ''
    if not state:
        continue
    state = str(state).strip()
    for st, pattern in state_in_name.items():
        if pattern in issuer:
            checked += 1
            if state == st:
                consistent += 1
            else:
                inconsistent += 1
                if inconsistent <= 10:
                    print(f"  INCONSISTENT: Row {row}: State='{state}' but issuer has '{st}': {issuer[:50]}")
            break

print(f"\nState consistency (where state in issuer name):")
print(f"  Checked: {checked}, Consistent: {consistent}, Inconsistent: {inconsistent}")

# 4. Sample first 10 rows
print(f"\nSample data (rows 2-6):")
for row in range(2, 7):
    cusip = ws.cell(row=row, column=1).value
    state = ws.cell(row=row, column=2).value
    issuer = str(ws.cell(row=row, column=3).value or '')[:40]
    yield_v = ws.cell(row=row, column=4).value
    print(f"  Row {row}: CUSIP={cusip} State={state} Yield={yield_v} | {issuer}")
