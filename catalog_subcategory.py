#!/usr/bin/env python3
"""Catalog Project Subcategory (col 20) values in the Excel."""
import openpyxl
from collections import Counter

wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

values = []
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=20).value
    values.append(val)

print(f"Total rows: {len(values)}")
print(f"Non-null: {sum(1 for v in values if v is not None and str(v).strip())}")
print(f"Null/empty: {sum(1 for v in values if v is None or not str(v).strip())}")

# Count unique values
counter = Counter()
for v in values:
    if v is not None and str(v).strip():
        counter[str(v).strip()] += 1

print(f"\nUnique values: {len(counter)}")
print(f"\nAll values (sorted by frequency):")
for val, cnt in counter.most_common():
    print(f"  [{cnt:3d}] '{val}'")
