#!/usr/bin/env python3
"""
Assess the quality of the current matching by checking consistency:
1. Does the State code match the issuer's actual state?
2. Are CUSIPs from the CSV matching what the BB_IDs suggest?
3. How many rows have data from a clearly different bond?
"""
import openpyxl
import re
from collections import Counter

wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

# Extract state from issuer name (many muni issuers have state in name)
state_map = {
    'AL': ['Alabama', ' AL '],
    'AK': ['Alaska', ' AK '],
    'AR': ['Arkansas', ' AR '],
    'AZ': ['Arizona', ' AZ '],
    'CA': ['California', ' CA '],
    'CO': ['Colorado', ' CO '],
    'CT': ['Connecticut', ' CT '],
    'DC': ['District of Columbia', ' DC '],
    'DE': ['Delaware', ' DE '],
    'FL': ['Florida', ' FL '],
    'GA': ['Georgia', ' GA '],
    'HI': ['Hawaii', ' HI '],
    'IA': ['Iowa', ' IA '],
    'ID': ['Idaho', ' ID '],
    'IL': ['Illinois', ' IL '],
    'IN': ['Indiana', ' IN '],
    'KS': ['Kansas', ' KS '],
    'KY': ['Kentucky', ' KY '],
    'LA': ['Louisiana', ' LA '],
    'MA': ['Massachusetts', ' MA '],
    'MD': ['Maryland', ' MD '],
    'ME': ['Maine', ' ME '],
    'MI': ['Michigan', ' MI '],
    'MN': ['Minnesota', ' MN '],
    'MO': ['Missouri', ' MO '],
    'MS': ['Mississippi', ' MS '],
    'MT': ['Montana', ' MT '],
    'NC': ['North Carolina', ' NC '],
    'ND': ['North Dakota', ' ND '],
    'NE': ['Nebraska', ' NE '],
    'NH': ['New Hampshire', ' NH '],
    'NJ': ['New Jersey', ' NJ '],
    'NM': ['New Mexico', ' NM '],
    'NV': ['Nevada', ' NV '],
    'NY': ['New York', ' NY '],
    'OH': ['Ohio', ' OH '],
    'OK': ['Oklahoma', ' OK '],
    'OR': ['Oregon', ' OR '],
    'PA': ['Pennsylvania', ' PA '],
    'PR': ['Puerto Rico'],
    'RI': ['Rhode Island', ' RI '],
    'SC': ['South Carolina', ' SC '],
    'SD': ['South Dakota', ' SD '],
    'TN': ['Tennessee', ' TN '],
    'TX': ['Texas', ' TX '],
    'UT': ['Utah', ' UT '],
    'VA': ['Virginia', ' VA '],
    'VT': ['Vermont', ' VT '],
    'WA': ['Washington', ' WA '],
    'WI': ['Wisconsin', ' WI '],
    'WV': ['West Virginia', ' WV '],
    'WY': ['Wyoming', ' WY '],
}

def infer_state(issuer):
    """Try to infer state from issuer name."""
    for state, patterns in state_map.items():
        for pat in patterns:
            if pat in issuer:
                return state
    return None

consistent = 0
inconsistent = 0
no_state = 0
no_data = 0
inconsistent_rows = []

for row in range(2, ws.max_row + 1):
    state_code = ws.cell(row=row, column=2).value
    issuer = ws.cell(row=row, column=3).value or ''

    if not state_code:
        no_data += 1
        continue

    state_code = str(state_code).strip()
    inferred = infer_state(issuer)

    if inferred is None:
        no_state += 1
    elif inferred == state_code:
        consistent += 1
    else:
        inconsistent += 1
        if len(inconsistent_rows) < 30:
            cusip = ws.cell(row=row, column=1).value
            inconsistent_rows.append((row, state_code, inferred, issuer[:50], cusip))

print(f"State consistency check:")
print(f"  Consistent (state matches issuer): {consistent}")
print(f"  Inconsistent (state != issuer): {inconsistent}")
print(f"  Could not infer state from issuer: {no_state}")
print(f"  No state data: {no_data}")

if inconsistent_rows:
    print(f"\nInconsistent rows (first 30):")
    for row, state, inferred, issuer, cusip in inconsistent_rows:
        print(f"  Row {row}: State='{state}' but issuer suggests '{inferred}': {issuer} (CUSIP: {cusip})")

# Also check: how many rows have null CUSIP (unmatched rows from original populate)
null_cusip = 0
bb_id_cusip = 0
csv_cusip = 0
for row in range(2, ws.max_row + 1):
    cusip = ws.cell(row=row, column=1).value
    state = ws.cell(row=row, column=2).value
    if cusip is None or not str(cusip).strip():
        null_cusip += 1
    elif state is None:
        bb_id_cusip += 1  # Has CUSIP but no state = likely BB_ID used as CUSIP
    else:
        csv_cusip += 1

print(f"\nCUSIP source analysis:")
print(f"  From CSV (has state): {csv_cusip}")
print(f"  BB_ID used as CUSIP (no state): {bb_id_cusip}")
print(f"  No CUSIP at all: {null_cusip}")
