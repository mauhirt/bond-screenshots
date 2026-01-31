#!/usr/bin/env python3
"""Analyze CUSIP column (col 4) in the output Excel."""
import openpyxl
from collections import Counter

wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

total = 0
null_count = 0
valid_9 = 0
lengths = Counter()
issues = []

for row in range(2, ws.max_row + 1):
    total += 1
    val = ws.cell(row=row, column=4).value
    if val is None or not str(val).strip():
        null_count += 1
        continue
    s = str(val).strip()
    lengths[len(s)] += 1
    if len(s) == 9:
        valid_9 += 1
    else:
        if len(issues) < 50:
            issuer = ws.cell(row=row, column=3).value or ''
            issues.append((row, s, len(s), str(issuer)[:30]))

print(f"Total rows: {total}")
print(f"Null/empty CUSIPs: {null_count}")
print(f"Valid 9-char CUSIPs: {valid_9}")
print(f"Non-9-char CUSIPs: {total - null_count - valid_9}")

print(f"\nLength distribution:")
for length, cnt in sorted(lengths.items()):
    print(f"  {length} chars: {cnt}")

print(f"\nSample non-9-char CUSIPs (first 50):")
for row, cusip, length, issuer in issues:
    print(f"  Row {row}: '{cusip}' (len={length}) | {issuer}")
