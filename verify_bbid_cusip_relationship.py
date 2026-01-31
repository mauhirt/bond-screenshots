#!/usr/bin/env python3
"""
Verify that Bloomberg BB_IDs are actually CUSIP bases (first 8 chars).

If BB_ID "237190AA" is the CUSIP without check digit, then:
- CSV CUSIP "22719XAA1" should be an OCR'd version of "237190AA" + check digit
- The OCR errors should all be explainable by OCR_PAIRS

This would mean we can compute correct CUSIPs directly from BB_IDs!
"""
import csv
import re
import openpyxl

OCR_PAIRS = {
    ('0','X'),('X','0'),('0','O'),('O','0'),('0','D'),('D','0'),
    ('1','I'),('I','1'),('1','L'),('L','1'),('1','7'),('7','1'),
    ('3','2'),('2','3'),('3','8'),('8','3'),
    ('4','8'),('8','4'),('4','A'),('A','4'),
    ('5','S'),('S','5'),('5','F'),('F','5'),('5','6'),('6','5'),
    ('6','G'),('G','6'),('6','8'),('8','6'),
    ('7','T'),('T','7'),
    ('8','B'),('B','8'),
    ('9','G'),('G','9'),('9','Q'),('Q','9'),
    ('D','K'),('K','D'),('D','H'),('H','D'),
    ('K','H'),('H','K'),('K','X'),('X','K'),
    ('M','N'),('N','M'),('M','W'),('W','M'),
    ('P','R'),('R','P'),
    ('U','V'),('V','U'),('W','V'),('V','W'),
}

def ocr_compatible(a, b):
    """Check if two characters are OCR-compatible (same or OCR pair)."""
    if a == b:
        return True
    if a.upper() == b.upper():
        return True
    return (a, b) in OCR_PAIRS or (b, a) in OCR_PAIRS

# Load Excel BB_IDs
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    formula = ws_orig.cell(row=row, column=2).value or ''
    m = re.search(r'"([^"]+)\s+Muni"', str(formula))
    bb_id = m.group(1) if m else ''
    excel_rows.append({'issuer': issuer, 'bb_id': bb_id})

# Load current output
wb_out = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws_out = wb_out.active

# Compare BB_ID (first 8 chars of real CUSIP) vs output CUSIP (from CSV)
print("BB_ID vs Output CUSIP - OCR compatibility analysis:")
print(f"{'Row':>4} {'BB_ID':>12} {'Out CUSIP':>12} {'Match':>8} {'Diffs':>30}")

compatible = 0
incompatible = 0
no_compare = 0

for i in range(len(excel_rows)):
    bb_id = excel_rows[i]['bb_id']
    row = i + 2
    cusip = str(ws_out.cell(row=row, column=1).value or '').strip()

    if not bb_id or not cusip:
        no_compare += 1
        continue

    # Compare first 8 chars
    bb8 = bb_id[:8]
    c8 = cusip[:8] if len(cusip) >= 8 else cusip.ljust(8)

    diffs = []
    all_compat = True
    for j, (a, b) in enumerate(zip(bb8, c8)):
        if a != b:
            compat = ocr_compatible(a, b)
            diffs.append(f"pos{j}:{a}->{b}({'OCR' if compat else 'BAD'})")
            if not compat:
                all_compat = False

    if all_compat:
        compatible += 1
    else:
        incompatible += 1
        if incompatible <= 40:
            issuer = excel_rows[i]['issuer'][:30]
            print(f"{row:4d} {bb8:>12} {cusip[:12]:>12} {'INCOMPAT':>8} {', '.join(diffs)[:50]}  | {issuer}")

print(f"\nResults:")
print(f"  OCR-compatible (BB_ID ≈ CUSIP): {compatible}")
print(f"  Incompatible (wrong bond matched): {incompatible}")
print(f"  No comparison possible: {no_compare}")
print(f"  Compatibility rate: {compatible/(compatible+incompatible)*100:.1f}%")

# CUSIP check digit algorithm
def cusip_check_digit(base8):
    """Compute CUSIP check digit for an 8-character base."""
    values = []
    for ch in base8:
        if ch.isdigit():
            values.append(int(ch))
        elif ch.isalpha():
            values.append(ord(ch.upper()) - ord('A') + 10)
        elif ch == '*':
            values.append(36)
        elif ch == '@':
            values.append(37)
        elif ch == '#':
            values.append(38)
        else:
            values.append(0)

    total = 0
    for i, v in enumerate(values):
        if i % 2 == 1:  # Double odd positions (0-indexed)
            v *= 2
        total += v // 10 + v % 10

    return str((10 - (total % 10)) % 10)

# Show what the correct CUSIPs should be
print(f"\nSample correct CUSIPs (from BB_ID + check digit):")
for i in range(min(20, len(excel_rows))):
    bb_id = excel_rows[i]['bb_id']
    if bb_id and len(bb_id) >= 8:
        check = cusip_check_digit(bb_id[:8])
        correct_cusip = bb_id[:8] + check
        output_cusip = str(ws_out.cell(row=i+2, column=1).value or '').strip()
        match = 'MATCH' if correct_cusip == output_cusip else 'DIFF'
        print(f"  BB_ID: {bb_id:>12} -> Correct CUSIP: {correct_cusip} | Output: {output_cusip:>12} [{match}]")
