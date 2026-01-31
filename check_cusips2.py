#!/usr/bin/env python3
"""Analyze CUSIP column (col 1) in the output Excel."""
import openpyxl
import re
from collections import Counter

wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

total = 0
null_count = 0
valid_9 = 0
lengths = Counter()
issues = []
all_cusips = []

for row in range(2, ws.max_row + 1):
    total += 1
    val = ws.cell(row=row, column=1).value
    if val is None or not str(val).strip():
        null_count += 1
        continue
    s = str(val).strip()
    all_cusips.append(s)
    lengths[len(s)] += 1

    # Valid CUSIP: 9 alphanumeric characters
    if len(s) == 9 and re.match(r'^[A-Z0-9]{9}$', s):
        valid_9 += 1
    else:
        if len(issues) < 80:
            issuer = ws.cell(row=row, column=3).value or ''
            issues.append((row, s, len(s), str(issuer)[:40]))

print(f"Total rows: {total}")
print(f"Null/empty CUSIPs: {null_count}")
print(f"Valid 9-char alphanumeric CUSIPs: {valid_9}")
print(f"Non-standard CUSIPs: {total - null_count - valid_9}")

print(f"\nLength distribution:")
for length, cnt in sorted(lengths.items()):
    print(f"  {length} chars: {cnt}")

# Check duplicates
cusip_counts = Counter(all_cusips)
dupes = [(c, cnt) for c, cnt in cusip_counts.items() if cnt > 1]
print(f"\nDuplicate CUSIPs: {len(dupes)}")
if dupes:
    for c, cnt in sorted(dupes, key=lambda x: -x[1])[:20]:
        print(f"  '{c}' appears {cnt} times")

if issues:
    print(f"\nNon-standard CUSIPs (first 80):")
    for row, cusip, length, issuer in issues:
        print(f"  Row {row}: '{cusip}' (len={length}) | {issuer}")
