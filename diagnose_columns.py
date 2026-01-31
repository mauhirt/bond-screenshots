#!/usr/bin/env python3
"""Diagnose column alignment by showing headers and sample data for all columns."""
import openpyxl

wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

# Show headers
print("Column headers:")
for col in range(1, ws.max_column + 1):
    print(f"  Col {col:2d}: '{ws.cell(row=1, column=col).value}'")

# Show sample data for rows 2-6
print("\nSample data (rows 2-6):")
for row in range(2, 7):
    print(f"\n--- Row {row} ---")
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        val = ws.cell(row=row, column=col).value
        print(f"  {header:25s} = {val}")

# Also check what the original Excel has for CUSIPs
print("\n\n=== Original Excel (first 5 rows) ===")
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active
print("Column headers:")
for col in range(1, ws_orig.max_column + 1):
    print(f"  Col {col:2d}: '{ws_orig.cell(row=1, column=col).value}'")

print("\nSample data (rows 2-6):")
for row in range(2, 7):
    print(f"\n--- Row {row} ---")
    for col in range(1, min(ws_orig.max_column + 1, 21)):
        header = ws_orig.cell(row=1, column=col).value
        val = ws_orig.cell(row=row, column=col).value
        if val and isinstance(val, str) and len(val) > 60:
            val = val[:60] + '...'
        print(f"  {str(header):25s} = {val}")
