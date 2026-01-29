#!/usr/bin/env python3
"""
Populate green bonds Excel file from extracted CSV data.
Handles various data quality issues from screenshot extraction including:
- Split Tax Prov fields (e.g., "FED R ST" + "TAX-EXEMPT")
- Merged BICS/Yes-No fields (e.g., "Financing -- No", "Transporta..Yes")
- Merged Fin Typ/Tax fields (e.g., "ST TAX REFUNDING")
- Missing trailing fields (ESG Project, Project Subcat)
- "MM"/"M" suffixes on Amt Issued
- Truncated values ending with ".."
- Non-data text rows from extraction artifacts
"""

import csv
import re
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers

# ── Constants ──────────────────────────────────────────────────────────────

EXCEL_HEADERS = [
    'CUSIP', 'State Code', 'Issuer Name', 'Yield at Issue', 'Amt Issued',
    'Issue Date', 'Maturity', 'Tax Prov', 'Fin Typ', 'BICS Level 2',
    'Self-reported Green', 'Mgmt of Proc', 'ESG Reporting',
    'ESG Assurance Providers', 'Proj Sel Proc', 'ESG Framework',
    'Industry', 'Issuer Type', 'ESG Project Categories', 'Project Subcategory'
]

BICS_PREFIXES = ['Utilit', 'Financ', 'Educat', 'Trans', 'General']

FIN_TYP_PATTERNS = [
    'NEW MONEY', 'NEH MONEY', 'NEH MNEY', 'NEH HANEY', 'NEH HANCY',
    'NEH HMEY', 'NEH HONEY', 'NEW HONEY', 'REH MONEY', 'REW MONEY',
    'REF MONEY', 'NEH MONEY..', 'REFUNDING', 'REFINANCING', 'REFINANC',
    'REFINANC.', 'REFINANC..', 'REFIN&MNG', 'REFINMNG', 'REFINCING',
    'REFINANG', 'REFINIDG', 'REFUNDING..', 'REV',
]

ESG_PROJECT_KEYWORDS = [
    'Sustainable', 'Energy', 'Clean', 'Pollution', 'Green', 'Biodiversity',
    'Climate', 'Renewable', 'LEED', 'Conservation', 'Biogas',
    'Infrastructure', 'Public', 'Natural', 'Water',
]

# ── Parsing helpers ────────────────────────────────────────────────────────

def clean_cusip(val):
    """Clean CUSIP: strip 'TH ' prefix and whitespace."""
    s = val.strip()
    if s.startswith('TH '):
        s = s[3:].strip()
    return s


def is_data_row(fields):
    """Check if a CSV row contains bond data (not extraction artifact text)."""
    if len(fields) < 16:
        return False
    # CUSIP should be alphanumeric (allow / for OCR artifacts), 4-14 chars
    cusip = clean_cusip(fields[0])
    if not cusip or cusip == '--':
        return False
    if len(cusip) < 4 or len(cusip) > 14:
        return False
    if not re.match(r'^[A-Za-z0-9/]+$', cusip):
        return False
    # Check for all-dashes row
    if all(f.strip() == '--' for f in fields[:5]):
        return False
    # State can be empty, --, or 2 letters (don't reject on state alone)
    return True


def starts_with_bics(val):
    """Check if a value looks like a BICS Level 2 field."""
    clean = val.strip().rstrip('.')
    return any(clean.startswith(p) for p in BICS_PREFIXES)


def looks_like_fin_typ(val):
    """Check if value looks like a Financing Type field."""
    v = val.strip().upper().rstrip('.')
    for pat in FIN_TYP_PATTERNS:
        if v.startswith(pat.upper().rstrip('.')):
            return True
    if v == '--':
        return True
    return False


def is_yes_no_dash(val):
    """Check if value is Yes, No, or --."""
    return val.strip() in ('Yes', 'No', '--')


def looks_like_esg_project(val):
    """Check if value looks like an ESG Project Category."""
    return any(kw.lower() in val.lower() for kw in ESG_PROJECT_KEYWORDS)


def split_bics_merge(bics_raw):
    """Split merged BICS fields like 'Financing -- No' or 'Transporta..Yes'."""
    val = bics_raw.strip()

    # Pattern: "BICS -- YesNo" (e.g., "Financing -- No")
    m = re.match(r'^(.+?)\s+--\s+(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)

    # Pattern: "BICS - YesNo" (e.g., "Financing - Yes")
    m = re.match(r'^(.+?)\s+-\s+(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)

    # Pattern: "BICS..Yes" or "BICS..No" (e.g., "Transporta..Yes")
    m = re.match(r'^(.+?\.\.+)(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)

    # Pattern: "BICSYes" or "BICSNo" without dots (rare)
    m = re.match(r'^(Utilit\w*|Financ\w*|Educat\w*|Trans\w*|General\s*\w*)(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)

    return val, None


def extract_fin_typ_from_merged(val):
    """Extract Fin Typ from a merged Tax Prov + Fin Typ field.
    E.g., 'ST TAX REFUNDING' → ('ST TAX', 'REFUNDING')
    E.g., 'ST TAX-NEW MONEY' → ('ST TAX', 'NEW MONEY')
    """
    v = val.strip()
    for pat in sorted(FIN_TYP_PATTERNS, key=len, reverse=True):
        # Check for "TAX PATTERN" or "TAX-PATTERN"
        idx = v.upper().find(pat.upper())
        if idx > 0:
            before = v[:idx].rstrip(' -')
            after = v[idx:]
            # Only split if the before part looks tax-related
            if any(kw in before.upper() for kw in ['TAX', 'FED', 'AMT', 'EXEMPT', 'EXMPT', 'ST']):
                return before, after
    return val, None


def parse_row(fields):
    """
    Parse a CSV row into 20 Excel columns.
    Returns a list of 20 values, or None if the row can't be parsed.
    """
    if not is_data_row(fields):
        return None

    # Fixed first 7 columns (clean CUSIP and State)
    base = [clean_cusip(fields[0])] + list(fields[1:7])  # CUSIP, State, Issuer, Yield, Amt, IssueDate, Maturity
    rest = list(fields[7:])

    # ── Find BICS Level 2 (anchor point) ──
    bics_idx = None
    for i, f in enumerate(rest):
        if starts_with_bics(f):
            bics_idx = i
            break

    if bics_idx is None:
        # Fallback: try to find BICS in merged fields
        for i, f in enumerate(rest):
            bics_val, _ = split_bics_merge(f)
            if starts_with_bics(bics_val):
                bics_idx = i
                break

    if bics_idx is None:
        # Can't find BICS - use positional fallback
        return parse_row_positional(fields)

    # ── Pre-BICS: TaxProv + FinTyp ──
    pre_bics = rest[:bics_idx]

    # ── BICS field (handle merges) ──
    bics_raw = rest[bics_idx]
    bics_value, extra_yesno = split_bics_merge(bics_raw)

    # ── Post-BICS fields ──
    post_bics = rest[bics_idx + 1:]

    # ── Extract TaxProv and FinTyp ──
    if len(pre_bics) == 0:
        tax_prov = '--'
        fin_typ = '--'
    elif len(pre_bics) == 1:
        # Single field: could be just TaxProv or just FinTyp
        if looks_like_fin_typ(pre_bics[0]):
            tax_prov = '--'
            fin_typ = pre_bics[0]
        else:
            tax_prov = pre_bics[0]
            fin_typ = '--'
    elif len(pre_bics) == 2:
        tax_prov = pre_bics[0]
        fin_typ = pre_bics[1]
        # Check if fin_typ has merged Tax + FinTyp
        if not looks_like_fin_typ(fin_typ):
            merged_tax, extracted_fin = extract_fin_typ_from_merged(fin_typ)
            if extracted_fin:
                tax_prov = tax_prov + ' ' + merged_tax
                fin_typ = extracted_fin
            else:
                # Both look like Tax Prov, concatenate
                tax_prov = pre_bics[0] + ' ' + pre_bics[1]
                fin_typ = '--'
    else:
        # 3+ fields: last one might be FinTyp, rest is TaxProv
        if looks_like_fin_typ(pre_bics[-1]):
            tax_prov = ' '.join(pre_bics[:-1])
            fin_typ = pre_bics[-1]
        else:
            # Try to extract fin_typ from last field
            merged_tax, extracted_fin = extract_fin_typ_from_merged(pre_bics[-1])
            if extracted_fin:
                tax_prov = ' '.join(pre_bics[:-1]) + ' ' + merged_tax
                fin_typ = extracted_fin
            else:
                tax_prov = ' '.join(pre_bics)
                fin_typ = '--'

    # ── Collect Yes/No block (6 values) ──
    yesno = []
    if extra_yesno:
        yesno.append(extra_yesno)

    remaining = []
    collecting_yesno = True
    for f in post_bics:
        if collecting_yesno and is_yes_no_dash(f) and len(yesno) < 6:
            yesno.append(f.strip())
        else:
            collecting_yesno = False
            remaining.append(f)

    # Pad Yes/No to 6
    while len(yesno) < 6:
        yesno.append('--')

    # ── Post-block: Industry, [IssuerType], ESGProject, ProjectSubcat ──
    industry = remaining[0] if len(remaining) > 0 else '--'
    issuer_type = '--'
    esg_project = '--'
    project_subcat = '--'

    if len(remaining) >= 4:
        # 4+ fields: check if remaining[1] looks like an ESG project
        if looks_like_esg_project(remaining[1]):
            # Pattern: Industry, ESGProject, SubCat, extra
            # (happens when Yes/No block was short, shifting fields left)
            esg_project = remaining[1]
            project_subcat = remaining[2]
            # remaining[3] is extra (could be additional subcat)
        elif remaining[1] == '--' or (remaining[1].isupper() and len(remaining[1]) <= 12):
            # Pattern: Industry, IssuerType, ESGProject, SubCat
            issuer_type = remaining[1]
            esg_project = remaining[2]
            project_subcat = remaining[3]
        else:
            # Default: treat as Industry, ESGProject, SubCat + extra
            esg_project = remaining[1]
            project_subcat = remaining[2]
    elif len(remaining) == 3:
        # 3 fields: check if middle one is IssuerType or ESGProject
        if remaining[1] == '--' or (remaining[1].isupper() and not looks_like_esg_project(remaining[1])):
            # Looks like IssuerType
            issuer_type = remaining[1]
            esg_project = remaining[2]
        else:
            # Looks like ESGProject
            esg_project = remaining[1]
            project_subcat = remaining[2]
    elif len(remaining) == 2:
        # Could be (Industry, ESGProject) or (Industry, IssuerType)
        if looks_like_esg_project(remaining[1]):
            esg_project = remaining[1]
        elif remaining[1] == '--':
            issuer_type = remaining[1]
        else:
            esg_project = remaining[1]
    # len(remaining) <= 1: only industry (or nothing)

    # ── Build 20-column result ──
    result = base + [tax_prov, fin_typ, bics_value] + yesno[:6] + [
        industry, issuer_type, esg_project, project_subcat
    ]
    return result


def parse_row_positional(fields):
    """
    Fallback parser when BICS can't be found.
    Uses field count to guess positions.
    """
    n = len(fields)
    if n < 16:
        return None

    # Take first 7 as standard (clean CUSIP)
    base = [clean_cusip(fields[0])] + list(fields[1:7])
    rest = list(fields[7:])

    # Assume standard layout: TaxProv, FinTyp, BICS, 6xYesNo, Industry, ...
    if len(rest) >= 9:
        tax_prov = rest[0]
        fin_typ = rest[1]
        bics = rest[2]

        # Check if BICS is actually a Yes/No (alignment shift)
        if bics in ('Yes', 'No', '--') and not starts_with_bics(bics):
            # Shift: TaxProv was split, adjust
            tax_prov = rest[0] + ' ' + rest[1]
            fin_typ = rest[2]
            bics = rest[3] if len(rest) > 3 else '--'
            yesno_start = 4
        else:
            yesno_start = 3

        yesno = []
        idx = yesno_start
        while idx < len(rest) and len(yesno) < 6:
            if is_yes_no_dash(rest[idx]):
                yesno.append(rest[idx].strip())
                idx += 1
            else:
                break
        while len(yesno) < 6:
            yesno.append('--')

        remaining = rest[idx:]
        industry = remaining[0] if len(remaining) > 0 else '--'
        issuer_type = '--'
        esg_project = '--'
        project_subcat = '--'

        if len(remaining) >= 4:
            issuer_type = remaining[1]
            esg_project = remaining[2]
            project_subcat = remaining[3]
        elif len(remaining) == 3:
            if remaining[1] == '--' or (remaining[1].isupper() and not looks_like_esg_project(remaining[1])):
                issuer_type = remaining[1]
                esg_project = remaining[2]
            else:
                esg_project = remaining[1]
                project_subcat = remaining[2]
        elif len(remaining) == 2:
            esg_project = remaining[1]

        result = base + [tax_prov, fin_typ, bics] + yesno[:6] + [
            industry, issuer_type, esg_project, project_subcat
        ]
        return result
    else:
        # Very few remaining fields - just pad
        result = base + list(rest) + ['--'] * (20 - len(base) - len(rest))
        return result[:20]


# ── Data cleaning ──────────────────────────────────────────────────────────

def clean_amt_issued(val):
    """Remove MM/M/B suffix and convert to number."""
    if not val or val.strip() == '--':
        return None
    s = val.strip()
    # Remove % suffix (extraction artifact)
    s = re.sub(r'%$', '', s)
    # Remove B (billion) suffix - convert to millions (*1000)
    if re.search(r'B$', s):
        s = re.sub(r'B$', '', s)
        s = s.replace(',', '')
        try:
            return float(s) * 1000  # Convert billions to millions for consistency
        except ValueError:
            return val
    # Remove MM or M suffix
    s = re.sub(r'MM$', '', s)
    s = re.sub(r'M$', '', s)
    # Remove commas
    s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return val  # Return as-is if can't convert


def clean_yield(val):
    """Convert yield to number."""
    if not val or val.strip() == '--':
        return None
    s = val.strip().rstrip('.')
    # Remove % suffix
    s = re.sub(r'%$', '', s)
    try:
        return float(s)
    except ValueError:
        return val


def clean_date(val):
    """Convert date string to datetime."""
    if not val or val.strip() == '--':
        return None
    s = val.strip()
    # Try MM/DD/YYYY
    for fmt in ['%m/%d/%Y', '%m/%d/%y']:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return val  # Return as string if can't parse


def clean_text(val):
    """Clean text values - handle -- as empty."""
    if not val or val.strip() == '--':
        return None
    return val.strip()


def clean_yes_no(val):
    """Clean Yes/No/-- fields."""
    if not val or val.strip() == '--':
        return None
    v = val.strip()
    if v in ('Yes', 'No'):
        return v
    return None


# ── Main processing ────────────────────────────────────────────────────────

def main():
    csv_path = '/home/user/bond-screenshots/claude_table_output_2025_new.csv'
    output_path = '/home/user/bond-screenshots/green_bonds_2025_final.xlsx'

    # Read CSV
    rows_raw = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            rows_raw.append(row)

    print(f"Total CSV rows: {len(rows_raw)}")

    # Parse rows
    parsed_rows = []
    skipped_rows = []
    for i, raw in enumerate(rows_raw):
        parsed = parse_row(raw)
        if parsed:
            parsed_rows.append((i + 1, parsed))  # (line_number, data)
        else:
            if len(raw) >= 5:  # Only log rows that might have been data
                skipped_rows.append(i + 1)

    print(f"Parsed data rows: {len(parsed_rows)}")
    print(f"Skipped rows: {len(skipped_rows)}")
    if skipped_rows:
        print(f"  Skipped line numbers (first 20): {skipped_rows[:20]}")

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Municipals'

    # Write headers
    header_font = Font(bold=True)
    for col, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Write data
    date_format = 'MM/DD/YYYY'
    for row_idx, (line_num, data) in enumerate(parsed_rows, 2):
        # Col 1: CUSIP (text, strip TH prefix)
        ws.cell(row=row_idx, column=1, value=clean_cusip(data[0]))

        # Col 2: State Code
        ws.cell(row=row_idx, column=2, value=clean_text(data[1]))

        # Col 3: Issuer Name
        ws.cell(row=row_idx, column=3, value=clean_text(data[2]))

        # Col 4: Yield at Issue (number)
        yield_val = clean_yield(data[3])
        cell = ws.cell(row=row_idx, column=4, value=yield_val)
        if isinstance(yield_val, float):
            cell.number_format = '0.000'

        # Col 5: Amt Issued (number, remove MM/M)
        amt_val = clean_amt_issued(data[4])
        cell = ws.cell(row=row_idx, column=5, value=amt_val)
        if isinstance(amt_val, float):
            cell.number_format = '#,##0.000'

        # Col 6: Issue Date
        date_val = clean_date(data[5])
        cell = ws.cell(row=row_idx, column=6, value=date_val)
        if isinstance(date_val, datetime):
            cell.number_format = date_format

        # Col 7: Maturity
        mat_val = clean_date(data[6])
        cell = ws.cell(row=row_idx, column=7, value=mat_val)
        if isinstance(mat_val, datetime):
            cell.number_format = date_format

        # Col 8: Tax Prov
        ws.cell(row=row_idx, column=8, value=clean_text(data[7]))

        # Col 9: Fin Typ
        ws.cell(row=row_idx, column=9, value=clean_text(data[8]))

        # Col 10: BICS Level 2
        ws.cell(row=row_idx, column=10, value=clean_text(data[9]))

        # Col 11-16: Yes/No fields
        for j in range(6):
            ws.cell(row=row_idx, column=11 + j, value=clean_yes_no(data[10 + j]))

        # Col 17: Industry
        ws.cell(row=row_idx, column=17, value=clean_text(data[16]))

        # Col 18: Issuer Type (leave empty or --)
        ws.cell(row=row_idx, column=18, value=None)

        # Col 19: ESG Project Categories
        ws.cell(row=row_idx, column=19, value=clean_text(data[18]))

        # Col 20: Project Subcategory
        ws.cell(row=row_idx, column=20, value=clean_text(data[19]))

    # Auto-adjust column widths
    for col in range(1, 21):
        max_width = len(EXCEL_HEADERS[col - 1])
        for row in range(2, min(50, ws.max_row + 1)):  # Sample first 50 rows
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                max_width = max(max_width, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_width + 2, 40)

    # Save
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"Total rows (including header): {ws.max_row}")
    print(f"Total data rows: {ws.max_row - 1}")

    return parsed_rows, skipped_rows


if __name__ == '__main__':
    parsed, skipped = main()
