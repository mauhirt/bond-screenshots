#!/usr/bin/env python3
"""Final comprehensive verification of v3 output."""
import openpyxl
import re
from collections import Counter
from datetime import datetime

wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

total = ws.max_row - 1
print(f"Total data rows: {total}")

# 1. CUSIP validation
valid_cusip = 0
for row in range(2, ws.max_row + 1):
    cusip = str(ws.cell(row=row, column=1).value or '')
    if len(cusip) == 9 and re.match(r'^[A-Z0-9]{9}$', cusip):
        valid_cusip += 1
print(f"\n1. CUSIPs: {valid_cusip}/{total} valid 9-char ({valid_cusip/total*100:.1f}%)")

# Check for duplicates
cusips = [str(ws.cell(row=r, column=1).value) for r in range(2, ws.max_row + 1)]
dupes = [c for c, cnt in Counter(cusips).items() if cnt > 1]
print(f"   Duplicate CUSIPs: {len(dupes)}")

# 2. Column completeness
cols = ['CUSIP', 'State Code', 'Issuer Name', 'Yield at Issue', 'Amt Issued',
        'Issue Date', 'Maturity', 'Tax Prov', 'Fin Typ', 'BICS Level 2',
        'Self-reported Green', 'Mgmt of Proc', 'ESG Reporting',
        'ESG Assurance Providers', 'Proj Sel Proc', 'ESG Framework',
        'Industry', 'Issuer Type', 'ESG Project Categories', 'Project Subcategory']

print(f"\n2. Column completeness:")
for i, name in enumerate(cols, 1):
    non_null = sum(1 for r in range(2, ws.max_row + 1)
                   if ws.cell(row=r, column=i).value is not None and str(ws.cell(row=r, column=i).value).strip())
    pct = non_null / total * 100
    print(f"   Col {i:2d} {name:25s}: {non_null:5d}/{total} ({pct:5.1f}%)")

# 3. Categorical column validation
print(f"\n3. Categorical validation:")

# Tax Prov
VALID_TAX = ['AMT/ST TAX-EXEMPT', 'AMT/ST TAXABLE', 'FED & ST TAX-EXEMPT',
    'FED AMT FOR INDIVIDUALS', 'FED BQ', 'FED BQ/ST TAX-EXEMPT',
    'FED BQ/ST TAXABLE', 'FED TAX-EXEMPT', 'FED TAX-EXEMPT/ST TAXABLE',
    'FED TAXABLE', 'FED TAXABLE/ST TAX-EXEMPT', 'FED TAXABLE/ST TAXABLE']
tax_counter = Counter()
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=8).value
    if v: tax_counter[str(v)] += 1
tax_inv = sum(c for v, c in tax_counter.items() if v not in VALID_TAX)
print(f"   Tax Prov: {sum(tax_counter.values())} non-null, {len(tax_counter)} categories, {tax_inv} invalid")

# Fin Typ
fin_counter = Counter()
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=9).value
    if v: fin_counter[str(v)] += 1
fin_inv = sum(c for v, c in fin_counter.items() if v not in ('NEW MONEY', 'REFUNDING'))
print(f"   Fin Typ: {sum(fin_counter.values())} non-null, {fin_inv} invalid")
for v, c in fin_counter.most_common():
    print(f"     [{c:4d}] {v}")

# BICS
VALID_BICS = ['Education', 'Financing', 'General Government', 'Health Care',
    'Housing', 'NA', 'Post Employment', 'Public Services', 'Transportation', 'Utilities']
bics_counter = Counter()
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=10).value
    if v: bics_counter[str(v)] += 1
bics_inv = sum(c for v, c in bics_counter.items() if v not in VALID_BICS)
print(f"   BICS: {sum(bics_counter.values())} non-null, {bics_inv} invalid")

# Industry
VALID_IND = ['APPROP', 'ARPT', 'BONDBK', 'CCRC', 'CDD', 'CHRT', 'CMNTYC', 'DEV',
    'EDU', 'EDULEASE', 'GARVEE', 'GASFWD', 'GO', 'GODIST', 'GOVLEASE',
    'GOVTGTD', 'HGR', 'HOSP', 'HOTELTAX', 'INCTAX', 'LMFH', 'LNGBDL',
    'LNPOOL', 'MDD', 'MEL', 'MISC', 'MISCTAX', 'MUNUTIL', 'NA', 'NFPCULT',
    'NFPRO', 'PILOT', 'PORTS', 'PUBPWR', 'PUBTRAN', 'PUBWTR', 'SALESTAX',
    'SCD', 'SCO', 'SELFAPP', 'SMFH', 'SOLWST', 'SPLASMT', 'STDHSG', 'TIF',
    'TOLL', 'TRIBES', 'TXMUD', 'WSGTD', 'WTRSWR']
ind_counter = Counter()
for r in range(2, ws.max_row + 1):
    v = ws.cell(row=r, column=17).value
    if v: ind_counter[str(v)] += 1
ind_inv = sum(c for v, c in ind_counter.items() if v not in VALID_IND)
print(f"   Industry: {sum(ind_counter.values())} non-null, {len(ind_counter)} categories, {ind_inv} invalid")

# 4. Date validation
print(f"\n4. Date validation:")
issue_dates = 0
issue_string = 0
maturities = 0
mat_string = 0
swapped = 0
for r in range(2, ws.max_row + 1):
    id_val = ws.cell(row=r, column=6).value
    mat_val = ws.cell(row=r, column=7).value
    if id_val:
        if isinstance(id_val, datetime):
            issue_dates += 1
        else:
            issue_string += 1
    if mat_val:
        if isinstance(mat_val, datetime):
            maturities += 1
        else:
            mat_string += 1
    if isinstance(id_val, datetime) and isinstance(mat_val, datetime):
        if id_val > mat_val:
            swapped += 1

print(f"   Issue dates: {issue_dates} datetime, {issue_string} string")
print(f"   Maturities: {maturities} datetime, {mat_string} string")
print(f"   Remaining swapped (issue > maturity): {swapped}")

# 5. Yes/No columns
print(f"\n5. Yes/No columns:")
for col in range(11, 17):
    vals = Counter()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v: vals[str(v)] += 1
    total_yn = sum(vals.values())
    valid_yn = vals.get('Yes', 0) + vals.get('No', 0)
    print(f"   Col {col} ({cols[col-1]}): {total_yn} non-null, {valid_yn} Yes/No, {total_yn - valid_yn} other")

# 6. Sample rows
print(f"\n6. Sample rows:")
for r in [2, 100, 500, 1000, 1500, 1826]:
    cusip = ws.cell(row=r, column=1).value
    state = ws.cell(row=r, column=2).value
    issuer = str(ws.cell(row=r, column=3).value or '')[:35]
    yld = ws.cell(row=r, column=4).value
    amt = ws.cell(row=r, column=5).value
    tax = ws.cell(row=r, column=8).value
    ind = ws.cell(row=r, column=17).value
    print(f"   Row {r:5d}: {cusip} | {state or '--':>2} | {issuer:35s} | yld={yld} | amt={amt} | tax={tax} | ind={ind}")
