#!/usr/bin/env python3
"""
Strict page-boundary deduplication.

True page-boundary duplicates are the SAME row captured in two overlapping screenshots.
They should have identical or near-identical raw field values.

Strategy: compare first 6 raw fields (CUSIP, State, Issuer, Yield, Amt, Issue Date)
as exact strings. If all match, it's a duplicate.
"""
import csv
import re
import openpyxl

# Load CSV
csv_rows = []
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    for row_data in reader:
        if len(row_data) >= 10:
            cusip = row_data[0].strip()
            if cusip.startswith('TH '):
                cusip = cusip[3:]
            if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                if not all(f.strip() == '--' for f in row_data[:5]):
                    csv_rows.append(row_data)

print(f"CSV rows: {len(csv_rows)}")

# Method 1: Exact match on CUSIP + first 5 key fields
def key_exact(row):
    """Generate exact key from first 6 fields."""
    return tuple(row[i].strip() for i in range(min(6, len(row))))

def key_cusip_issuer(row):
    """Key = CUSIP + issuer prefix."""
    return (row[0].strip(), row[2].strip()[:15] if len(row) > 2 else '')

# Try several dedup strategies
print("\n--- Strategy 1: Consecutive exact CUSIP dedup ---")
dedup1 = [csv_rows[0]]
for i in range(1, len(csv_rows)):
    if csv_rows[i][0].strip() != csv_rows[i-1][0].strip():
        dedup1.append(csv_rows[i])
print(f"Result: {len(dedup1)} (removed {len(csv_rows) - len(dedup1)})")

print("\n--- Strategy 2: Consecutive exact first-6-fields dedup ---")
dedup2 = [csv_rows[0]]
for i in range(1, len(csv_rows)):
    if key_exact(csv_rows[i]) != key_exact(csv_rows[i-1]):
        dedup2.append(csv_rows[i])
print(f"Result: {len(dedup2)} (removed {len(csv_rows) - len(dedup2)})")

print("\n--- Strategy 3: Window-based exact CUSIP dedup (window=5) ---")
dedup3 = [csv_rows[0]]
for i in range(1, len(csv_rows)):
    cusip = csv_rows[i][0].strip()
    is_dupe = False
    for j in range(max(0, len(dedup3)-5), len(dedup3)):
        if dedup3[j][0].strip() == cusip:
            is_dupe = True
            break
    if not is_dupe:
        dedup3.append(csv_rows[i])
print(f"Result: {len(dedup3)} (removed {len(csv_rows) - len(dedup3)})")

print("\n--- Strategy 4: Window-based exact first-6-fields dedup (window=8) ---")
dedup4 = [csv_rows[0]]
for i in range(1, len(csv_rows)):
    key = key_exact(csv_rows[i])
    is_dupe = False
    for j in range(max(0, len(dedup4)-8), len(dedup4)):
        if key_exact(dedup4[j]) == key:
            is_dupe = True
            break
    if not is_dupe:
        dedup4.append(csv_rows[i])
print(f"Result: {len(dedup4)} (removed {len(csv_rows) - len(dedup4)})")

print("\n--- Strategy 5: Window-based CUSIP+Issuer dedup (window=8) ---")
dedup5 = [csv_rows[0]]
for i in range(1, len(csv_rows)):
    key = key_cusip_issuer(csv_rows[i])
    is_dupe = False
    for j in range(max(0, len(dedup5)-8), len(dedup5)):
        if key_cusip_issuer(dedup5[j]) == key:
            is_dupe = True
            break
    if not is_dupe:
        dedup5.append(csv_rows[i])
print(f"Result: {len(dedup5)} (removed {len(csv_rows) - len(dedup5)})")

# For the best strategy, test positional matching
best_dedup = dedup1  # Start with simplest
best_name = "Strategy 1 (consecutive CUSIP)"

for name, dd in [("1: consec CUSIP", dedup1), ("2: consec 6-fields", dedup2),
                  ("3: window CUSIP", dedup3), ("4: window 6-fields", dedup4),
                  ("5: window CUSIP+issuer", dedup5)]:
    if abs(len(dd) - 1825) < abs(len(best_dedup) - 1825):
        best_dedup = dd
        best_name = name

print(f"\nBest strategy: {best_name} with {len(best_dedup)} rows (diff from 1825: {len(best_dedup) - 1825})")

# Positional matching test with best dedup
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    excel_rows.append(issuer)

# Test with dedup1 (simplest, closest to 1825)
for name, dd in [("1: consec CUSIP", dedup1), ("3: window CUSIP", dedup3), ("5: window CUSIP+issuer", dedup5)]:
    match = 0
    mismatch = 0
    first_mm = None
    for i in range(min(len(excel_rows), len(dd))):
        e = excel_rows[i].lower()[:12]
        c = dd[i][2].strip().lower().rstrip('.')[:12] if len(dd[i]) > 2 else ''
        if e == c:
            match += 1
        else:
            mismatch += 1
            if first_mm is None:
                first_mm = i
    print(f"\n{name}: {match}/{min(len(excel_rows), len(dd))} issuer match (first mismatch at {first_mm})")
    if first_mm is not None:
        for j in range(max(0, first_mm-2), min(len(dd), first_mm+5)):
            e_iss = excel_rows[j][:35] if j < len(excel_rows) else 'N/A'
            c_iss = dd[j][2].strip()[:25] if j < len(dd) and len(dd[j]) > 2 else 'N/A'
            marker = ' <<<' if j == first_mm else ''
            print(f"    [{j:4d}] Excel: '{e_iss:35s}' | CSV: '{c_iss:25s}'{marker}")
