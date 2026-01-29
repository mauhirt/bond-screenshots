#!/usr/bin/env python3
"""
compare_rows.py
Replay the matching from populate_excel_v2.py, then compare parsed CSV data
against what is actually stored in green_bonds_2025_final.xlsx.
Reports all field-level discrepancies.
"""

import csv
import re
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

# ============================================================================
# Copied verbatim from populate_excel_v2.py -- all parsing/matching functions
# ============================================================================

BICS_PREFIXES = ['Utilit', 'Financ', 'Educat', 'Trans', 'General']

FIN_TYP_PATTERNS = [
    'NEW MONEY', 'NEH MONEY', 'NEH MNEY', 'NEH HANEY', 'NEH HANCY',
    'NEH HMEY', 'NEH HONEY', 'NEW HONEY', 'REH MONEY', 'REW MONEY',
    'REF MONEY', 'NEH MONEY..', 'REFUNDING', 'REFINANCING', 'REFINANC',
    'REFINANC.', 'REFINANC..', 'REFIN&MNG', 'REFINMNG', 'REFINCING',
    'REFINANG', 'REFINIDG', 'REFUNDING..', 'REV',
]

ESG_PROJECT_KEYWORDS = [
    'Sustainable', 'Energy', 'Clean', 'Pollution', 'Green', 'Biodiversity',
    'Climate', 'Renewable', 'LEED', 'Conservation', 'Biogas',
    'Infrastructure', 'Public', 'Natural', 'Water',
]

OCR_PAIRS = {
    ('0','X'),('X','0'),('0','O'),('O','0'),('0','D'),('D','0'),
    ('1','I'),('I','1'),('1','L'),('L','1'),('1','7'),('7','1'),
    ('3','2'),('2','3'),('3','8'),('8','3'),
    ('4','8'),('8','4'),('4','A'),('A','4'),
    ('5','S'),('S','5'),('5','F'),('F','5'),('5','6'),('6','5'),
    ('6','G'),('G','6'),('6','8'),('8','6'),
    ('7','T'),('T','7'),
    ('8','B'),('B','8'),
    ('9','G'),('G','9'),('9','Q'),('Q','9'),
    ('D','K'),('K','D'),('D','H'),('H','D'),
    ('E','F'),('F','E'),
    ('K','H'),('H','K'),('K','X'),('X','K'),
    ('M','N'),('N','M'),('M','W'),('W','M'),
    ('P','R'),('R','P'),
    ('U','V'),('V','U'),('W','V'),('V','W'),
}


def bb_to_cusip_score(bb_id, cusip):
    bb8 = bb_id[:8].ljust(8)
    c8 = cusip[:8].ljust(8) if len(cusip) >= 8 else cusip.ljust(8)
    score = 0
    for a, b in zip(bb8, c8):
        if a == b:
            score += 3
        elif a.upper() == b.upper():
            score += 2
        elif (a, b) in OCR_PAIRS:
            score += 1
    return score


def issuer_prefix_score(excel_issuer, csv_issuer):
    e = excel_issuer.lower().strip()[:15]
    c = csv_issuer.lower().strip()[:15]
    if not e or not c:
        return 0
    match = 0
    for a, b in zip(e, c):
        if a == b:
            match += 1
        else:
            break
    return match


def clean_cusip(val):
    s = val.strip()
    if s.startswith('TH '):
        s = s[3:].strip()
    return s

def starts_with_bics(val):
    clean = val.strip().rstrip('.')
    return any(clean.startswith(p) for p in BICS_PREFIXES)

def looks_like_fin_typ(val):
    v = val.strip().upper().rstrip('.')
    for pat in FIN_TYP_PATTERNS:
        if v.startswith(pat.upper().rstrip('.')):
            return True
    if v == '--':
        return True
    return False

def is_yes_no_dash(val):
    return val.strip() in ('Yes', 'No', '--')

def looks_like_esg_project(val):
    return any(kw.lower() in val.lower() for kw in ESG_PROJECT_KEYWORDS)

def split_bics_merge(bics_raw):
    val = bics_raw.strip()
    m = re.match(r'^(.+?)\s+--\s+(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)
    m = re.match(r'^(.+?)\s+-\s+(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)
    m = re.match(r'^(.+?\.\.+)(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)
    m = re.match(r'^(Utilit\w*|Financ\w*|Educat\w*|Trans\w*|General\s*\w*)(Yes|No)$', val)
    if m:
        return m.group(1).strip(), m.group(2)
    return val, None

def extract_fin_typ_from_merged(val):
    v = val.strip()
    for pat in sorted(FIN_TYP_PATTERNS, key=len, reverse=True):
        idx = v.upper().find(pat.upper())
        if idx > 0:
            before = v[:idx].rstrip(' -')
            after = v[idx:]
            if any(kw in before.upper() for kw in ['TAX', 'FED', 'AMT', 'EXEMPT', 'EXMPT', 'ST']):
                return before, after
    return val, None

def parse_csv_row(fields):
    if len(fields) < 10:
        return None

    cusip = clean_cusip(fields[0])
    base = [cusip] + list(fields[1:7])
    rest = list(fields[7:])

    bics_idx = None
    for i, f in enumerate(rest):
        if starts_with_bics(f):
            bics_idx = i
            break
    if bics_idx is None:
        for i, f in enumerate(rest):
            bv, _ = split_bics_merge(f)
            if starts_with_bics(bv):
                bics_idx = i
                break

    if bics_idx is None:
        if len(rest) >= 9:
            tax_prov = rest[0]
            fin_typ = rest[1]
            bics = rest[2]
            yesno_start = 3
            yesno = []
            idx = yesno_start
            while idx < len(rest) and len(yesno) < 6:
                if is_yes_no_dash(rest[idx]):
                    yesno.append(rest[idx].strip())
                    idx += 1
                else:
                    break
            while len(yesno) < 6:
                yesno.append('--')
            remaining = rest[idx:]
            industry = remaining[0] if remaining else '--'
            esg_project = remaining[1] if len(remaining) > 1 else '--'
            project_subcat = remaining[2] if len(remaining) > 2 else '--'
            result = base + [tax_prov, fin_typ, bics] + yesno[:6] + [
                industry, '--', esg_project, project_subcat]
            return {i: result[i] for i in range(20)}
        return None

    pre_bics = rest[:bics_idx]
    bics_raw = rest[bics_idx]
    bics_value, extra_yesno = split_bics_merge(bics_raw)
    post_bics = rest[bics_idx + 1:]

    if len(pre_bics) == 0:
        tax_prov, fin_typ = '--', '--'
    elif len(pre_bics) == 1:
        if looks_like_fin_typ(pre_bics[0]):
            tax_prov, fin_typ = '--', pre_bics[0]
        else:
            tax_prov, fin_typ = pre_bics[0], '--'
    elif len(pre_bics) == 2:
        tax_prov, fin_typ = pre_bics[0], pre_bics[1]
        if not looks_like_fin_typ(fin_typ):
            mt, ef = extract_fin_typ_from_merged(fin_typ)
            if ef:
                tax_prov = tax_prov + ' ' + mt
                fin_typ = ef
            else:
                tax_prov = pre_bics[0] + ' ' + pre_bics[1]
                fin_typ = '--'
    else:
        if looks_like_fin_typ(pre_bics[-1]):
            tax_prov = ' '.join(pre_bics[:-1])
            fin_typ = pre_bics[-1]
        else:
            mt, ef = extract_fin_typ_from_merged(pre_bics[-1])
            if ef:
                tax_prov = ' '.join(pre_bics[:-1]) + ' ' + mt
                fin_typ = ef
            else:
                tax_prov = ' '.join(pre_bics)
                fin_typ = '--'

    yesno = []
    if extra_yesno:
        yesno.append(extra_yesno)
    remaining = []
    collecting = True
    for f in post_bics:
        if collecting and is_yes_no_dash(f) and len(yesno) < 6:
            yesno.append(f.strip())
        else:
            collecting = False
            remaining.append(f)
    while len(yesno) < 6:
        yesno.append('--')

    industry = remaining[0] if remaining else '--'
    issuer_type = '--'
    esg_project = '--'
    project_subcat = '--'
    if len(remaining) >= 4:
        if looks_like_esg_project(remaining[1]):
            esg_project = remaining[1]
            project_subcat = remaining[2]
        elif remaining[1] == '--' or (remaining[1].isupper() and len(remaining[1]) <= 12):
            issuer_type = remaining[1]
            esg_project = remaining[2]
            project_subcat = remaining[3]
        else:
            esg_project = remaining[1]
            project_subcat = remaining[2]
    elif len(remaining) == 3:
        if remaining[1] == '--' or (remaining[1].isupper() and not looks_like_esg_project(remaining[1])):
            issuer_type = remaining[1]
            esg_project = remaining[2]
        else:
            esg_project = remaining[1]
            project_subcat = remaining[2]
    elif len(remaining) == 2:
        if looks_like_esg_project(remaining[1]):
            esg_project = remaining[1]
        else:
            esg_project = remaining[1]

    result = base + [tax_prov, fin_typ, bics_value] + yesno[:6] + [
        industry, issuer_type, esg_project, project_subcat]
    return {i: result[i] for i in range(min(20, len(result)))}


def clean_text(val):
    if not val or val.strip() == '--':
        return None
    return val.strip()

def clean_yield(val):
    if not val or val.strip() == '--':
        return None
    s = val.strip().rstrip('.')
    s = re.sub(r'%$', '', s)
    try:
        return float(s)
    except ValueError:
        return val

def clean_date(val):
    if not val or val.strip() == '--':
        return None
    s = val.strip()
    for fmt in ['%m/%d/%Y', '%m/%d/%y']:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return val

def clean_yes_no(val):
    if not val or val.strip() == '--':
        return None
    v = val.strip()
    return v if v in ('Yes', 'No') else None


# ============================================================================
# Comparison logic
# ============================================================================

COLUMN_NAMES = {
    1:  'CUSIP',
    2:  'State Code',
    3:  'Issuer Name',
    4:  'Yield at Issue',
    5:  'Amt Issued',
    6:  'Issue Date',
    7:  'Maturity',
    8:  'Tax Prov',
    9:  'Fin Typ',
    10: 'BICS Level 2',
    11: 'Self-reported Green',
    12: 'Mgmt of Proc',
    13: 'ESG Reporting',
    14: 'ESG Assurance Providers',
    15: 'Proj Sel Proc',
    16: 'ESG Framework',
    17: 'Industry',
    18: 'Issuer Type',
    19: 'ESG Project Categories',
    20: 'Project Subcategory',
}


def normalize_for_compare(val):
    """Normalize a value for comparison purposes."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, float):
        # Round to 6 decimals to avoid float precision issues
        return round(val, 6)
    if isinstance(val, (int,)):
        return val
    s = str(val).strip()
    if s == '' or s == '--' or s == 'None':
        return None
    return s


def main():
    orig_path = '/home/user/bond-screenshots/green bonds excel.xlsx'
    csv_path = '/home/user/bond-screenshots/claude_table_output_2025_new.csv'
    final_path = '/home/user/bond-screenshots/green_bonds_2025_final.xlsx'

    # ── Step 1: Load original Excel (to get BB_IDs, issuers, amts) ──
    print("=" * 80)
    print("STEP 1: Loading original Excel (ground truth for BB_ID, Issuer, Amt)")
    print("=" * 80)
    orig_wb = load_workbook(orig_path)
    orig_ws = orig_wb.active

    excel_rows = []
    for row in range(2, orig_ws.max_row + 1):
        issuer = (orig_ws.cell(row=row, column=3).value or '').strip()
        amt = orig_ws.cell(row=row, column=5).value
        formula = orig_ws.cell(row=row, column=2).value or ''
        m = re.search(r'"([^"]+)\s+Muni"', str(formula))
        bb_id = m.group(1) if m else ''
        excel_rows.append({'idx': row - 2, 'row': row, 'bb_id': bb_id,
                           'issuer': issuer, 'amt': amt})

    print(f"  Original Excel rows: {len(excel_rows)}")

    # ── Step 2: Load and parse CSV ──
    print("\n" + "=" * 80)
    print("STEP 2: Loading and parsing CSV")
    print("=" * 80)
    csv_all = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader, 1):
            if len(row) >= 10:
                cusip = row[0].strip()
                if cusip.startswith('TH '):
                    cusip = cusip[3:]
                if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                    if not all(f.strip() == '--' for f in row[:5]):
                        parsed = parse_csv_row(row)
                        csv_all.append({
                            'idx': len(csv_all), 'line': i, 'cusip': cusip,
                            'issuer': row[2].strip(), 'fields': row,
                            'parsed': parsed
                        })

    print(f"  CSV data rows: {len(csv_all)}")

    # ── Step 3: Replay matching (exact same logic as populate_excel_v2.py) ──
    print("\n" + "=" * 80)
    print("STEP 3: Replaying multi-phase matching")
    print("=" * 80)

    matched = {}   # excel_idx -> csv_idx
    matched_csv = set()

    # Phase 1: BB_ID -> CUSIP (OCR-aware)
    print("  Phase 1: BB_ID matching...")
    pairs = []
    for ei, e in enumerate(excel_rows):
        bb = e['bb_id']
        if not bb:
            continue
        for ci, c in enumerate(csv_all):
            score = bb_to_cusip_score(bb, c['cusip'])
            if score >= 10:
                pairs.append((score, ei, ci))

    pairs.sort(key=lambda x: -x[0])
    for score, ei, ci in pairs:
        if ei in matched or ci in matched_csv:
            continue
        matched[ei] = ci
        matched_csv.add(ci)
    print(f"    Matched: {len(matched)}")

    # Phase 2: Issuer name + BB_ID
    print("  Phase 2: Issuer + BB_ID matching...")
    csv_by_issuer = defaultdict(list)
    for c in csv_all:
        if c['idx'] not in matched_csv:
            key = c['issuer'].lower()[:10]
            csv_by_issuer[key].append(c)

    new_matches = 0
    for e in excel_rows:
        if e['idx'] in matched:
            continue
        e_key = e['issuer'].lower()[:10]
        best_score = 0
        best_ci = None
        for csv_key, candidates in csv_by_issuer.items():
            issuer_sim = issuer_prefix_score(e_key, csv_key)
            if issuer_sim < 5:
                continue
            for c in candidates:
                if c['idx'] in matched_csv:
                    continue
                bb_score = bb_to_cusip_score(e['bb_id'], c['cusip']) if e['bb_id'] else 0
                total = issuer_sim * 2 + bb_score
                if total > best_score:
                    best_score = total
                    best_ci = c['idx']
        if best_ci is not None and best_score >= 15:
            matched[e['idx']] = best_ci
            matched_csv.add(best_ci)
            new_matches += 1
    print(f"    New matches: {new_matches}, Total: {len(matched)}")

    unmatched_count = len(excel_rows) - len(matched)
    print(f"  Unmatched Excel rows: {unmatched_count}")

    # ── Step 4: Load the final output Excel ──
    print("\n" + "=" * 80)
    print("STEP 4: Loading final Excel (green_bonds_2025_final.xlsx)")
    print("=" * 80)
    final_wb = load_workbook(final_path)
    final_ws = final_wb.active
    print(f"  Final Excel rows (incl header): {final_ws.max_row}")
    print(f"  Final Excel data rows: {final_ws.max_row - 1}")

    # ── Step 5: Compare field by field ──
    print("\n" + "=" * 80)
    print("STEP 5: Field-by-field comparison")
    print("=" * 80)

    # Tracking per column: col_num -> { 'match': count, 'mismatch': count,
    #                                     'both_none': count, 'samples': [...] }
    col_stats = {}
    for col in range(1, 21):
        col_stats[col] = {'match': 0, 'mismatch': 0, 'both_none': 0,
                          'samples': [], 'compared': 0}

    total_matched_with_parsed = 0
    total_matched_no_parsed = 0

    for e in excel_rows:
        ei = e['idx']
        out_row = ei + 2  # Excel row number in final output (1-indexed, row 1 = header)
        excel_issuer = e['issuer']
        excel_amt = e['amt']
        bb_id = e['bb_id']

        if ei not in matched:
            # Unmatched row -- skip comparison (only has BB_ID/issuer/amt)
            continue

        ci = matched[ei]
        c = csv_all[ci]
        p = c['parsed']

        if p is None:
            total_matched_no_parsed += 1
            continue

        total_matched_with_parsed += 1

        # Build the EXPECTED values using the same logic as populate_excel_v2.py
        expected = {}
        # Col 1: CUSIP
        expected[1] = clean_text(p.get(0, ''))
        # Col 2: State
        expected[2] = clean_text(p.get(1, ''))
        # Col 3: Issuer Name (from original Excel)
        expected[3] = excel_issuer
        # Col 4: Yield
        expected[4] = clean_yield(p.get(3, ''))
        # Col 5: Amt (from original Excel)
        expected[5] = excel_amt
        # Col 6: Issue Date
        expected[6] = clean_date(p.get(5, ''))
        # Col 7: Maturity
        expected[7] = clean_date(p.get(6, ''))
        # Col 8: Tax Prov
        expected[8] = clean_text(p.get(7, ''))
        # Col 9: Fin Typ
        expected[9] = clean_text(p.get(8, ''))
        # Col 10: BICS Level 2
        expected[10] = clean_text(p.get(9, ''))
        # Col 11-16: Yes/No fields
        for j in range(6):
            expected[11 + j] = clean_yes_no(p.get(10 + j, ''))
        # Col 17: Industry
        expected[17] = clean_text(p.get(16, ''))
        # Col 18: Issuer Type (always None in populate_excel_v2.py)
        expected[18] = None
        # Col 19: ESG Project Categories
        expected[19] = clean_text(p.get(18, ''))
        # Col 20: Project Subcategory
        expected[20] = clean_text(p.get(19, ''))

        # Read actual values from final Excel
        for col in range(1, 21):
            actual_raw = final_ws.cell(row=out_row, column=col).value
            expected_raw = expected[col]

            actual_norm = normalize_for_compare(actual_raw)
            expected_norm = normalize_for_compare(expected_raw)

            col_stats[col]['compared'] += 1

            if actual_norm is None and expected_norm is None:
                col_stats[col]['both_none'] += 1
                col_stats[col]['match'] += 1
            elif actual_norm == expected_norm:
                col_stats[col]['match'] += 1
            else:
                col_stats[col]['mismatch'] += 1
                if len(col_stats[col]['samples']) < 5:
                    col_stats[col]['samples'].append({
                        'excel_row': out_row,
                        'excel_idx': ei,
                        'csv_idx': ci,
                        'bb_id': bb_id,
                        'expected_raw': expected_raw,
                        'actual_raw': actual_raw,
                        'expected_norm': expected_norm,
                        'actual_norm': actual_norm,
                    })

    # ── Print results ──
    print(f"\n  Matched rows with parsed CSV data: {total_matched_with_parsed}")
    print(f"  Matched rows where parse returned None: {total_matched_no_parsed}")
    print(f"  Unmatched rows (skipped): {unmatched_count}")

    print("\n" + "=" * 80)
    print("RESULTS: Per-column comparison summary")
    print("=" * 80)

    total_mismatches_all = 0
    cols_with_issues = []

    for col in range(1, 21):
        s = col_stats[col]
        name = COLUMN_NAMES[col]
        compared = s['compared']
        match_count = s['match']
        mismatch_count = s['mismatch']
        both_none = s['both_none']
        total_mismatches_all += mismatch_count

        pct = (match_count / compared * 100) if compared > 0 else 0.0

        status = "OK" if mismatch_count == 0 else "MISMATCHES"
        print(f"\n  Col {col:2d} ({name:30s}): "
              f"compared={compared:5d}  match={match_count:5d}  "
              f"mismatch={mismatch_count:5d}  both_none={both_none:5d}  "
              f"match%={pct:6.2f}%  [{status}]")

        if mismatch_count > 0:
            cols_with_issues.append(col)
            print(f"         Sample mismatches (up to 5):")
            for i, samp in enumerate(s['samples']):
                print(f"           [{i+1}] Row {samp['excel_row']}, "
                      f"BB_ID='{samp['bb_id']}', CSV_idx={samp['csv_idx']}")
                print(f"               Expected: {repr(samp['expected_norm'])}  "
                      f"(raw: {repr(samp['expected_raw'])})")
                print(f"               Actual:   {repr(samp['actual_norm'])}  "
                      f"(raw: {repr(samp['actual_raw'])})")

    # ── Summary ──
    print("\n" + "=" * 80)
    print("OVERALL SUMMARY")
    print("=" * 80)
    print(f"  Total rows compared:          {total_matched_with_parsed}")
    print(f"  Total field comparisons:      {total_matched_with_parsed * 20}")
    print(f"  Total field mismatches:       {total_mismatches_all}")
    print(f"  Columns with mismatches:      {len(cols_with_issues)}")
    if cols_with_issues:
        print(f"  Affected columns:")
        for col in cols_with_issues:
            s = col_stats[col]
            print(f"    Col {col:2d} ({COLUMN_NAMES[col]}): "
                  f"{s['mismatch']} mismatches out of {s['compared']} compared")
    else:
        print("  ALL COLUMNS MATCH PERFECTLY.")

    print("\n  Done.")


if __name__ == '__main__':
    main()
