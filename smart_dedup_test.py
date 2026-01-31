#!/usr/bin/env python3
"""
Smart page-boundary deduplication + positional matching.

Logic:
- 70 screenshots, ~26 rows each, ~3-4 overlapping rows per boundary
- CSV has 2079 rows, Excel has 1825 → 254 duplicates to remove
- At page boundaries, the last few rows of page N reappear at the start of page N+1
- Detect duplicates by comparing each row to recent rows (sliding window)
- After dedup, positional matching gives correct row order

Dedup strategy:
- For each CSV row, compare its CUSIP to CUSIPs in a recent window (last 8 rows)
- If the CUSIP is very similar (OCR-aware, high threshold) to a recent row AND
  the issuer name is similar, it's a page boundary duplicate → skip
"""
import csv
import re
import openpyxl
from collections import defaultdict

OCR_PAIRS = set()
_pairs = [
    ('0','X'),('0','O'),('0','D'),('0','Q'),
    ('1','I'),('1','L'),('1','7'),('1','J'),
    ('3','2'),('3','8'),('3','5'),
    ('4','8'),('4','A'),('4','9'),
    ('5','S'),('5','F'),('5','6'),('5','3'),
    ('6','G'),('6','8'),('6','A'),('6','5'),
    ('7','T'),('7','1'),
    ('8','B'),('8','6'),('8','3'),
    ('9','G'),('9','Q'),('9','4'),
    ('A','4'),('A','6'),
    ('D','K'),('D','H'),('D','0'),
    ('E','F'),('E','C'),
    ('F','5'),('F','P'),
    ('G','6'),('G','9'),('G','C'),
    ('H','K'),('H','D'),('H','N'),
    ('J','1'),('J','U'),
    ('K','X'),('K','H'),('K','D'),
    ('L','1'),('L','I'),
    ('M','N'),('M','W'),('M','H'),
    ('N','M'),('N','H'),
    ('O','0'),('O','Q'),('O','D'),
    ('P','R'),('P','F'),
    ('Q','9'),('Q','0'),('Q','G'),
    ('R','P'),
    ('S','5'),
    ('T','7'),
    ('U','V'),('U','J'),
    ('V','U'),('V','W'),
    ('W','M'),('W','V'),('W','A'),
    ('X','K'),('X','0'),
    ('Y','V'),
    ('Z','2'),
]
for a, b in _pairs:
    OCR_PAIRS.add((a, b))
    OCR_PAIRS.add((b, a))


def cusip_similarity(c1, c2):
    """How similar are two CUSIPs? Returns ratio 0-1."""
    # Compare up to min length
    ml = min(len(c1), len(c2), 9)
    if ml == 0:
        return 0
    exact = 0
    ocr = 0
    for i in range(ml):
        a, b = c1[i], c2[i]
        if a == b:
            exact += 1
        elif (a, b) in OCR_PAIRS:
            ocr += 1
    return (exact + ocr * 0.5) / ml


def issuer_similarity(i1, i2):
    """How similar are two issuer names? Returns ratio 0-1."""
    a = i1.lower().strip().rstrip('.')[:20]
    b = i2.lower().strip().rstrip('.')[:20]
    if not a or not b:
        return 0
    # Prefix match
    prefix = 0
    for x, y in zip(a, b):
        if x == y:
            prefix += 1
        else:
            break
    return prefix / max(len(a), len(b))


def fields_similarity(f1, f2):
    """Compare all fields of two CSV rows to detect duplicates."""
    # Compare: CUSIP, State, Issuer, Yield, Amt, and a few more
    cusip_sim = cusip_similarity(f1[0].strip(), f2[0].strip())
    issuer_sim = issuer_similarity(f1[2] if len(f1) > 2 else '', f2[2] if len(f2) > 2 else '')

    # State match
    state_match = 1.0 if (len(f1) > 1 and len(f2) > 1 and f1[1].strip() == f2[1].strip()) else 0.0

    # Yield similarity
    yield_sim = 0.0
    if len(f1) > 3 and len(f2) > 3:
        y1 = f1[3].strip()
        y2 = f2[3].strip()
        if y1 == y2:
            yield_sim = 1.0
        elif y1 and y2:
            # Compare as numbers
            try:
                yv1, yv2 = float(y1), float(y2)
                if abs(yv1 - yv2) < 0.1:
                    yield_sim = 0.8
            except ValueError:
                pass

    # Combined score
    return cusip_sim * 0.4 + issuer_sim * 0.3 + state_match * 0.1 + yield_sim * 0.2


# Load CSV
csv_rows = []
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    for row_data in reader:
        if len(row_data) >= 10:
            cusip = row_data[0].strip()
            if cusip.startswith('TH '):
                cusip = cusip[3:]
            if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                if not all(f.strip() == '--' for f in row_data[:5]):
                    csv_rows.append(row_data)

print(f"CSV rows before dedup: {len(csv_rows)}")

# Smart dedup: sliding window comparison
WINDOW_SIZE = 8  # Look back up to 8 rows
DUPE_THRESHOLD = 0.65  # Similarity threshold to consider a duplicate

deduped = []
removed = 0

for i, row in enumerate(csv_rows):
    is_dupe = False

    # Compare with recent rows
    start = max(0, len(deduped) - WINDOW_SIZE)
    for j in range(start, len(deduped)):
        sim = fields_similarity(row, deduped[j])
        if sim >= DUPE_THRESHOLD:
            is_dupe = True
            break

    if not is_dupe:
        deduped.append(row)
    else:
        removed += 1

print(f"Removed {removed} duplicates")
print(f"CSV rows after dedup: {len(deduped)}")
print(f"Target: 1825")
print(f"Difference: {len(deduped) - 1825}")

# Load Excel
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    amt = ws_orig.cell(row=row, column=5).value
    formula = ws_orig.cell(row=row, column=2).value or ''
    m = re.search(r'"([^"]+)\s+Muni"', str(formula))
    bb_id = m.group(1) if m else ''
    excel_rows.append({'issuer': issuer, 'amt': amt, 'bb_id': bb_id})

print(f"Excel rows: {len(excel_rows)}")

# Test positional matching
print(f"\nPositional matching after smart dedup:")
match_count = 0
mismatch_count = 0
mismatches = []

for i in range(min(len(excel_rows), len(deduped))):
    e_issuer = excel_rows[i]['issuer'].lower()[:12]
    c_issuer = deduped[i][2].strip().lower().rstrip('.')[:12] if len(deduped[i]) > 2 else ''

    if e_issuer == c_issuer:
        match_count += 1
    else:
        mismatch_count += 1
        if len(mismatches) < 30:
            mismatches.append((i, excel_rows[i]['issuer'][:40], deduped[i][2].strip()[:30] if len(deduped[i]) > 2 else 'N/A'))

print(f"  Issuer matches: {match_count}/{min(len(excel_rows), len(deduped))}")
print(f"  Mismatches: {mismatch_count}")

if mismatches:
    print(f"\n  First mismatches:")
    for pos, e_iss, c_iss in mismatches[:20]:
        print(f"    [{pos:4d}] Excel: '{e_iss}' vs CSV: '{c_iss}'")

# If deduped count != 1825, try adjusting threshold
if len(deduped) != 1825:
    print(f"\n\nTrying different thresholds:")
    for thresh in [0.55, 0.60, 0.65, 0.70, 0.75, 0.80]:
        deduped_test = []
        for i, row in enumerate(csv_rows):
            is_dupe = False
            start = max(0, len(deduped_test) - WINDOW_SIZE)
            for j in range(start, len(deduped_test)):
                sim = fields_similarity(row, deduped_test[j])
                if sim >= thresh:
                    is_dupe = True
                    break
            if not is_dupe:
                deduped_test.append(row)
        print(f"  Threshold {thresh:.2f}: {len(deduped_test)} rows (target 1825, diff {len(deduped_test) - 1825})")
