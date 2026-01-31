#!/usr/bin/env python3
"""Debug remaining unmatched rows - check why they can't match."""
import csv
import re
import openpyxl
from collections import defaultdict

OCR_PAIRS = set()
_pairs = [
    ('0','X'),('0','O'),('0','D'),
    ('1','I'),('1','L'),('1','7'),
    ('3','2'),('3','8'),
    ('4','8'),('4','A'),
    ('5','S'),('5','F'),('5','6'),
    ('6','G'),('6','8'),
    ('7','T'),
    ('8','B'),
    ('9','G'),('9','Q'),
    ('D','K'),('D','H'),
    ('E','F'),
    ('K','H'),('K','X'),
    ('M','N'),('M','W'),
    ('P','R'),
    ('U','V'),('W','V'),
]
for a, b in _pairs:
    OCR_PAIRS.add((a, b))
    OCR_PAIRS.add((b, a))

def cusip_check_digit(base8):
    values = []
    for ch in base8.upper():
        if ch.isdigit(): values.append(int(ch))
        elif ch.isalpha(): values.append(ord(ch) - ord('A') + 10)
        else: values.append(0)
    total = 0
    for i, v in enumerate(values):
        if i % 2 == 1: v *= 2
        total += v // 10 + v % 10
    return str((10 - (total % 10)) % 10)

def cusip_ocr_score(correct9, csv_cusip):
    best = 0
    def sp(a, b):
        s = 0
        for x, y in zip(a, b):
            if x == y: s += 3
            elif x.upper() == y.upper(): s += 2
            elif (x, y) in OCR_PAIRS: s += 1
        return s
    clen = len(csv_cusip)
    if clen == 9: best = sp(correct9, csv_cusip)
    elif clen == 8: best = sp(correct9[:8], csv_cusip)
    elif clen == 10:
        for skip in range(10):
            trimmed = csv_cusip[:skip] + csv_cusip[skip+1:]
            best = max(best, sp(correct9, trimmed))
    else:
        ml = min(9, clen)
        best = sp(correct9[:ml], csv_cusip[:ml])
    return best

# Load data
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    amt = ws_orig.cell(row=row, column=5).value
    formula = ws_orig.cell(row=row, column=2).value or ''
    m = re.search(r'"([^"]+)\s+Muni"', str(formula))
    bb_id = m.group(1) if m else ''
    correct_cusip = ''
    if bb_id and len(bb_id) >= 8:
        correct_cusip = bb_id[:8] + cusip_check_digit(bb_id[:8])
    excel_rows.append({'issuer': issuer, 'amt': amt, 'bb_id': bb_id,
                       'correct_cusip': correct_cusip})

csv_rows = []
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    for row_data in reader:
        if len(row_data) >= 10:
            cusip = row_data[0].strip()
            if cusip.startswith('TH '): cusip = cusip[3:]
            if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                if not all(f.strip() == '--' for f in row_data[:5]):
                    csv_rows.append({'cusip': cusip, 'issuer': row_data[2].strip(), 'fields': row_data})

# For each of the top unmatched issuers, find their correct CUSIPs and best CSV matches
test_issuers = [
    'Massachusetts Clean Water Trust',
    'Oklahoma Water Resources Board',
    'Metropolitan Transportation Authority',
    'Ohio Water Development Authority',
]

for test_issuer in test_issuers:
    print(f"\n{'='*60}")
    print(f"Issuer: {test_issuer}")

    # Find Excel rows for this issuer
    ei_list = [i for i, e in enumerate(excel_rows)
               if e['issuer'].lower().startswith(test_issuer.lower()[:15])]
    print(f"  Excel rows: {len(ei_list)}")

    if ei_list:
        # Show correct CUSIPs
        for ei in ei_list[:5]:
            e = excel_rows[ei]
            print(f"    Correct CUSIP: {e['correct_cusip']} (BB_ID: {e['bb_id']})")

            # Find best CSV matches for this CUSIP
            scores = []
            for ci, c in enumerate(csv_rows):
                sc = cusip_ocr_score(e['correct_cusip'], c['cusip'])
                if sc >= 8:  # Very low threshold
                    scores.append((sc, ci, c['cusip'], c['issuer'][:30]))
            scores.sort(reverse=True)
            if scores:
                print(f"    Top CSV matches:")
                for sc, ci, csv_c, csv_iss in scores[:5]:
                    print(f"      Score {sc}: CUSIP={csv_c} Issuer={csv_iss}")
            else:
                print(f"    NO CSV matches with score >= 8!")

    # Also search CSV by issuer
    matching_csv = [ci for ci, c in enumerate(csv_rows)
                    if c['issuer'].lower().startswith(test_issuer.lower()[:8])]
    print(f"  CSV rows with matching issuer prefix: {len(matching_csv)}")
    for ci in matching_csv[:5]:
        c = csv_rows[ci]
        print(f"    CSV CUSIP: {c['cusip']} Issuer: {c['issuer'][:40]}")
