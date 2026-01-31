#!/usr/bin/env python3
"""Test matching by issuer name prefix + amount."""
import csv
import re
import openpyxl
from collections import defaultdict

# Load original Excel
wb_orig = openpyxl.load_workbook('/home/user/bond-screenshots/green bonds excel.xlsx')
ws_orig = wb_orig.active

excel_rows = []
for row in range(2, ws_orig.max_row + 1):
    issuer = (ws_orig.cell(row=row, column=3).value or '').strip()
    amt = ws_orig.cell(row=row, column=5).value
    formula = ws_orig.cell(row=row, column=2).value or ''
    m = re.search(r'"([^"]+)\s+Muni"', str(formula))
    bb_id = m.group(1) if m else ''
    excel_rows.append({'issuer': issuer, 'amt': amt, 'bb_id': bb_id, 'row': row})

# Load CSV
csv_rows = []
with open('/home/user/bond-screenshots/claude_table_output_2025_new.csv', 'r') as f:
    reader = csv.reader(f)
    for row_data in reader:
        if len(row_data) >= 10:
            cusip = row_data[0].strip()
            if cusip.startswith('TH '):
                cusip = cusip[3:]
            if len(cusip) >= 4 and re.match(r'^[A-Za-z0-9/]+$', cusip):
                if not all(f.strip() == '--' for f in row_data[:5]):
                    csv_rows.append({
                        'cusip': cusip,
                        'issuer': row_data[2].strip(),
                        'amt_raw': row_data[4].strip() if len(row_data) > 4 else '',
                        'yield_raw': row_data[3].strip() if len(row_data) > 3 else '',
                        'fields': row_data
                    })

print(f"Excel rows: {len(excel_rows)}")
print(f"CSV rows: {len(csv_rows)}")

# Parse CSV amounts
def parse_csv_amt(raw):
    """Parse CSV amount to numeric value."""
    s = raw.strip().replace(',', '').replace(' ', '')
    if not s or s == '--':
        return None
    # Remove trailing dots
    s = s.rstrip('.')
    try:
        val = float(s)
        # CSV amounts may be in millions (e.g., 2.150 = 2,150,000)
        # or in thousands (e.g., 2150 = 2,150,000)
        # or raw (e.g., 2145000)
        return val
    except ValueError:
        return None

# Check amount format
print("\nCSV amount samples (first 20):")
for i in range(min(20, len(csv_rows))):
    raw = csv_rows[i]['amt_raw']
    parsed = parse_csv_amt(raw)
    excel_amt = excel_rows[i]['amt'] if i < len(excel_rows) else 'N/A'
    print(f"  CSV: '{raw}' -> {parsed}, Excel: {excel_amt}")

# Figure out amount scale
print("\nAmount scale analysis:")
for i in range(min(20, len(csv_rows))):
    parsed = parse_csv_amt(csv_rows[i]['amt_raw'])
    excel_amt = excel_rows[i]['amt']
    if parsed and excel_amt and parsed > 0:
        ratio = excel_amt / parsed
        print(f"  Excel {excel_amt} / CSV {parsed} = {ratio:.1f}")
