#!/usr/bin/env python3
"""Fix the 3 remaining invalid Project Subcategory values."""
import openpyxl

wb = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws = wb.active

fixes = {
    324: ('Greenhouse Gas Control, ?Energy', 'Green House Gas Control, Energy Storage'),
    529: ('Greenhouse Gas Control, ?Solar...', 'Green House Gas Control, Solar'),
    1488: ('?Sub.and, ?or Energy Storage', 'Energy Storage'),
}

for row, (expected, new_val) in fixes.items():
    val = ws.cell(row=row, column=20).value
    print(f"Row {row}: '{val}' -> '{new_val}'")
    ws.cell(row=row, column=20).value = new_val

wb.save('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
print("\nFixed 3 remaining values.")

# Final verification
from collections import Counter
wb2 = openpyxl.load_workbook('/home/user/bond-screenshots/green_bonds_2025_final.xlsx')
ws2 = wb2.active

VALID_BASE = {
    'Bioenergy', 'BREEAM Certified', 'Circular Design and Production',
    'Circular Value Recovery', 'Conservation', 'Energy Star Certified',
    'Energy Storage', 'Geothermal', 'Green House Gas Control',
    'Greenhouse Gas Control', 'Hydro', 'Hydrogen', 'Infrastructure',
    'Information Support', 'LEED Certified', 'Marine', 'Multimodal',
    'Non Motorized', 'Plumbing System', 'Pollution Control', 'Public',
    'Rail (Non Passenger)', 'Smart Grids', 'Soil Remediation', 'Solar',
    'Sustainable Forestry', 'Vehicles', 'Waste Management',
    'WELL Certified', 'Wind',
}

counter = Counter()
non_null = 0
invalid = []
for row in range(2, ws2.max_row + 1):
    val = ws2.cell(row=row, column=20).value
    if val is not None and str(val).strip():
        s = str(val).strip()
        counter[s] += 1
        non_null += 1
        # Validate each part
        parts = [p.strip() for p in s.split(',')]
        for p in parts:
            if p not in VALID_BASE:
                invalid.append((row, s, p))

print(f"\nFinal: {non_null} non-null values, {len(counter)} unique categories")
if invalid:
    print(f"\nINVALID parts found ({len(invalid)}):")
    for row, val, part in invalid:
        print(f"  Row {row}: '{val}' -> invalid part '{part}'")
else:
    print("ALL values are valid!")

print(f"\nDistribution:")
for val, cnt in counter.most_common():
    print(f"  [{cnt:3d}] {val}")
